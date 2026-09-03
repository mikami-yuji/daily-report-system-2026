import os
import shutil
import tempfile
import time
import logging
import threading
from collections import defaultdict
from contextlib import contextmanager
from typing import Optional
import openpyxl
from fastapi import HTTPException

import cache

# ファイルパスごとの再帰的排他ロック（スレッドセーフ）
_file_locks = defaultdict(threading.RLock)
_global_lock = threading.Lock()

@contextmanager
def get_file_write_lock(file_path: str, timeout: float = 30.0):
    """
    同一Excelファイルへの並行書き込みを直列化（キューイング）するコンテキストマネージャー。
    指定タイムアウト内にロックが取得できない場合は 409 Conflict を送出します。
    """
    norm_path = os.path.abspath(file_path).lower()
    with _global_lock:
        lock = _file_locks[norm_path]
    
    acquired = lock.acquire(timeout=timeout)
    if not acquired:
        logging.warning(f"File write lock timeout ({timeout}s) on {file_path}")
        raise HTTPException(
            status_code=409,
            detail="現在、別のユーザーまたは処理がこのExcelファイルを更新中です。数秒待ってから再試行してください。"
        )
    try:
        logging.debug(f"Acquired write lock for {file_path}")
        yield
    finally:
        lock.release()
        logging.debug(f"Released write lock for {file_path}")

def safe_save_workbook_with_retry(
    wb: openpyxl.Workbook,
    target_file_path: str,
    max_retries: int = 4,
    base_delay: float = 0.3,
    create_backup_task: bool = True
) -> None:
    """
    一時ファイルに保存・整合性検証を行った後、対象ファイルを安全に置き換えます。
    同一ボリュームであれば os.replace によるアトミック置換を優先し、
    ファイルロック(PermissionError)や一時的I/Oエラーが発生した場合は指数バックオフでリトライします。
    """
    filename = os.path.basename(target_file_path)
    temp_dir = tempfile.gettempdir()
    temp_file = os.path.join(temp_dir, f"temp_{int(time.time() * 1000)}_{filename}")

    # 1. 一時ファイルへ保存
    try:
        wb.save(temp_file)
    except Exception as e:
        logging.error(f"Failed to save temporary workbook {temp_file}: {e}")
        if os.path.exists(temp_file):
            try:
                os.remove(temp_file)
            except Exception:
                pass
        raise HTTPException(status_code=500, detail=f"一時ファイルへの保存に失敗しました: {str(e)}")

    # 2. 一時ファイルが破損していないか読み込みテスト
    try:
        test_wb = openpyxl.load_workbook(temp_file, read_only=True)
        test_wb.close()
    except Exception as e:
        logging.error(f"Temporary file validation failed for {temp_file}: {e}")
        if os.path.exists(temp_file):
            try:
                os.remove(temp_file)
            except Exception:
                pass
        raise HTTPException(status_code=500, detail="保存データの検証に失敗しました。ファイルが破損している可能性があります。")

    # 3. リトライ付きで元ファイルへ置換
    last_error: Optional[Exception] = None
    for attempt in range(max_retries):
        try:
            try:
                # 同一ボリューム等の場合は os.replace によるアトミック置換を試行
                os.replace(temp_file, target_file_path)
                logging.info(f"Successfully atomic-replaced {target_file_path} (attempt {attempt + 1})")
                last_error = None
                break
            except OSError:
                # 異なるドライブ等の場合は copy2 でフォールバック
                shutil.copy2(temp_file, target_file_path)
                logging.info(f"Successfully copied and replaced {target_file_path} (attempt {attempt + 1})")
                last_error = None
                break
        except (PermissionError, OSError) as e:
            last_error = e
            wait_time = base_delay * (2 ** attempt)
            logging.warning(
                f"File locked or write error on {target_file_path} (attempt {attempt + 1}/{max_retries}). "
                f"Retrying in {wait_time:.2f}s... Error: {e}"
            )
            time.sleep(wait_time)
        except Exception as e:
            last_error = e
            logging.error(f"Unexpected error replacing {target_file_path}: {e}")
            break

    # 一時ファイルのクリーンアップ
    if os.path.exists(temp_file):
        try:
            os.remove(temp_file)
        except Exception:
            pass

    if last_error is not None:
        if isinstance(last_error, PermissionError):
            raise HTTPException(
                status_code=409,
                detail="Excelファイルが別のプロセスで開かれているため保存できませんでした。ファイルを閉じてから再試行してください。"
            )
        raise HTTPException(status_code=500, detail=f"ファイル保存に失敗しました: {str(last_error)}")

    # 4. バックアップの作成
    if create_backup_task:
        try:
            cache.create_backup(target_file_path)
        except Exception as e:
            logging.warning(f"Backup creation warning for {target_file_path}: {e}")


def safe_load_workbook_with_retry(
    file_path: str,
    keep_vba: bool = True,
    read_only: bool = False,
    data_only: bool = False,
    max_retries: int = 3,
    base_delay: float = 0.2
) -> openpyxl.Workbook:
    """
    ファイルロックやネットワーク遅延を考慮したリトライ付きopenpyxlロード関数。
    """
    if not os.path.exists(file_path):
        raise HTTPException(status_code=404, detail=f"ファイルが見つかりません: {file_path}")

    last_error: Optional[Exception] = None
    for attempt in range(max_retries):
        try:
            wb = openpyxl.load_workbook(
                file_path,
                keep_vba=keep_vba,
                read_only=read_only,
                data_only=data_only
            )
            return wb
        except (PermissionError, OSError) as e:
            last_error = e
            wait_time = base_delay * (2 ** attempt)
            logging.warning(
                f"File access error reading {file_path} (attempt {attempt + 1}/{max_retries}). "
                f"Retrying in {wait_time:.2f}s... Error: {e}"
            )
            time.sleep(wait_time)
        except Exception as e:
            logging.error(f"Failed to load workbook {file_path}: {e}")
            raise HTTPException(status_code=500, detail=f"Excelファイルの読み込みに失敗しました: {str(e)}")

    if isinstance(last_error, PermissionError):
        raise HTTPException(
            status_code=409,
            detail="Excelファイルがロックされているため読み込めませんでした。しばらく待ってから再試行してください。"
        )
    raise HTTPException(status_code=500, detail=f"Excelファイルの読み込みに失敗しました: {str(last_error)}")
