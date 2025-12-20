import openpyxl
import os
import sys

EXCEL_DIR = r'\\Asahipack02\社内書類ｎｅｗ\01：部署別　営業部\02：営業日報\2025年度'

print(f"Scanning directory: {EXCEL_DIR}")

try:
    files = [f for f in os.listdir(EXCEL_DIR) if f.endswith('.xlsm') and not f.startswith('~$')]
    if not files:
        print("No .xlsm files found.")
        sys.exit(1)
        
    target_file = os.path.join(EXCEL_DIR, files[0])
    print(f"Loading workbook from: {target_file}")

    wb = openpyxl.load_workbook(target_file, keep_vba=True, read_only=False, data_only=True)
    if 'デザイン種別' in wb.defined_names:
        dn = wb.defined_names['デザイン種別']
        print("Found named range 'デザイン種別'")
        for title, coord in dn.destinations:
            print(f"Sheet: {title}, Coord: {coord}")
            ws = wb[title]
            # Handle range vs single cell
            coord = coord.replace('$', '')
            if ':' in coord:
                cells = ws[coord]
                values = []
                for row in cells:
                    for cell in row:
                        if cell.value:
                            values.append(cell.value)
                print(f"VALUES: {values}")
            else:
                print(f"VALUE: {ws[coord].value}")
    else:
        print("Named range 'デザイン種別' NOT FOUND.")
    
    wb.close()
    
except Exception as e:
    print(f"ERROR: {e}")

