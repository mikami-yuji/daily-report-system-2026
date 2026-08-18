import pytest
import os
import sys
import tempfile
import openpyxl

sys.path.insert(0, os.path.abspath(os.path.join(os.path.dirname(__file__), '..')))

import excel_io

def test_safe_save_and_load_workbook():
    """ワークブックの安全な保存と再読み込みのテスト"""
    temp_dir = tempfile.mkdtemp()
    test_file = os.path.join(temp_dir, "test_report.xlsx")
    
    # ワークブック作成
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "営業日報"
    ws.cell(row=1, column=1, value="管理番号")
    ws.cell(row=2, column=1, value=100)
    
    # 初回保存
    excel_io.safe_save_workbook_with_retry(wb, test_file, create_backup_task=False)
    assert os.path.exists(test_file)
    
    # 読み込みテスト
    loaded_wb = excel_io.safe_load_workbook_with_retry(test_file, keep_vba=False)
    assert "営業日報" in loaded_wb.sheetnames
    loaded_ws = loaded_wb["営業日報"]
    assert loaded_ws.cell(row=2, column=1).value == 100
    
    # 更新テスト
    loaded_ws.cell(row=2, column=2, value="2026-08-18")
    excel_io.safe_save_workbook_with_retry(loaded_wb, test_file, create_backup_task=False)
    
    # 再度読み込み確認
    reloaded_wb = excel_io.safe_load_workbook_with_retry(test_file, keep_vba=False)
    assert reloaded_wb["営業日報"].cell(row=2, column=2).value == "2026-08-18"
    reloaded_wb.close()
    
    # クリーンアップ
    try:
        os.remove(test_file)
        os.rmdir(temp_dir)
    except Exception:
        pass
