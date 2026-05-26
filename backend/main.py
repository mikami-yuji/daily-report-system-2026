from fastapi import FastAPI, HTTPException, UploadFile, File, BackgroundTasks, Request
from fastapi.staticfiles import StaticFiles
from fastapi.responses import FileResponse, JSONResponse
from fastapi.middleware.cors import CORSMiddleware
from fastapi.exceptions import RequestValidationError
from pydantic import BaseModel, Field, field_validator, model_validator
import pandas as pd
import openpyxl
from datetime import datetime, timedelta
import os
import shutil
import json
import pickle
import hashlib
from typing import Optional, List, Dict, Any


import logging
import sys
import os

def get_base_path():
    """EXEの実行ディレクトリ（config.json、ログ等の外部ファイル用）"""
    if getattr(sys, 'frozen', False):
        return os.path.dirname(sys.executable)
    return os.path.dirname(os.path.abspath(__file__))

def get_bundle_path():
    """バンドルデータのディレクトリ（PyInstaller展開先のstatic等用）"""
    if getattr(sys, '_MEIPASS', None):
        return sys._MEIPASS
    return os.path.dirname(os.path.abspath(__file__))

BASE_DIR = get_base_path()
BUNDLE_DIR = get_bundle_path()

# Ensure multipart libraries are bundled by PyInstaller
try:
    import python_multipart
    import multipart
except ImportError:
    pass

# Setup logging - ファイルとコンソール両方に出力
logging.basicConfig(
    level=logging.DEBUG,
    format='%(asctime)s - %(levelname)s - %(message)s',
    handlers=[
        logging.FileHandler('server_debug.log'),
        logging.StreamHandler()  # コンソール出力
    ]
)
logging.info("Server starting up...")
logging.info(f"DEBUG PATHS: BASE_DIR={BASE_DIR}")
logging.info(f"DEBUG PATHS: BUNDLE_DIR={BUNDLE_DIR}")
STATIC_DIR = os.path.join(BUNDLE_DIR, "static")
logging.info(f"DEBUG PATHS: STATIC_DIR={STATIC_DIR}")
logging.info(f"DEBUG PATHS: Static Exists={os.path.exists(STATIC_DIR)}")

logging.info(f"Loaded python_multipart: {python_multipart.__file__ if 'python_multipart' in locals() else 'Not found'}")

app = FastAPI()

# Enable CORS for frontend communication
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],  # In production, specify the frontend URL
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

# バリデーションエラーの詳細をログ出力
@app.exception_handler(RequestValidationError)
async def validation_exception_handler(request: Request, exc: RequestValidationError):
    logging.error(f"Validation error on {request.url.path}: {exc.errors()}")
    logging.error(f"Request body: {exc.body}")
    return JSONResponse(
        status_code=422,
        content={"detail": exc.errors(), "body": str(exc.body)[:500]}
    )

from fastapi import Request

# ミドルウェアは削除済み - APIエンドポイントは直接 /api/ プレフィックス付きで定義されている


# Load configuration
def load_config() -> str:
    config_path = os.path.join(BASE_DIR, 'config.json')
    # 2026年度版のデフォルトパス
    default_path = r'\\Asahipack02\社内書類ｎｅｗ\01：部署別　営業部\02：営業日報\2026年度'
    
    path = None
    
    if os.path.exists(config_path):
        try:
            with open(config_path, 'r', encoding='utf-8-sig') as f:
                config = json.load(f)
                path = config.get('excel_dir')
                if path:
                    logging.info(f"Loaded raw config path: {path}")
        except Exception as e:
            logging.warning(f"Failed to load config.json: {e}")

    # もしpathが設定されている場合は、パスの正規化と存在チェックを行う
    if path:
        # 相対パスの場合はBASE_DIR基準で絶対パス化する
        if not os.path.isabs(path):
            path = os.path.abspath(os.path.join(BASE_DIR, path))
            logging.info(f"Resolved relative path to absolute: {path}")
            
        # パスが存在するかチェック
        if os.path.exists(path):
            logging.info(f"Using configured Excel path: {path}")
            return path
        else:
            logging.warning(f"Configured Excel path does not exist: {path}")

    # パスが存在しないか指定されていない場合、順次フォールバック探索を行う
    # フォールバック1: デフォルトの共有フォルダ（社内LAN接続時を最優先とする）
    if os.path.exists(default_path):
        logging.info(f"Fallback 1: Using network shared directory: {default_path}")
        return default_path

    # フォールバック2: 実行ディレクトリ配下の 'data'（オフライン・ローカル動作時）
    local_data_path = os.path.abspath(os.path.join(BASE_DIR, 'data'))
    if os.path.exists(local_data_path):
        logging.info(f"Fallback 2: Using local data directory: {local_data_path}")
        return local_data_path

    # フォールバック3: パッケージ配下の 'DailyReportSystem_2026_Simple/data' (あれば)
    simple_data_path = os.path.abspath(os.path.join(BASE_DIR, 'DailyReportSystem_2026_Simple', 'data'))
    if os.path.exists(simple_data_path):
        logging.info(f"Fallback 3: Using simple package data directory: {simple_data_path}")
        return simple_data_path

    # 最終的なフォールバック（作成可能なようにlocal_data_pathを返して自動で作成させる）
    logging.info(f"No existing paths found. Fallback to default local data path: {local_data_path}")
    return local_data_path

EXCEL_DIR = load_config()
logging.info(f"STARTUP: Working with EXCEL_DIR: {EXCEL_DIR}")

# --- Global Sales Data Storage ---
DATA_DIR = os.path.join(BASE_DIR, 'data')
SALES_CSV_PATH = os.path.join(DATA_DIR, 'sales_data.csv')
os.makedirs(DATA_DIR, exist_ok=True)

# Global DataFrame to hold sales data
global_sales_df = None

def load_sales_data():
    """Loads sales data from CSV into global DataFrame."""
    global global_sales_df
    if not os.path.exists(SALES_CSV_PATH):
        logging.info("No existing sales data found.")
        return

    try:
        logging.info("Loading sales data from disk...")
        # Try cp932 first
        try:
            df = pd.read_csv(SALES_CSV_PATH, encoding='cp932')
        except:
            df = pd.read_csv(SALES_CSV_PATH, encoding='utf-8')
        
        # Clean columns
        df.columns = [str(col).strip() for col in df.columns]
        
        # Ensure customer code exists
        if '得意先コード' in df.columns:
            # Normalize customer code (remove decimals, convert to string)
            df['得意先コード'] = df['得意先コード'].astype(str).str.split('.').str[0]
            global_sales_df = df
            logging.info(f"Sales data loaded successfully. {len(df)} rows.")
        else:
            logging.error("Sales CSV missing '得意先コード' column.")
    except Exception as e:
        logging.error(f"Failed to load sales data: {e}")

# Load on startup
load_sales_data()

# Find a default Excel file dynamically
DEFAULT_EXCEL_FILE = "daily_report_template.xlsm" # Fallback
if os.path.exists(EXCEL_DIR):
    files = [f for f in os.listdir(EXCEL_DIR) if f.endswith('.xlsm') and not f.startswith('~$')]
    if files:
        DEFAULT_EXCEL_FILE = files[0]
        logging.info(f"Set default Excel file to: {DEFAULT_EXCEL_FILE}")
    else:
        logging.warning("No .xlsm files found in directory. Using fallback default.")


class ReportInput(BaseModel):
    model_config = {"populate_by_name": True, "extra": "ignore"}  # 余分なフィールドを無視
    
    日付: str = ""
    行動内容: str = ""
    エリア: str = ""
    得意先CD: str = ""
    直送先CD: str = ""
    訪問先名: str = ""
    直送先名: str = ""
    重点顧客: str = ""
    ランク: str = ""
    面談者: str = ""
    滞在時間: str = ""
    商談内容: str = ""
    提案物: str = ""
    次回プラン: str = ""
    競合他社情報: str = ""
    デザイン提案有無: str = ""
    デザイン種別: str = ""
    デザイン名: str = ""
    デザイン進捗状況: str = ""
    デザイン依頼No: str = Field("", alias="デザイン依頼No.")
    上長コメント: str = ""
    コメント返信欄: str = ""
    上長: str = ""
    山澄常務: str = ""
    岡本常務: str = ""
    中野次長: str = ""
    既読チェック: str = ""
    original_values: Optional[Any] = None # For optimistic locking

    @model_validator(mode='before')
    @classmethod
    def convert_all_to_string(cls, data: Any) -> Any:
        # すべての文字列フィールドに対して、数値やnullを安全に文字列へ変換する
        if isinstance(data, dict):
            for key, value in data.items():
                if key == 'original_values':
                    continue
                if value is None:
                    data[key] = ""
                elif isinstance(value, (int, float)):
                    if isinstance(value, float) and value.is_integer():
                        data[key] = str(int(value))
                    else:
                        data[key] = str(value)
                elif not isinstance(value, str):
                    data[key] = str(value)
        return data

    @field_validator('得意先CD', '直送先CD', mode='before')
    @classmethod
    def convert_to_string(cls, v):
        if v is None:
            return ""
        return str(v)

@app.get("/api/health")
def read_root():
    return {"message": "Daily Report API is running", "excel_dir": EXCEL_DIR}

@app.get("/api/files")
def list_excel_files():
    """List all Excel files in the directory"""
    logging.debug(f"Listing files in {EXCEL_DIR}")
    if not os.path.exists(EXCEL_DIR):
         logging.error(f"Directory not found: {EXCEL_DIR}")
         raise HTTPException(status_code=500, detail=f"Excel Directory not found: {EXCEL_DIR}")
         
    try:
        files = []
        # Add timeout protection or more verbose logging? 
        # listing network drive can appear to hang.
        
        items = os.listdir(EXCEL_DIR)
        logging.debug(f"Found {len(items)} items in directory")
        
        for file in items:
            if file.endswith(('.xlsx', '.xlsm')):
                file_path = os.path.join(EXCEL_DIR, file)
                try:
                    file_size = os.path.getsize(file_path)
                    file_mtime = os.path.getmtime(file_path)
                    files.append({
                        "name": file,
                        "size": file_size,
                        "modified": datetime.fromtimestamp(file_mtime).isoformat()
                    })
                except Exception as file_err:
                    logging.warning(f"Error processing file {file}: {file_err}")
                    continue
                    
        return {"files": files, "default": DEFAULT_EXCEL_FILE}
    except Exception as e:
        logging.critical(f"CRITICAL ERROR in list_excel_files: {str(e)}")
        import traceback
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=f"Failed to list files: {str(e)}")


# Cache for Excel dataframes: {(filename, sheet_name): {'mtime': float, 'df': pd.DataFrame}}
CACHE = {}

def create_backup(file_path):
    try:
        backup_dir = os.path.join(os.path.dirname(file_path), 'backup')
        if not os.path.exists(backup_dir):
            os.makedirs(backup_dir)
        
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        filename = os.path.basename(file_path)
        name, ext = os.path.splitext(filename)
        backup_filename = f"{name}_{timestamp}{ext}"
        backup_path = os.path.join(backup_dir, backup_filename)
        
        shutil.copy2(file_path, backup_path)
        logging.info(f"Backup created: {backup_path}")
    except Exception as e:
        logging.warning(f"Failed to create backup: {e}")


def get_cached_dataframe(filename: str, sheet_name: str) -> pd.DataFrame:
    """
    Get dataframe from cache or read from Excel file if modified or not in cache.
    """
    excel_file = os.path.join(EXCEL_DIR, filename)
    
    if not os.path.exists(excel_file):
        logging.error(f"File not found: {excel_file}")
        raise HTTPException(status_code=404, detail=f"Excel file '{filename}' not found at {excel_file}")
    
    current_mtime = os.path.getmtime(excel_file)
    cache_key = (filename, sheet_name)
    
    # --- In-Memory Cache Check ---
    if cache_key in CACHE:
        cached_data = CACHE[cache_key]
        if cached_data['mtime'] == current_mtime:
            return cached_data['df'].copy() # Return copy to prevent mutation of cached data

    # --- Disk Cache Check ---
    # Create cache directory if needed
    CACHE_DIR = os.path.join(os.path.dirname(os.path.abspath(__file__)), ".cache")
    if not os.path.exists(CACHE_DIR):
        os.makedirs(CACHE_DIR)
        
    # Create unique cache filename based on file path and sheet
    cache_id = hashlib.md5(f"{filename}_{sheet_name}".encode('utf-8')).hexdigest()
    cache_path = os.path.join(CACHE_DIR, f"{cache_id}.pkl")
    
    try:
        if os.path.exists(cache_path):
            with open(cache_path, 'rb') as f:
                disk_cache = pickle.load(f)
            
            # Use disk cache if timestamp matches
            if disk_cache.get('mtime') == current_mtime:
                logging.debug(f"Loaded {filename} ({sheet_name}) from disk cache")
                df = disk_cache['df']
                # Update in-memory cache
                CACHE[cache_key] = {'mtime': current_mtime, 'df': df}
                return df.copy()
    except Exception as e:
        logging.warning(f"Failed to load from disk cache: {e}")

    # --- Read from Excel (Expensive Operation) ---
    try:
        logging.debug(f"Reading Excel {excel_file}, sheet={sheet_name}")
        df = pd.read_excel(excel_file, sheet_name=sheet_name, header=0)
        
        # Update in-memory cache
        CACHE[cache_key] = {'mtime': current_mtime, 'df': df}
        
        # Update disk cache
        try:
            with open(cache_path, 'wb') as f:
                pickle.dump({'mtime': current_mtime, 'df': df}, f)
            logging.debug(f"Saved {filename} ({sheet_name}) to disk cache")
        except Exception as e:
            logging.warning(f"Failed to save disk cache: {e}")
            
        return df.copy()
    except Exception as e:
        logging.error(f"Reading Excel failed: {e}")
        import traceback
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=f"Error reading Excel file: {str(e)}")


@app.get("/api/customers")
def get_customers(filename: str = DEFAULT_EXCEL_FILE):
    """Get customer list from the Excel file"""
    try:
        # Get dataframe from cache
        df = get_cached_dataframe(filename, '得意先_List')
        
        # Clean up column names
        df.columns = [str(col).replace('\n', '').strip() for col in df.columns]
        
        # デバッグログ: カラム名を出力
        logging.info(f"得意先_List columns: {list(df.columns)}")
        
        # 現目標カラムの存在確認
        target_col = None
        for col in df.columns:
            if '現目標' in col or '目標' in col:
                target_col = col
                logging.info(f"Found target column: '{col}'")
        
        # Rename specific columns
        # 都道府県（カラムF）→ エリア: フロントエンドが c.エリア でアクセスするため
        df = df.rename(columns={
            '得意先CD.': '得意先CD',
            '直送先CD.': '直送先CD',
            '都道府県': 'エリア',
        })
        
        # Fill NaN values with empty strings
        df = df.fillna(value='')
        
        # Convert to dict
        records = df.to_dict(orient="records")
        
        # デバッグ: 最初のレコードの現目標値を確認
        if records and len(records) > 0:
            sample = records[0]
            logging.info(f"Sample record keys: {list(sample.keys())}")
            if '現目標' in sample:
                logging.info(f"Sample 現目標 value: '{sample.get('現目標')}'")
            else:
                logging.warning("現目標 column not found in record!")
        
        # Clean the records
        import math
        cleaned_records = []
        for record in records:
            cleaned_record = {}
            for key, value in record.items():
                if isinstance(value, float):
                    if math.isnan(value) or math.isinf(value):
                        cleaned_record[key] = None
                    else:
                        cleaned_record[key] = value
                elif value == '':
                    cleaned_record[key] = None
                elif isinstance(value, str):
                    import re
                    cleaned_value = re.sub(r'_x000D_', '\n', value)
                    cleaned_value = cleaned_value.replace('\r', '')
                    cleaned_record[key] = cleaned_value
                else:
                    cleaned_record[key] = value
            cleaned_records.append(cleaned_record)

        return cleaned_records
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.get("/api/priority-customers")
def get_priority_customers(filename: str = DEFAULT_EXCEL_FILE):
    """得意先_Listからカラム H (重点顧客) が「重点」の顧客を取得。カラム I の担当者情報も含める"""
    try:
        # 得意先_Listを読み込み
        df = get_cached_dataframe(filename, '得意先_List')
        
        # カラム名をクリーンアップ
        df.columns = [str(col).replace('\n', '').strip() for col in df.columns]
        
        # カラム名を保存
        col_customer_cd = df.columns[0]  # 得意先CD
        col_customer_name = df.columns[1] if len(df.columns) > 1 else None  # 得意先名
        col_priority = df.columns[7] if len(df.columns) > 7 else None  # カラムH: 重点顧客
        col_staff = df.columns[8] if len(df.columns) > 8 else None  # カラムI: 担当者
        
        # 得意先CDがある行のみ抽出（ヘッダー行や空行を除外）
        df = df.dropna(subset=[col_customer_cd])
        
        # カラム H の名前を特定
        priority_col = col_priority
        if not priority_col:
            # フォールバック: 「重点顧客」という名前のカラムを探す
            for col in df.columns:
                if '重点' in str(col):
                    priority_col = col
                    break
            if not priority_col:
                logging.warning(f"Priority column not found. Columns: {list(df.columns)}")
                return []
        
        logging.info(f"Priority column: {priority_col}, Staff column: {col_staff}")
        
        # 「重点」と記載されている行のみ抽出
        priority_df = df[df[priority_col].astype(str).str.contains('重点', na=False)]
        
        logging.info(f"Found {len(priority_df)} priority customers")
        
        # レコードを作成
        records = []
        for _, row in priority_df.iterrows():
            customer_cd = row[col_customer_cd]
            customer_name = row[col_customer_name] if col_customer_name else ''
            staff = row[col_staff] if col_staff else ''
            
            # CDをクリーンアップ
            if isinstance(customer_cd, float):
                import math
                if math.isnan(customer_cd):
                    continue
                customer_cd = str(int(customer_cd))
            else:
                customer_cd = str(customer_cd).strip()
            
            if not customer_cd:
                continue
            
            # 担当者をクリーンアップ
            if isinstance(staff, float):
                import math
                if math.isnan(staff):
                    staff = ''
                else:
                    staff = str(staff)
            else:
                staff = str(staff).strip() if staff else ''
                
            records.append({
                '得意先CD': customer_cd,
                '得意先名': str(customer_name).strip() if customer_name else '',
                '担当者': staff
            })
        
        return records
    except Exception as e:
        logging.error(f"Error in get_priority_customers: {e}")
        import traceback
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=str(e))



@app.get("/api/interviewers")
def get_interviewers(customer_code: str, filename: str = DEFAULT_EXCEL_FILE):
    """Get list of interviewers for a specific customer"""
    excel_file = os.path.join(EXCEL_DIR, filename)
    if not os.path.exists(excel_file):
        raise HTTPException(status_code=404, detail=f"Excel file '{filename}' not found")
    
    try:
        # Read the '営業日報' sheet
        df = pd.read_excel(excel_file, sheet_name='営業日報', header=0)
        
        # Clean up column names
        df.columns = [str(col).replace('\n', '').strip() for col in df.columns]
        
        # Rename specific columns to match frontend expectations
        df = df.rename(columns={
            '得意先CD.': '得意先CD',
        })
        
        # Filter by customer code
        customer_reports = df[df['得意先CD'] == customer_code]
        
        # Get unique interviewers, excluding NaN and '-'
        interviewers = customer_reports['面談者'].dropna().unique().tolist()
        interviewers = [i for i in interviewers if i and str(i).strip() not in ['-', 'nan', '']]
        
        return interviewers
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@app.get("/api/reports/{management_number}")
def get_report_by_id(management_number: int, filename: str = DEFAULT_EXCEL_FILE):
    """指定された管理番号の日報を取得"""
    try:
        # Get dataframe from cache
        df = get_cached_dataframe(filename, '営業日報')
        
        # Clean up column names
        df.columns = [str(col).replace('\n', '') for col in df.columns]
        
        # Rename specific columns
        df = df.rename(columns={
            '得意先CD.': '得意先CD',
            '訪問先名得意先名': '訪問先名',
            'コメント': '上長コメント',  # Excel uses 'コメント' for manager comment
            'コメント返信欄': 'コメント返信欄'  # Keep as-is for reply field
        })
        
        # Filter by management number
        report_df = df[df['管理番号'] == management_number]
        
        if report_df.empty:
            raise HTTPException(status_code=404, detail=f"Report with management number {management_number} not found")
        
        # Get the first (and should be only) record
        record = report_df.iloc[0].to_dict()
        
        # Clean the record
        import math
        cleaned_record = {}
        for key, value in record.items():
            if isinstance(value, float):
                if math.isnan(value) or math.isinf(value):
                    cleaned_record[key] = None
                elif key == '得意先CD' and not math.isnan(value):
                    cleaned_record[key] = str(int(value))
                else:
                    cleaned_record[key] = value
            elif value == '':
                cleaned_record[key] = None
            elif isinstance(value, str):
                import re
                cleaned_value = re.sub(r'_x000D_', '\n', value)
                cleaned_value = cleaned_value.replace('\r', '')
                cleaned_record[key] = cleaned_value
            else:
                cleaned_record[key] = value
        
        return cleaned_record
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@app.get("/api/reports")
def get_reports(filename: str = DEFAULT_EXCEL_FILE):
    try:
        logging.debug(f"Fetching reports for {filename} from {EXCEL_DIR}")
        # Get dataframe from cache
        df = get_cached_dataframe(filename, '営業日報')
        
        # Clean up column names (remove newlines and strip)
        df.columns = [str(col).replace('\n', '').strip() for col in df.columns]
        
        # Rename specific columns to match frontend expectations
        df = df.rename(columns={
            '得意先CD.': '得意先CD',
            '訪問先名得意先名': '訪問先名',
            '直送先CD.': '直送先CD',
            '直送先名.': '直送先名',
            'コメント': '上長コメント',  # Excel uses 'コメント' for manager comment
            'コメント返信欄': 'コメント返信欄'  # Keep as-is for reply field
        })
        
        # Replace all NaN, infinity, and null values with None
        # Use fillna to replace NaN with None
        df = df.fillna(value='')
        
        # Convert dates to string to avoid serialization issues
        if '日付' in df.columns:
             df['日付'] = df['日付'].astype(str)
             
        # システム確認用デザインNo.が空の場合、デザイン依頼No.で補完する
        if 'システム確認用デザインNo.' in df.columns and 'デザイン依頼No.' in df.columns:
            import numpy as np
            df['システム確認用デザインNo.'] = np.where(
                (df['システム確認用デザインNo.'] == '') | df['システム確認用デザインNo.'].isna(), 
                df['デザイン依頼No.'], 
                df['システム確認用デザインNo.']
            )
        
        # Convert to dict and manually clean any remaining problematic values
        records = df.to_dict(orient="records")
        
        # Clean the records to ensure JSON compatibility
        import math
        cleaned_records = []
        for record in records:
            cleaned_record = {}
            for key, value in record.items():
                if isinstance(value, float):
                    if math.isnan(value) or math.isinf(value):
                        cleaned_record[key] = None
                    # Convert customer code to string without decimal
                    elif key in ['得意先CD', '直送先CD'] and not math.isnan(value):
                        cleaned_record[key] = str(int(value))
                    else:
                        cleaned_record[key] = value
                elif value == '':
                    cleaned_record[key] = None
                elif isinstance(value, str):
                    # Replace Excel's carriage return artifacts with proper newlines
                    import re
                    cleaned_value = re.sub(r'_x000D_', '\n', value)
                    cleaned_value = cleaned_value.replace('\r', '')
                    cleaned_record[key] = cleaned_value
                else:
                    cleaned_record[key] = value
            cleaned_records.append(cleaned_record)

        return cleaned_records
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@app.get("/api/interviewers/{customer_cd}")
def get_interviewers(
    customer_cd: str, 
    filename: str = DEFAULT_EXCEL_FILE,
    customer_name: Optional[str] = None,
    delivery_name: Optional[str] = None
):
    """Get list of interviewers for a specific customer with optional name filtering"""
    try:
        # Get dataframe from cache
        df = get_cached_dataframe(filename, '営業日報')
        
        # Clean up column names
        df.columns = [str(col).replace('\n', '') for col in df.columns]
        
        # Rename specific columns to standard names
        df = df.rename(columns={
            '得意先CD.': '得意先CD',
            '直送先CD.': '直送先CD',
            '訪問先名得意先名': '訪問先名', # Case where \n was removed
            '訪問先名\n得意先名': '訪問先名', # Just in case
        })
        
        # Convert customer_cd to float for matching (Excel stores as float)
        try:
            customer_cd_float = float(customer_cd)
        except ValueError:
            # If conversion fails, try string matching
            customer_cd_float = customer_cd
        
        # Base filter: customer code
        # Handle potential float/string mismatch by trying both if needed, but usually float is correct for number-like CDs
        customer_reports = df[df['得意先CD'] == customer_cd_float]
        
        # Name filtering logic
        if delivery_name:
            # If delivery_name is provided, filter by it
            # Ensure '直送先名' column exists
            if '直送先名' in customer_reports.columns:
                customer_reports = customer_reports[customer_reports['直送先名'] == delivery_name]
        
        # If no delivery_name, or if we want to also filter by customer_name when delivery_name is NOT set
        # The requirement is "past interviewers matching ... Direct Delivery Name (if provided) OR Customer Name (if matching)"
        # If we selected a delivery destination, we used the block above.
        # If we selected a customer (without direct delivery), we should strictly filter by that customer name if possible, 
        # to avoid showing interviewers from other branches if they share the same Customer CD (rare but possible).
        # OR more importantly, if we are in "Direct Delivery" mode, we already filtered.
        # If we are NOT in "Direct Delivery" mode (delivery_name is None), we might want to filter by customer_name
        # to ensure we don't pick up rows that HAVE a direct delivery name (i.e. different destination).
        elif customer_name:
             if '訪問先名' in customer_reports.columns:
                 # Filter rows where Visit Name contains customer_name OR matches exactly
                 # Since '訪問先名' in DB might contain 'Customer Name Direct Delivery Name' now,
                 # strict matching might be tricky.
                 # But usually '訪問先名' == 'Customer Name' for standard records.
                 # Let's try exact match or contains.
                 # Also, we might want to EXCLUDE rows that have a '直送先名' if we are selecting a generic customer?
                 # User said: " 直近のログではなく、得意先名や直送先名が一致する過去に入力した面談者のみ"
                 
                 # If delivery_name is NOT provided, it effectively means we want records for the main customer.
                 # So we should probably match rows where Visit Name is exactly Customer Name, OR contains it.
                 # Let's use simple filtering: if customer_name is passed, try to filter by it.
                 # But keep in mind 訪問先名 might include 直送先名 suffix now.
                 pass

        # Updated Strategy for clarity:
        # 1. Base: Customer CD
        # 2. If delivery_name provided: matches '直送先名' == delivery_name
        # 3. If delivery_name NOT provided but customer_name provided: matches '訪問先名' == customer_name OR '直送先名' is Empty/NaN
        #    (This covers the case where we selected the main customer and want to exclude direct delivery branches)
        
        if delivery_name and '直送先名' in customer_reports.columns:
            customer_reports = customer_reports[customer_reports['直送先名'] == delivery_name]
        elif '直送先名' in customer_reports.columns:
             # If no delivery name specified, we prefer rows that ALSO have no delivery name specified
             # This avoids suggesting interviewers specific to a branch when we selected the HQ
             customer_reports = customer_reports[
                 customer_reports['直送先名'].isna() | 
                 (customer_reports['直送先名'] == '') | 
                 (customer_reports['直送先名'] == None)
             ]

        interviewers = customer_reports['面談者'].dropna().unique().tolist()
        
        # Remove empty strings and sort
        interviewers = [str(i).strip() for i in interviewers if str(i).strip() and str(i).strip() != 'nan']
        interviewers = sorted(set(interviewers))
        
        return {"customer_cd": customer_cd, "interviewers": interviewers}
    except Exception as e:
        logging.error(f"Error in get_interviewers: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))

@app.get("/api/designs/{customer_cd}")
def get_designs(customer_cd: str, delivery_name: Optional[str] = None, filename: str = DEFAULT_EXCEL_FILE):
    """Get list of design requests for a specific customer (optionally filtered by delivery destination)"""
    try:
        logging.info(f"get_designs called: customer_cd={customer_cd}, delivery_name={delivery_name}")
        
        # Get dataframe from cache
        df = get_cached_dataframe(filename, '営業日報')
        
        # Clean up column names
        df.columns = [str(col).replace('\n', '') for col in df.columns]
        
        # Rename specific columns
        df = df.rename(columns={
            '得意先CD.': '得意先CD',
            '直送先名.': '直送先名'
        })
        
        logging.info(f"Columns in dataframe: {list(df.columns)[:20]}...")  # Log first 20 columns
        
        # Convert customer_cd to float for matching
        try:
            customer_cd_float = float(customer_cd)
        except ValueError:
            customer_cd_float = customer_cd
        
        # Filter by customer code
        customer_reports = df[df['得意先CD'] == customer_cd_float].copy()
        logging.info(f"After customer filter: {len(customer_reports)} rows")
        
        # Filter by delivery destination if provided
        if delivery_name and '直送先名' in customer_reports.columns:
            before_count = len(customer_reports)
            customer_reports = customer_reports[customer_reports['直送先名'] == delivery_name]
            logging.info(f"After delivery_name filter ({delivery_name}): {len(customer_reports)} rows (was {before_count})")
        else:
            logging.info(f"Skipping delivery_name filter: delivery_name={delivery_name}, has_column={'直送先名' in customer_reports.columns}")
        
        # Filter for records with design request number
        design_reports = customer_reports[customer_reports['デザイン依頼No.'].notna()]
        logging.info(f"After design no filter: {len(design_reports)} rows")
        
        # Get unique design request numbers
        unique_design_nos = design_reports['デザイン依頼No.'].unique()
        
        designs = []
        for design_no in unique_design_nos:
            # Get all records for this design number
            design_records = design_reports[design_reports['デザイン依頼No.'] == design_no]
            
            # Get the latest record (assuming lower down in Excel is newer, or we could sort by date if available)
            # Here we just take the last one in the dataframe which corresponds to the last row in Excel
            latest_record = design_records.iloc[-1]
            
            # Get the status
            status = str(latest_record['デザイン進捗状況']) if pd.notna(latest_record['デザイン進捗状況']) else ""
            
            # Skip designs with completed/rejected statuses: 出稿, 不採用(コンペ負け), 不採用(企画倒れ)
            if '出稿' in status or 'コンペ負け' in status or '企画倒れ' in status:
                continue
            
            design_info = {
                "デザイン依頼No": design_no,
                "デザイン名": str(latest_record['デザイン名']) if pd.notna(latest_record['デザイン名']) else "",
                "デザイン種別": str(latest_record['デザイン種別']) if pd.notna(latest_record['デザイン種別']) else "",
                "デザイン進捗状況": status,
                "デザイン提案有無": str(latest_record['デザイン提案有無']) if pd.notna(latest_record['デザイン提案有無']) else ""
            }
            designs.append(design_info)
            
        return {"customer_cd": customer_cd, "designs": designs}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.post("/api/reports")
def add_report(report: ReportInput, background_tasks: BackgroundTasks, filename: str = DEFAULT_EXCEL_FILE):
    logging.info(f"DEBUG ADD_REPORT: Received payload: {report.model_dump()}")
    excel_file = os.path.join(EXCEL_DIR, filename)
    if not os.path.exists(excel_file):
        raise HTTPException(status_code=404, detail=f"Excel file '{filename}' not found")
    
    try:
        # Load workbook with openpyxl to preserve formulas and macros
        wb = openpyxl.load_workbook(excel_file, keep_vba=True)
        ws = wb['営業日報']
        
        # 得意先_Listから現目標を取得
        current_target = ""
        if report.得意先CD and '得意先_List' in wb.sheetnames:
            customer_ws = wb['得意先_List']
            customer_cd = str(report.得意先CD).strip()
            direct_delivery_cd = str(report.直送先CD).strip() if report.直送先CD else ""
            
            # 得意先_Listの構造: A=得意先CD, B=直送先CD, ..., J=現目標
            for row in range(2, customer_ws.max_row + 1):
                cell_customer_cd = customer_ws.cell(row=row, column=1).value
                cell_dd_cd = customer_ws.cell(row=row, column=2).value
                
                if cell_customer_cd is not None:
                    # 得意先CDを文字列に変換（floatの場合は整数に）
                    if isinstance(cell_customer_cd, float):
                        cell_customer_cd = str(int(cell_customer_cd))
                    else:
                        cell_customer_cd = str(cell_customer_cd).strip()
                    
                    # 直送先CDも同様に処理
                    if cell_dd_cd is not None:
                        if isinstance(cell_dd_cd, float):
                            cell_dd_cd = str(int(cell_dd_cd))
                        else:
                            cell_dd_cd = str(cell_dd_cd).strip()
                    else:
                        cell_dd_cd = ""
                    
                    # マッチング条件: 得意先CDが一致 AND (直送先CDが一致 OR 両方空)
                    if cell_customer_cd == customer_cd:
                        if direct_delivery_cd:
                            # 直送先が指定されている場合は直送先CDも一致する必要がある
                            if cell_dd_cd == direct_delivery_cd:
                                target_value = customer_ws.cell(row=row, column=10).value  # J列=現目標
                                if target_value:
                                    current_target = str(target_value).strip()
                                break
                        else:
                            # 直送先が指定されていない場合は、直送先CDが空の行を探す
                            if not cell_dd_cd:
                                target_value = customer_ws.cell(row=row, column=10).value  # J列=現目標
                                if target_value:
                                    current_target = str(target_value).strip()
                                break
            
            logging.debug(f"Found current_target for {customer_cd}: {current_target}")
        
        # Find the maximum management number and its row by scanning all rows
        max_mgmt_num = 0
        max_mgmt_row = 1  # Default to header row if no data found
        for row in range(2, ws.max_row + 1):  # Start from row 2 (skip header)
            mgmt_num = ws.cell(row=row, column=1).value
            try:
                if mgmt_num is not None:
                    val = int(mgmt_num)
                    if val > max_mgmt_num:
                        max_mgmt_num = val
                        max_mgmt_row = row
            except (ValueError, TypeError):
                continue
        
        logging.debug(f"Max Mgmt Num: {max_mgmt_num}, Max Mgmt Row: {max_mgmt_row}")

        # Increment to get new management number
        new_mgmt_num = max_mgmt_num + 1
        
        # Insert at the row immediately after the last management number
        next_row = max_mgmt_row + 1
        logging.debug(f"Writing to Row: {next_row}")
        
        # Prepare the data to write
        # Adjust column indices based on actual Excel structure (251113_2026-_-_008.xlsm)
        # Also copy styles from the previous row (max_mgmt_row) to the new row (next_row)
        
        from copy import copy

        def copy_style(source_cell, target_cell):
            if source_cell.has_style:
                target_cell.font = copy(source_cell.font)
                target_cell.border = copy(source_cell.border)
                target_cell.fill = copy(source_cell.fill)
                target_cell.number_format = copy(source_cell.number_format)
                target_cell.protection = copy(source_cell.protection)
                target_cell.alignment = copy(source_cell.alignment)

        # Define the columns to write to and their values
        # 2026年度版カラム構造（K列に「得意先目標」が追加）
        
        # 自由入力（得意先CDが空で、訪問先名が入っている場合）は、ダミーの得意先CD「999999」を入れる
        write_customer_cd = report.得意先CD
        if not write_customer_cd and report.訪問先名:
            write_customer_cd = "999999"

        columns_to_write = {
            1: new_mgmt_num,        # A: 管理番号
            2: report.日付,          # B: 日付
            3: report.行動内容,       # C: 行動内容
            4: report.エリア,         # D: エリア
            5: write_customer_cd,    # E: 得意先CD.
            6: report.直送先CD,       # F: 直送先CD.
            7: report.訪問先名,       # G: 訪問先名/得意先名
            8: report.直送先名,       # H: 直送先名
            9: report.重点顧客,       # I: 重点顧客
            10: report.ランク,        # J: ランク
            11: current_target,       # K: 得意先目標（得意先_Listから自動取得）
            12: report.面談者,        # L: 面談者
            13: report.滞在時間,      # M: 滞在時間
            14: report.デザイン提案有無,  # N: デザイン提案有無
            15: report.デザイン種別,   # O: デザイン種別
            16: report.デザイン名,     # P: デザイン名
            17: report.デザイン進捗状況, # Q: デザイン進捗状況
            18: report.デザイン依頼No,  # R: デザイン依頼No.
            19: report.商談内容,       # S: 商談内容
            20: report.提案物,         # T: 提案物
            21: report.次回プラン,     # U: 次回プラン
            22: report.競合他社情報,   # V: 競合他社情報
            23: report.上長コメント,   # W: 上長コメント
            24: report.コメント返信欄  # X: コメント返信欄
        }

        for col_idx, value in columns_to_write.items():
            target_cell = ws.cell(row=next_row, column=col_idx)
            target_cell.value = value
            
            # Copy style from the row above (max_mgmt_row)
            # Ensure we are copying from a valid row
            if max_mgmt_row >= 2:
                source_cell = ws.cell(row=max_mgmt_row, column=col_idx)
                copy_style(source_cell, target_cell)


        
        # Save the workbook (Critical path - blocking)
        wb.save(excel_file)
        wb.close()

        # Create backup in background
        background_tasks.add_task(create_backup, excel_file)
        
        
        # Clear cache
        cache_key = (filename, '営業日報')
        if cache_key in CACHE:
            del CACHE[cache_key]
        
        return {
            "message": "Report added successfully", 
            "management_number": new_mgmt_num,
            "file_path": os.path.abspath(excel_file)
        }
    except PermissionError:
        raise HTTPException(
            status_code=409,
            detail="ファイルが開かれているため保存できません。Excelファイルを閉じてから再度実行してください。"
        )
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

# コメント更新専用エンドポイント（楽観的ロックなし）
class CommentInput(BaseModel):
    上長コメント: Optional[str] = None
    コメント返信欄: Optional[str] = None

# 承認チェック更新専用
class ApprovalInput(BaseModel):
    上長: Optional[str] = None
    山澄常務: Optional[str] = None
    岡本常務: Optional[str] = None
    中野次長: Optional[str] = None
    既読チェック: Optional[str] = None

# 後方互換性のため
class ReplyInput(BaseModel):
    コメント返信欄: str

@app.patch("/api/reports/{management_number}/reply")
def update_report_reply(management_number: int, reply: ReplyInput, background_tasks: BackgroundTasks, filename: str = DEFAULT_EXCEL_FILE):
    """コメント返信欄のみを更新（安全な保存）"""
    import tempfile
    import shutil
    
    logging.debug(f"update_report_reply: management_number={management_number}, reply={reply.コメント返信欄}")
    try:
        excel_file = os.path.join(EXCEL_DIR, filename)
        logging.debug(f"excel_file: {excel_file}")
        
        wb = openpyxl.load_workbook(excel_file, keep_vba=True)
        logging.debug("workbook loaded")
        if '営業日報' not in wb.sheetnames:
            raise HTTPException(status_code=404, detail="Sheet '営業日報' not found")
        
        ws = wb['営業日報']
        logging.debug(f"worksheet max_row: {ws.max_row}")
        
        # Find the row
        target_row = None
        for row in range(2, ws.max_row + 1):
            val = ws.cell(row=row, column=1).value
            if val is not None:
                try:
                    if int(float(str(val))) == int(management_number):
                        target_row = row
                        break
                except (ValueError, TypeError):
                    if val == management_number:
                        target_row = row
                        break
        
        logging.debug(f"target_row: {target_row}")
        if not target_row:
            wb.close()
            raise HTTPException(status_code=404, detail=f"Report {management_number} not found")
        
        # 2026年度版: X列(24) = コメント返信欄
        ws.cell(row=target_row, column=24, value=reply.コメント返信欄)
        logging.debug("cell set, saving to temp file...")
        
        # 安全な保存: 一時ファイルに保存してから置き換え
        temp_dir = tempfile.gettempdir()
        temp_file = os.path.join(temp_dir, f"temp_{filename}")
        
        try:
            # 一時ファイルに保存
            wb.save(temp_file)
            wb.close()
            logging.debug(f"saved to temp: {temp_file}")
            
            # 一時ファイルが正常か確認（読み込みテスト）
            test_wb = openpyxl.load_workbook(temp_file, read_only=True)
            test_wb.close()
            logging.debug("temp file verified")
            
            # 元のファイルを一時ファイルで置き換え
            shutil.copy2(temp_file, excel_file)
            logging.debug("replaced original file")
            
        finally:
            # 一時ファイルを削除
            if os.path.exists(temp_file):
                try:
                    os.remove(temp_file)
                except:
                    pass
        
        # Clear cache
        cache_key = (filename, '営業日報')
        if cache_key in CACHE:
            del CACHE[cache_key]
        
        return {"success": True, "management_number": management_number}
    except HTTPException:
        raise
    except Exception as e:
        import traceback
        logging.error(f"update_report_reply: {e}")
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=str(e))

@app.patch("/api/reports/{management_number}/comment")
def update_report_comment(management_number: int, comment: CommentInput, background_tasks: BackgroundTasks, filename: str = DEFAULT_EXCEL_FILE):
    """上長コメントとコメント返信欄を個別に更新（安全な保存）"""
    import tempfile
    import shutil
    
    logging.debug(f"update_report_comment: management_number={management_number}, 上長コメント={comment.上長コメント}, コメント返信欄={comment.コメント返信欄}")
    try:
        excel_file = os.path.join(EXCEL_DIR, filename)
        
        wb = openpyxl.load_workbook(excel_file, keep_vba=True)
        if '営業日報' not in wb.sheetnames:
            raise HTTPException(status_code=404, detail="Sheet '営業日報' not found")
        
        ws = wb['営業日報']
        
        # Find the row
        target_row = None
        for row in range(2, ws.max_row + 1):
            val = ws.cell(row=row, column=1).value
            if val is not None:
                try:
                    if int(float(str(val))) == int(management_number):
                        target_row = row
                        break
                except (ValueError, TypeError):
                    if val == management_number:
                        target_row = row
                        break
        
        if not target_row:
            wb.close()
            raise HTTPException(status_code=404, detail=f"Report {management_number} not found")
        
        # Update only provided fields
        # 2026年度版: W列(23) = 上長コメント, X列(24) = コメント返信欄
        if comment.上長コメント is not None:
            ws.cell(row=target_row, column=23, value=comment.上長コメント)
        if comment.コメント返信欄 is not None:
            ws.cell(row=target_row, column=24, value=comment.コメント返信欄)
        
        # 安全な保存: 一時ファイルに保存してから置き換え
        temp_dir = tempfile.gettempdir()
        temp_file = os.path.join(temp_dir, f"temp_{filename}")
        
        try:
            wb.save(temp_file)
            wb.close()
            
            # 一時ファイルが正常か確認
            test_wb = openpyxl.load_workbook(temp_file, read_only=True)
            test_wb.close()
            
            # 元のファイルを置き換え
            shutil.copy2(temp_file, excel_file)
        finally:
            if os.path.exists(temp_file):
                try:
                    os.remove(temp_file)
                except:
                    pass
        
        # Clear cache
        cache_key = (filename, '営業日報')
        if cache_key in CACHE:
            del CACHE[cache_key]
        
        # Create backup in background
        background_tasks.add_task(create_backup, excel_file)
        
        return {"success": True, "management_number": management_number}
    except HTTPException:
        raise
    except Exception as e:
        import traceback
        logging.error(f"update_report_comment: {e}")
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=str(e))

@app.patch("/api/reports/{management_number}/approval")
def update_report_approval(management_number: int, approval: ApprovalInput, background_tasks: BackgroundTasks, filename: str = DEFAULT_EXCEL_FILE):
    """承認チェック（上長、山澄常務、岡本常務、中野次長、既読チェック）を個別に更新"""
    import tempfile
    import shutil
    
    logging.debug(f"update_report_approval: management_number={management_number}")
    try:
        excel_file = os.path.join(EXCEL_DIR, filename)
        
        wb = openpyxl.load_workbook(excel_file, keep_vba=True)
        if '営業日報' not in wb.sheetnames:
            raise HTTPException(status_code=404, detail="Sheet '営業日報' not found")
        
        ws = wb['営業日報']
        
        # Find the row
        target_row = None
        for row in range(2, ws.max_row + 1):
            val = ws.cell(row=row, column=1).value
            if val is not None:
                try:
                    if int(float(str(val))) == int(management_number):
                        target_row = row
                        break
                except (ValueError, TypeError):
                    if val == management_number:
                        target_row = row
                        break
        
        if not target_row:
            wb.close()
            raise HTTPException(status_code=404, detail=f"Report {management_number} not found")
        
        # Update only provided fields
        # 2026年度版カラムマッピング: Y=上長(25), Z=山澄常務(26), AA=岡本常務(27), AB=中野次長(28), AC=既読チェック(29)
        column_mapping = {
            '上長': 25,           # Y列
            '山澄常務': 26,        # Z列
            '岡本常務': 27,        # AA列
            '中野次長': 28,        # AB列
            '既読チェック': 29     # AC列
        }
        
        if approval.上長 is not None:
            ws.cell(row=target_row, column=column_mapping['上長'], value=approval.上長)
        if approval.山澄常務 is not None:
            ws.cell(row=target_row, column=column_mapping['山澄常務'], value=approval.山澄常務)
        if approval.岡本常務 is not None:
            ws.cell(row=target_row, column=column_mapping['岡本常務'], value=approval.岡本常務)
        if approval.中野次長 is not None:
            ws.cell(row=target_row, column=column_mapping['中野次長'], value=approval.中野次長)
        if approval.既読チェック is not None:
            ws.cell(row=target_row, column=column_mapping['既読チェック'], value=approval.既読チェック)
        
        # 安全な保存
        temp_dir = tempfile.gettempdir()
        temp_file = os.path.join(temp_dir, f"temp_{filename}")
        
        try:
            wb.save(temp_file)
            wb.close()
            
            test_wb = openpyxl.load_workbook(temp_file, read_only=True)
            test_wb.close()
            
            shutil.copy2(temp_file, excel_file)
        finally:
            if os.path.exists(temp_file):
                try:
                    os.remove(temp_file)
                except:
                    pass
        
        # Clear cache
        cache_key = (filename, '営業日報')
        if cache_key in CACHE:
            del CACHE[cache_key]
        
        # Create backup in background
        background_tasks.add_task(create_backup, excel_file)
        
        return {"success": True, "management_number": management_number}
    except HTTPException:
        raise
    except Exception as e:
        import traceback
        logging.error(f"update_report_approval: {e}")
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=str(e))

@app.post("/api/reports/{management_number}")
def update_report(management_number: int, report: ReportInput, background_tasks: BackgroundTasks, filename: str = DEFAULT_EXCEL_FILE):
    """既存の日報を更新（全項目対応）"""
    logging.info(f"update_report called: management_number={management_number}, original_values={report.original_values}")
    try:
        excel_file = os.path.join(EXCEL_DIR, filename)
        
        # Load the workbook
        wb = openpyxl.load_workbook(excel_file, keep_vba=True)
        if '営業日報' not in wb.sheetnames:
            raise HTTPException(status_code=404, detail="Sheet '営業日報' not found")
             
        ws = wb['営業日報']
        
        # Find the row with the matching management number
        target_row = None
        for row in range(2, ws.max_row + 1):
            val = ws.cell(row=row, column=1).value
            if val is not None:
                try:
                    if int(float(str(val))) == int(management_number):
                        target_row = row
                        break
                except (ValueError, TypeError):
                    if val == management_number:
                        target_row = row
                        break
        
        if not target_row:
            raise HTTPException(status_code=404, detail=f"Report with management number {management_number} not found")

        # --- Optimistic Locking Check ---
        if report.original_values:
            logging.debug(f"Performing conflict check for Report {management_number}")
            
            # Fields to check for conflicts (critical text fields)
            check_fields = {
                23: '上長コメント',
                24: 'コメント返信欄',
                19: '商談内容'
            }
            
            conflicts = []
            for col_idx, field_name in check_fields.items():
                current_val = ws.cell(row=target_row, column=col_idx).value
                current_str = str(current_val) if current_val is not None else ""
                
                original_val = report.original_values.get(field_name, "")
                original_str = str(original_val) if original_val is not None else ""
                
                # Normalize newlines for comparison
                current_str = current_str.replace('\r\n', '\n').replace('\r', '\n').strip()
                original_str = original_str.replace('\r\n', '\n').replace('\r', '\n').strip()
                
                if current_str != original_str:
                    logging.warning(f"CONFLICT: Field '{field_name}' changed. Current: '{current_str}' vs Original: '{original_str}'")
                    conflicts.append(field_name)
            
            if conflicts:
                conflict_msg = ", ".join(conflicts)
                raise HTTPException(
                    status_code=409, 
                    detail=f"他の方が編集しました（{conflict_msg}）。最新の情報を読み込んでからやり直してください。"
                )
        # --------------------------------
        
        # Update all fields (2026年度版カラム構造 - K列に得意先目標が追加)
        columns_to_write = {
            2: report.日付,              # B: 日付
            3: report.行動内容,           # C: 行動内容
            4: report.エリア,             # D: エリア
            5: report.得意先CD,           # E: 得意先CD.
            6: report.直送先CD,           # F: 直送先CD.
            7: report.訪問先名,           # G: 訪問先名/得意先名
            8: report.直送先名,           # H: 直送先名
            9: report.重点顧客,           # I: 重点顧客
            10: report.ランク,            # J: ランク
            # 11: 得意先目標 (K) - 手動入力のため省略
            12: report.面談者,            # L: 面談者
            13: report.滞在時間,          # M: 滞在時間
            14: report.デザイン提案有無,   # N: デザイン提案有無
            15: report.デザイン種別,       # O: デザイン種別
            16: report.デザイン名,         # P: デザイン名
            17: report.デザイン進捗状況,   # Q: デザイン進捗状況
            18: report.デザイン依頼No,     # R: デザイン依頼No.
            19: report.商談内容,           # S: 商談内容
            20: report.提案物,             # T: 提案物
            21: report.次回プラン,         # U: 次回プラン
            22: report.競合他社情報,       # V: 競合他社情報
            23: report.上長コメント,       # W: 上長コメント
            24: report.コメント返信欄      # X: コメント返信欄
        }

        for col_idx, value in columns_to_write.items():
            ws.cell(row=target_row, column=col_idx, value=value)
        
        # Save the workbook (Critical path - blocking)
        wb.save(excel_file)
        wb.close()
        
        # Create backup in background
        background_tasks.add_task(create_backup, excel_file)
        
        # Clear cache for this file
        cache_key = (filename, '営業日報')
        if cache_key in CACHE:
            del CACHE[cache_key]
        
        return {"message": "Report updated successfully", "management_number": management_number}
    except HTTPException:
        raise
    except PermissionError:
        raise HTTPException(
            status_code=409,
            detail="ファイルが開かれているため保存できません。Excelファイルを閉じてから再度実行してください。"
        )
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@app.delete("/api/reports/{management_number}")
def delete_report(management_number: int, filename: str = DEFAULT_EXCEL_FILE):
    """指定された管理番号の日報を削除"""
    try:
        excel_file = os.path.join(EXCEL_DIR, filename)
        
        # Load the workbook
        wb = openpyxl.load_workbook(excel_file, keep_vba=True)
        if '営業日報' not in wb.sheetnames:
            raise HTTPException(status_code=404, detail="Sheet '営業日報' not found")
             
        ws = wb['営業日報']
        
        # Find the row with the matching management number
        target_row = None
        for row in range(2, ws.max_row + 1):
            val = ws.cell(row=row, column=1).value
            if val is not None:
                try:
                    if int(float(str(val))) == int(management_number):
                        target_row = row
                        break
                except (ValueError, TypeError):
                    if val == management_number:
                        target_row = row
                        break
        
        if not target_row:
            raise HTTPException(status_code=404, detail=f"Report with management number {management_number} not found")
        
        # Delete the row
        ws.delete_rows(target_row, 1)
        
        # Save the workbook
        wb.save(excel_file)
        wb.close()
        
        # Clear cache for this file
        cache_key = (filename, '営業日報')
        if cache_key in CACHE:
            del CACHE[cache_key]
        
        return {"message": "Report deleted successfully", "management_number": management_number}
    except HTTPException:
        raise
    except PermissionError:
        raise HTTPException(
            status_code=409,
            detail="ファイルが開かれているため保存できません。Excelファイルを閉じてから再度実行してください。"
        )
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.post("/api/upload")
async def upload_file(file: UploadFile = File(...)):
    """Upload an Excel file to the backend directory"""
    try:
        # Validate file extension
        if not file.filename.endswith(('.xlsx', '.xlsm')):
            raise HTTPException(status_code=400, detail="Only .xlsx and .xlsm files are allowed")
        
        # Save the uploaded file
        file_path = os.path.join(EXCEL_DIR, file.filename)
        
        with open(file_path, "wb") as buffer:
            shutil.copyfileobj(file.file, buffer)
        
        return {
            "message": "File uploaded successfully",
            "filename": file.filename,
            "path": file_path
        }
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


# Set up logging
import logging

logging.basicConfig(
    filename='debug.log',
    level=logging.DEBUG,
    format='%(asctime)s - %(levelname)s - %(message)s',
    encoding='utf-8' # Ensure we can log Japanese characters
)

@app.get("/api/images/list")
def get_design_images(filename: str):
    """
    Get list of images from the matching folder in Design Data directory.
    Target directory: \\Asahipack02\\社内書類ｎｅｗ\\01：部署別　営業部\\03：デザインデータ
    Logic: Extract name from filename '...【Name】.xlsm' -> Search folder containing 'Name'
    """
    DESIGN_DIR = r"\\Asahipack02\社内書類ｎｅｗ\01：部署別　営業部\03：デザインデータ"
    
    logging.info(f"--- get_design_images called with filename: {filename} ---")
    
    # 0. 手動マッピングチェック
    # 特定のファイル名を特定のフォルダ名または検索語句にマッピング
    FOLDER_MAPPING = {
        # 正確なファイル名 -> 正確なターゲットフォルダ名（または検索する部分文字列）
        "本社006　2025年度用日報【木村（拓）MGR】.xlsm": "大阪本社　05：木村（拓）",
    }
    
    matched_dir = None

    if filename in FOLDER_MAPPING:
        mapped_target = FOLDER_MAPPING[filename]
        logging.info(f"Manual mapping found for {filename}: {mapped_target}")
        
        # マッピングされたディレクトリが存在するか確認
        path_check = os.path.join(DESIGN_DIR, mapped_target)
        if os.path.isdir(path_check):
            matched_dir = mapped_target
            logging.info(f"Mapped directory verified: {matched_dir}")
        else:
             logging.warning(f"Mapped directory not found: {path_check}")
             # 通常の検索にフォールバックするか失敗させるか？ここではフォールバックします。
    
    if not matched_dir:
        try:
            # Extract name from filename (e.g., 本社009　2025年度用日報【沖本】.xlsm -> 沖本)
            import re
            # Helper for normalization
            def normalize_text(text):
                # Convert full-width parens and space to half-width
                text = text.replace('（', '(').replace('）', ')').replace('　', ' ')
                # Strip whitespace
                return text.strip()

            match = re.search(r'【(.*?)】', filename)
            if not match:
                logging.warning("Regex match failed for filename")
                # Fallback extraction?
                target_name = os.path.splitext(os.path.basename(filename))[0]
                logging.info(f"Fallback extracted name: {target_name}")
            else:
                target_name = match.group(1)
                logging.info(f"Regex extracted name: {target_name}")

            normalized_target = normalize_text(target_name)
            logging.info(f"Normalized target: {normalized_target}")
            
            # 接尾辞を削除
            # MGR/Mgr をリストに追加
            stripped_target = re.sub(r'(MGR|Mgr|次長|課長|部長|係長|主任|担当|顧問|専務|常務|社長)$', '', normalized_target, flags=re.IGNORECASE)
            logging.info(f"Stripped target: {stripped_target}")

            logging.debug(f"Searching for folder containing '{target_name}' (Norm: {normalized_target}) in {DESIGN_DIR}")
            
            if not os.path.exists(DESIGN_DIR):
                 logging.error(f"Design directory not found: {DESIGN_DIR}")
                 return {"message": "Design directory not found", "images": []}

            # Find matching directory
            
            try:
                dir_list = os.listdir(DESIGN_DIR)
                # logging.debug(f"Directory listing (first 5): {dir_list[:5]}")
            except Exception as e:
                logging.error(f"Failed to list directory: {e}")
                return {"message": f"Failed to access design dir: {e}", "images": []}

            # 1. Try exact match (normalized)
            for item in dir_list:
                if not os.path.isdir(os.path.join(DESIGN_DIR, item)):
                    continue
                    
                norm_item = normalize_text(item)
                if normalized_target in norm_item:
                    matched_dir = item
                    logging.info(f"Match found (Normalized): {item}")
                    break
            
            # 2. If no match, try suffix stripping (e.g. 山下(和)次長 -> 山下(和))
            if not matched_dir:
                if stripped_target != normalized_target:
                     logging.info("Retrying with stripped name...")
                     for item in dir_list:
                        if not os.path.isdir(os.path.join(DESIGN_DIR, item)):
                            continue
                        norm_item = normalize_text(item)
                        if stripped_target in norm_item:
                            matched_dir = item
                            logging.info(f"Match found (Stripped): {item}")
                            break
            
            if not matched_dir:
                logging.warning(f"No folder found for target: {normalized_target} / {stripped_target}")
                return {"message": f"No folder found for '{target_name}'", "images": []}
                
        except Exception as e:
             logging.error(f"Error during folder search logic: {e}")
             raise HTTPException(status_code=500, detail=str(e))
            
    try:
        target_path = os.path.join(DESIGN_DIR, matched_dir)
        logging.info(f"Target path: {target_path}")
        
        # List images (recursively or just top level? Starting with top level + shallow)
        # Extensions to look for
        valid_extensions = ('.jpg', '.jpeg', '.png', '.gif', '.bmp', '.webp', '.pdf')
        
        image_files = []
        for root, dirs, files in os.walk(target_path):
            for file in files:
                if file.lower().endswith(valid_extensions):
                    # Create a relative path from DESIGN_DIR for the client to request
                    # e.g., "大阪本社　09：沖本\image.jpg"
                    full_path = os.path.join(root, file)
                    rel_path = os.path.relpath(full_path, DESIGN_DIR)
                    try:
                        mtime = os.path.getmtime(full_path)
                    except:
                        mtime = 0
                    
                    image_files.append({
                        "name": file,
                        "path": rel_path, # Path identifier to send back to serve endpoint
                        "folder": matched_dir,
                        "mtime": mtime
                    })
            if len(image_files) > 100:
                break
        
        # Sort by mtime descending (newest first)
        image_files.sort(key=lambda x: x['mtime'], reverse=True)
        
        return {"images": image_files, "folder": matched_dir}

    except Exception as e:
        logging.exception("Error in get_design_images")
        logging.error(f"Error listing images: {e}")
        raise HTTPException(status_code=500, detail=str(e))

@app.get("/api/images/content")
def serve_design_image(path: str):
    """
    Serve the image file content.
    path: Relative path from DESIGN_DIR (e.g., "大阪本社　09：沖本\image.jpg")
    """
    DESIGN_DIR = r"\\Asahipack02\社内書類ｎｅｗ\01：部署別　営業部\03：デザインデータ"
    
    try:
        # Security check: Prevent directory traversal
        # relpath needs to be safe. 
        # Since we construct it ourselves in /images/list, it should be fine, but good to check.
        safe_path = os.path.normpath(os.path.join(DESIGN_DIR, path))
        
        if not safe_path.startswith(DESIGN_DIR):
            raise HTTPException(status_code=403, detail="Access denied")
            
        if not os.path.exists(safe_path):
            raise HTTPException(status_code=404, detail="Image not found")
            
        from fastapi.responses import FileResponse
        return FileResponse(safe_path)
    except Exception as e:
         raise HTTPException(status_code=500, detail=str(e))

@app.get("/api/images/search")
def search_design_images(query: str, filename: Optional[str] = None):
    """
    Search for images matching the query (Design No) in the Design Data directory.
    If filename is provided (e.g. '見上.xlsm'), it tries to find a matching user folder first (e.g. '08：見上').
    Recursively searches subfolders.
    """
    DESIGN_DIR = r"\\Asahipack02\社内書類ｎｅｗ\01：部署別　営業部\03：デザインデータ"
    
    logging.info(f"--- search_design_images called. Query: {query}, Filename: {filename} ---")

    if not query or len(query.strip()) < 2:
        return {"message": "Query too short", "images": []}
        
    # Helper for normalization
    def normalize_text(text):
        # Convert full-width parens and space to half-width
        text = text.replace('（', '(').replace('）', ')').replace('　', ' ')
        # Strip whitespace
        return text.strip()

    try:
        if not os.path.exists(DESIGN_DIR):
             return {"message": "Design directory not found", "images": []}

        search_roots = [DESIGN_DIR]
        
        # Optimize: Try to find specific user folder based on filename
        found_folder = None
        if filename:
            try:
                # Extract name part
                # 1. Try to extract from 【】
                import re
                match = re.search(r'【(.*?)】', filename)
                if match:
                    name_part = match.group(1)
                else:
                    # 拡張子削除にフォールバック
                    name_part = os.path.splitext(os.path.basename(filename))[0]
                
                # 名前部分を正規化
                normalized_name = normalize_text(name_part)
                # 接尾辞（サフィックス）を削除したバージョンも用意
                stripped_name = re.sub(r'(MGR|Mgr|次長|課長|部長|係長|主任|担当|顧問|専務|常務|社長)$', '', normalized_name, flags=re.IGNORECASE)
                
                logging.info(f"Search optimization - Extracted: {name_part}, Norm: {normalized_name}, Stripped: {stripped_name}")

                # Use scandir for better performance on network drive for top-level listing
                with os.scandir(DESIGN_DIR) as it:
                    for entry in it:
                        if entry.is_dir():
                            norm_entry_name = normalize_text(entry.name)
                            # 抽出された名前（正規化済み）がフォルダ名（正規化済み）に含まれているか確認
                            if normalized_name in norm_entry_name:
                                found_folder = entry.path
                                logging.info(f"Optimization - Found folder (Norm): {entry.name}")
                                break
                            # 見つからない場合、接尾辞なしの名前を試す
                            if stripped_name != normalized_name and stripped_name in norm_entry_name:
                                found_folder = entry.path
                                logging.info(f"Optimization - Found folder (Stripped): {entry.name}")
                                break

                if found_folder:
                    logging.debug(f"Search target set to: {found_folder}")
                    search_roots = [found_folder]
            except Exception as e:
                logging.error(f"Failed to optimize search folder: {e}")
                logging.warning(f"Failed to optimize search folder: {e}")
                
        
        if found_folder:
            search_roots = [found_folder]
        else:
            if filename:
                 # print(f"WARN: Could not find user folder for {filename}. Aborting full scan to prevent timeout.")
                 logging.info("User folder not found from filename")
                 return {"message": "User folder not found from filename", "images": []}
            
            # If no filename provided, we search everything? That's dangerous too.
            search_roots = [DESIGN_DIR]

        image_files = []
        valid_extensions = ('.jpg', '.jpeg', '.png', '.gif', '.bmp', '.webp', '.pdf') 
        
        count = 0
        MAX_RESULTS = 50
        
        # Helper for safer walking on finicky network drives
        def safe_walk(directory, query_lower, extensions, max_depth=3, current_depth=0, parent_matches_query=False):
            results = []
            try:
                # Use listdir instead of scandir/walk to avoid hanging
                items = os.listdir(directory)
            except Exception as e:
                logging.warning(f"Failed to listdir {directory}: {e}")
                return results

            dirs_to_visit = []

            for name in items:
                full_path = os.path.join(directory, name)
                name_lower = name.lower()
                
                # Check file match
                # IF parent folder matched, we take ALL images.
                # IF not, we only take images matching query.
                is_file_match = False
                if parent_matches_query and name_lower.endswith(extensions):
                    is_file_match = True
                elif query_lower in name_lower and name_lower.endswith(extensions):
                    is_file_match = True
                
                if is_file_match:
                    try:
                        if os.path.isfile(full_path):
                            try:
                                rel_path = os.path.relpath(full_path, DESIGN_DIR)
                                folder_name = os.path.basename(directory)
                                try:
                                    mtime = os.path.getmtime(full_path)
                                except:
                                    mtime = 0
                                results.append({
                                    "name": name,
                                    "path": rel_path,
                                    "folder": folder_name,
                                    "mtime": mtime
                                })
                            except ValueError:
                                pass
                    except Exception:
                        pass
                
                # Identify directories for recursion
                if current_depth < max_depth:
                    next_parent_matches = parent_matches_query
                    
                    # Logic:
                    # 1. If parent already matched, we continue down (inheriting True).
                    # 2. If parent didn't match, verify if THIS folder matches.
                    if not next_parent_matches:
                        if query_lower in name_lower:
                            next_parent_matches = True
                    
                    # Recursion Filter:
                    # If we are NOT in a matching tree yet, we MUST skip unrelated folders 
                    # to avoid massive scan (timeout).
                    if not next_parent_matches:
                         # Skip unrelated folders
                         # BUT we must handle "Generic" folders like "Data", "Images", "Design", etc.
                         # AND we must skip "Specific" folders that don't match (e.g. "12345-1").
                         
                         # Heuristic:
                         # 1. If folder name looks like a Design ID (5+ digits, maybe hyphens), assume it's Specific.
                         #    If query is NOT in it, SKIP.
                         # 2. Otherwise, assume it's Generic (e.g. "Data", "2025", "01_Sales").
                         #    ENTER.
                         
                         is_specific_id = False
                         # Simple regex for design ID: typically 5-8 digits, optionally followed by hyphens/more digits
                         # e.g. 117675, 117675-1, 101219-106074-3
                         # But NOT "2025" (Year) or "01" (Section).
                         # Let's say "Specific" if it has 5 or more consecutive digits.
                         import re
                         if re.search(r'\d{5,}', name):
                             is_specific_id = True
                         
                         # Refinement: What if the query is SMALL? e.g. "555"?
                         # The query is usually a Design No (5-6 digits).
                         # If query is matches regex, it's specific.
                         
                         if is_specific_id and query_lower not in name_lower:
                             # It looks like a DIFFERENT specific ID. Skip.
                             continue
                         
                         # If it is generic (not super long digits) OR it matched query, we enter.
                         pass

                    if '.' in name and not name.startswith('.'):
                        continue 
                        
                    try:
                        if os.path.isdir(full_path):
                            dirs_to_visit.append((full_path, next_parent_matches))
                    except Exception:
                        pass
            
            # Recurse
            for subdir_path, matches_status in dirs_to_visit:
                 sub_results = safe_walk(subdir_path, query_lower, extensions, max_depth, current_depth + 1, matches_status)
                 results.extend(sub_results)
                 if len(results) >= 50: 
                     break
            
            return results

        # Main search loop
        try:
            for search_root in search_roots:
                logging.debug(f"Searching root: {search_root}")
                # Use custom walker
                # Initial parent_matches logic:
                # If the search_root ITSELF matches query (e.g. we targeted a specific user folder matched by filename, but query is DesignNo),
                # expected behavior: search for DesignNo INSIDE user folder.
                # So initial parent_matches = False usually.
                found_images = safe_walk(search_root, query.lower(), valid_extensions)
                image_files.extend(found_images)
                if len(image_files) >= MAX_RESULTS:
                    image_files = image_files[:MAX_RESULTS]
                    break
        except Exception as e:
            logging.error(f"Search loop failed: {e}")
            # Return whatever we found so far instead of 500
            pass
            
        # Sort by mtime descending (newest first)
        image_files.sort(key=lambda x: x['mtime'], reverse=True)

        return {"images": image_files, "query": query}

    except Exception as e:
        # print(f"Error searching images: {e}")
        # Log to file safely if needed, or just return detailed 500
        # Ensure 'e' conversion to string doesn't fail
        error_msg = "Unknown error"
        try:
            error_msg = str(e)
        except:
            pass
        raise HTTPException(status_code=500, detail=error_msg)


# --- Sales Data Integration ---
# --- Sales Data Integration (Global) ---
DATA_DIR = os.path.join(os.path.dirname(__file__), 'data')
SALES_CSV_PATH = os.path.join(DATA_DIR, 'sales_data.csv')
os.makedirs(DATA_DIR, exist_ok=True)

global_sales_df = None

def load_sales_data():
    """Loads sales data from CSV into global DataFrame."""
    global global_sales_df
    if not os.path.exists(SALES_CSV_PATH):
        logging.info("No existing sales data found.")
        return

    try:
        logging.info("Loading sales data from disk...")
        try:
            df = pd.read_csv(SALES_CSV_PATH, encoding='cp932')
        except:
            df = pd.read_csv(SALES_CSV_PATH, encoding='utf-8')
        
        df.columns = [str(col).strip() for col in df.columns]
        
        if '得意先コード' in df.columns:
            df['得意先コード'] = df['得意先コード'].astype(str).str.split('.').str[0]
            global_sales_df = df
            logging.info(f"Sales data loaded successfully. {len(df)} rows.")
        else:
            logging.error("Sales CSV missing '得意先コード' column.")
    except Exception as e:
        logging.error(f"Failed to load sales data: {e}")

# Load on startup
load_sales_data()

@app.post("/api/sales/upload")
async def upload_sales_csv(file: UploadFile = File(...)):
    """
    Uploads a global sales data CSV file, saves it, and unloads it into memory.
    """
    try:
        logging.info(f"Receiving sales CSV: {file.filename}")
        contents = await file.read()
        
        import io
        try:
            pd.read_csv(io.BytesIO(contents), encoding='cp932')
        except:
            try:
                pd.read_csv(io.BytesIO(contents), encoding='utf-8')
            except Exception as e:
                raise HTTPException(status_code=400, detail="Invalid CSV format. Please use Shift-JIS or UTF-8.")

        with open(SALES_CSV_PATH, "wb") as f:
            f.write(contents)
        
        logging.info("Sales CSV saved to disk.")
        load_sales_data()
        
        return {"message": "Sales data uploaded and processed successfully."}

    except Exception as e:
        logging.error(f"Error uploading sales CSV: {e}")
        raise HTTPException(status_code=500, detail=f"Upload failed: {str(e)}")


@app.get("/api/sales/all")
async def get_all_sales_data():
    """
    Retrieves ALL sales data as a list.
    """
    if global_sales_df is None:
        return []

    try:
        # Convert NaN to None for JSON compliance
        df_clean = global_sales_df.where(pd.notnull(global_sales_df), None)
        
        # Select relevant columns and rename for consistency
        records = []
        for _, row in df_clean.iterrows():
            records.append({
                "rank": row.get('順位'),
                "rank_class": row.get('ランク'),
                "customer_code": row.get('得意先コード'),
                "customer_name": row.get('得意先名称'),
                "sales_amount": row.get('売上金額'),
                "gross_profit": row.get('粗利金額'),
                "sales_yoy": row.get('前年対比率'),
                "sales_last_year": row.get('前年売上'),
                "profit_last_year": row.get('前年粗利'),
                "sales_2y_ago": row.get('前々年売上'),
                "profit_2y_ago": row.get('前々年粗利'),
                # Attempt to get area from '地域名称' or '地域' or Column M (index 12)
                "area": row.get('地域名称') or row.get('地域') or (row.iloc[12] if len(row) > 12 else None),
                # 担当者 from Column I (index 8)
                "sales_rep": row.get('担当者') or (row.iloc[8] if len(row) > 8 else None),
            })
            
        return records

    except Exception as e:
        logging.error(f"Error retrieving all sales data: {e}")
        raise HTTPException(status_code=500, detail=f"Error retrieving data: {str(e)}")


@app.get("/api/sales/{customer_code}")
async def get_sales_data(customer_code: str):
    """
    Retrieves sales data for a specific customer from the global dataset.
    """
    if global_sales_df is None:
        return {"found": False, "message": "Sales data not yet uploaded."}
    
    try:
        target_code = str(customer_code).split('.')[0]
        matched_row = global_sales_df[global_sales_df['得意先コード'] == target_code]
        
        if matched_row.empty:
            return {"found": False, "message": "Customer not found in sales data."}
        
        row = matched_row.iloc[0]
        
        def get_val(col):
            val = row.get(col)
            if pd.isna(val):
                return None
            if hasattr(val, 'item'): 
                return val.item() 
            return val

        from datetime import datetime
        data = {
            "found": True,
            "rank": get_val('順位'),
            "rank_class": get_val('ランク'),
            "sales_amount": get_val('売上金額'),
            "gross_profit": get_val('粗利金額'),
            "sales_yoy": get_val('前年対比率'),
            "sales_last_year": get_val('前年売上'),
            "profit_last_year": get_val('前年粗利'),
            "sales_2y_ago": get_val('前々年売上'),
            "profit_2y_ago": get_val('前々年粗利'),
            "customer_name": get_val('得意先名称'),
            "updated_at": datetime.now().isoformat()
        }
        return data

    except Exception as e:
        logging.error(f"Error retrieving sales data: {e}")
        raise HTTPException(status_code=500, detail=f"Error retrieving data: {str(e)}")

@app.get("/api/analytics/team-summary")
def get_team_summary(month: str = None):
    """
    複数ファイルから全メンバーの活動状況を集計する。
    month: "YY/MM" 形式 (例: "26/03")。
    """
    logging.info(f"Team Summary triggered for month: {month}")
    if not os.path.exists(EXCEL_DIR):
        raise HTTPException(status_code=500, detail="Excel directory not found")

    # 除外リスト
    EXCLUDE_FILES = [
        "●20260117_2026年度用_日報【原本_2】.xlsm",
        "【電話営業資料】電話　メール話題情報.xlsx",
        "支店999_【津田_操作確認用】_2026年度用日報.xlsm",
        "本社001_【中野次長】_2026年度用日報.xlsm"
    ]

    try:
        # ファイル一覧の取得
        target_files = [
            f for f in os.listdir(EXCEL_DIR) 
            if f.endswith('.xlsm') and f not in EXCLUDE_FILES and not f.startswith('~$')
        ]
        logging.info(f"Scanning {len(target_files)} files for aggregation.")

        def extract_staff_name_py(filename: str) -> str:
            import re
            match = re.search(r'【(.+?)】', filename)
            if not match: return "不明"
            content = match.group(1)
            name_with_paren = re.search(r'^(.+?)（(.+?)）', content)
            if name_with_paren: return name_with_paren.group(1) + name_with_paren.group(2)
            surname = re.search(r'^([^\s\u4e00-\u9fa5]*[\u4e00-\u9fa5]+?)(?:課長|次長|部長|常務|社長|主任|係長|専務|取締役|マネージャー|リーダー|担当|氏)?$', content)
            if surname: return surname.group(1)
            return content[:4]

        summary_results = []

        # 月次ターゲットの正規化 (YY/MM -> 20YY-MM)
        target_month_iso = ""
        if month and '/' in month:
            parts = month.split('/')
            target_month_iso = f"20{parts[0]}-{parts[1]}"
            logging.info(f"Normalized target month to: {target_month_iso}")

        # 柔軟な日付解析ヘルパー
        def parse_dt(x):
            if pd.isna(x): return pd.NaT
            s = str(x).strip()
            if s == '' or s == 'nan' or s == '-': return pd.NaT
            try:
                return pd.to_datetime(s, format='%y/%m/%d')
            except:
                pass
            try:
                return pd.to_datetime(s, format='%Y/%m/%d')
            except:
                pass
            try:
                return pd.to_datetime(s, format='%Y-%m-%d')
            except:
                pass
            try:
                return pd.to_datetime(s, errors='coerce')
            except:
                return pd.NaT

        for filename in target_files:
            try:
                staff_name = extract_staff_name_py(filename)
                df = get_cached_dataframe(filename, '営業日報')
                if df is None or df.empty:
                    logging.debug(f"File {filename}: Empty dataframe.")
                    continue

                # カラムクリーンアップ
                df.columns = [str(col).replace('\n', '').strip() for col in df.columns]
                
                date_col = next((c for c in df.columns if '日付' in c), '日付')
                action_col = next((c for c in df.columns if '行動内容' in c), '行動内容')
                priority_col = next((c for c in df.columns if '重点' in c), '重点顧客')
                area_col = next((c for c in df.columns if 'エリア' in c), 'エリア')

                # 日付パースとフィルタ
                df['dt'] = df[date_col].apply(parse_dt)
                df = df.dropna(subset=['dt'])

                if month:
                    # '26/03' などの文字列と比較するマスク
                    mask1 = df['dt'].dt.strftime('%y/%m') == month
                    mask2 = df['dt'].dt.strftime('%Y-%m') == target_month_iso if target_month_iso else pd.Series([False]*len(df))
                    df = df[mask1 | mask2]

                if df.empty:
                    logging.debug(f"File {filename}: No records found after filtering for {month}.")
                    continue

                # 集計
                df['エリア'] = df[area_col].apply(lambda x: str(x).strip() if pd.notnull(x) and str(x).strip() != '' else '未設定')
                
                def get_category(row):
                    val = row.get(priority_col, '')
                    if pd.notnull(val) and str(val).strip() != '' and str(val).strip() != '-':
                        return '重点'
                    return '一般'
                
                df['区分'] = df.apply(get_category, axis=1)
                df['is_visit'] = df[action_col].apply(lambda x: 1 if pd.notnull(x) and '訪問' in str(x) else 0)
                df['is_call'] = df[action_col].apply(lambda x: 1 if pd.notnull(x) and '電話' in str(x) else 0)

                grouped = df.groupby(['エリア', '区分']).agg({
                    'is_visit': 'sum',
                    'is_call': 'sum'
                }).reset_index()

                for _, row in grouped.iterrows():
                    summary_results.append({
                        "staff": staff_name,
                        "area": row['エリア'],
                        "category": row['区分'],
                        "visits": int(row['is_visit']),
                        "calls": int(row['is_call']),
                        "file": filename
                    })
                
                logging.debug(f"File {filename}: Successfully aggregated {len(grouped)} area/category pairs.")

            except Exception as file_err:
                logging.warning(f"Error aggregating file {filename}: {file_err}")
                continue

        logging.info(f"Aggregation complete. Total records: {len(summary_results)}.")
        return {"records": summary_results, "count": len(target_files)}

    except Exception as e:
        logging.error(f"Error in get_team_summary: {e}")
        import traceback
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=str(e))

@app.get("/api/stats/dashboard")
def get_dashboard_stats(filename: str = DEFAULT_EXCEL_FILE):
    """ダッシュボード統計をバックエンドで集計して返す"""
    # 空レスポンスのテンプレート（エラー時のフォールバック用）
    empty_response = {
        "summary": {"totalReports": 0, "thisMonth": 0, "visits": 0, "calls": 0},
        "priority": {"uniqueCustomers": 0, "visits": 0, "calls": 0},
        "monthly": [], "ranking": [], "updatedAt": datetime.now().isoformat()
    }
    try:
        logging.info(f"--- Dashboard Stats for {filename} ---")
        # Excelファイルが読めない場合（他のユーザーが開いている等）は空レスポンスを返す
        try:
            df = get_cached_dataframe(filename, '営業日報')
        except Exception as read_err:
            logging.warning(f"Dashboard: Cannot read file {filename}: {read_err}")
            return empty_response

        if df is None or df.empty:
            logging.info(f"Dashboard: Empty dataframe for {filename}")
            return empty_response

        logging.info(f"Dashboard: Raw rows={len(df)}")

        # カラム名のクリーンアップ
        df.columns = [str(col).replace('\n', '').strip() for col in df.columns]
        date_col = next((c for c in df.columns if '日付' in c or '年月日' in c), '日付')
        action_col = next((c for c in df.columns if '行動内容' in c), '行動内容')
        priority_col = next((c for c in df.columns if '重点' in c), '重点顧客')
        customer_col = next((c for c in df.columns if '得意先CD' in c), '得意先CD')
        customer_name_col = next((c for c in df.columns if '訪問先名' in c or '得意先名' in c), '訪問先名')
        area_col = next((c for c in df.columns if 'エリア' in c), 'エリア')

        logging.info(f"Dashboard: Cols detected - date={date_col}, action={action_col}")

        # 日付パース（YY/MM/DD テキスト形式に対応）
        def parse_dt(x):
            if pd.isna(x): return pd.NaT
            s = str(x).strip()
            if s == '' or s == 'nan' or s == '-': return pd.NaT
            # 形式1: YY/MM/DD (26/05/13)
            try:
                d = pd.to_datetime(s, format='%y/%m/%d')
                return d
            except:
                pass
            # 形式2: 汎用パース
            try:
                return pd.to_datetime(s, errors='coerce')
            except:
                return pd.NaT

        df['dt'] = df[date_col].apply(parse_dt)
        vdf = df.dropna(subset=['dt']).copy()

        logging.info(f"Dashboard: Valid date rows={len(vdf)}")

        if vdf.empty:
            logging.warning(f"Dashboard: No valid dates. Sample: {df[date_col].head(3).tolist()}")
            return empty_response

        # 基本統計
        total = len(vdf)
        now = datetime.now()
        this_month_mask = (vdf['dt'].dt.year == now.year) & (vdf['dt'].dt.month == now.month)
        tm_cnt = int(this_month_mask.sum())

        # 訪問・電話フラグ
        is_v = vdf[action_col].astype(str).str.contains('訪問', na=False)
        is_c = vdf[action_col].astype(str).str.contains('電話', na=False)
        visits_total = int(is_v.sum())
        calls_total = int(is_c.sum())

        # 重点顧客
        is_p = vdf[priority_col].fillna('').astype(str).apply(lambda x: x.strip() != '' and x.strip() != '-')
        pdf = vdf[is_p]
        p_unique = int(pdf[customer_col].nunique()) if customer_col in pdf.columns else 0
        p_visits = int((pdf[action_col].astype(str).str.contains('訪問', na=False)).sum())
        p_calls = int((pdf[action_col].astype(str).str.contains('電話', na=False)).sum())

        # 月別推移
        vdf['m'] = vdf['dt'].dt.strftime('%y/%m')
        monthly = []
        for m, g in vdf.groupby('m'):
            mv = int(g[action_col].astype(str).str.contains('訪問', na=False).sum())
            mc = int(g[action_col].astype(str).str.contains('電話', na=False).sum())
            g_p = g[priority_col].fillna('').astype(str).apply(lambda x: x.strip() != '' and x.strip() != '-')
            pv = int((g[action_col].astype(str).str.contains('訪問', na=False) & g_p).sum())
            pc = int((g[action_col].astype(str).str.contains('電話', na=False) & g_p).sum())
            # エリア別
            area_stats = []
            if area_col in g.columns:
                ga = g.copy()
                ga[area_col] = ga[area_col].fillna('未設定').astype(str).str.strip()
                ga.loc[ga[area_col] == '', area_col] = '未設定'
                for an, ag in ga.groupby(area_col):
                    av = int(ag[action_col].astype(str).str.contains('訪問', na=False).sum())
                    ac = int(ag[action_col].astype(str).str.contains('電話', na=False).sum())
                    ap = ag[priority_col].fillna('').astype(str).apply(lambda x: x.strip() != '' and x.strip() != '-')
                    area_stats.append({
                        "area": str(an), "visits": av, "calls": ac,
                        "priorityVisits": int((ag[action_col].astype(str).str.contains('訪問', na=False) & ap).sum()),
                        "priorityCalls": int((ag[action_col].astype(str).str.contains('電話', na=False) & ap).sum())
                    })
            monthly.append({
                "month": m, "visits": mv, "calls": mc,
                "priorityVisits": pv, "priorityCalls": pc,
                "areaBreakdown": area_stats
            })
        monthly.sort(key=lambda x: x['month'], reverse=True)

        # 得意先ランキング (Top 10)
        ranking = []
        rank_base = vdf[is_v | is_c].copy()
        if not rank_base.empty and customer_name_col in rank_base.columns:
            rank_base[customer_name_col] = rank_base[customer_name_col].fillna('名称不明').astype(str)
            rank_base['_rv'] = rank_base[action_col].astype(str).str.contains('訪問', na=False)
            rank_base['_rc'] = rank_base[action_col].astype(str).str.contains('電話', na=False)
            rdf = rank_base.groupby(customer_name_col).agg(v=('_rv', 'sum'), c=('_rc', 'sum'))
            rdf['t'] = rdf['v'] + rdf['c']
            rdf = rdf.sort_values('t', ascending=False).head(10)
            for name, row in rdf.iterrows():
                ranking.append({"name": str(name), "visits": int(row['v']), "calls": int(row['c']), "total": int(row['t'])})

        logging.info(f"Dashboard: Total={total}, ThisMonth={tm_cnt}, Visits={visits_total}, Calls={calls_total}")
        return {
            "summary": {"totalReports": total, "thisMonth": tm_cnt, "visits": visits_total, "calls": calls_total},
            "priority": {"uniqueCustomers": p_unique, "visits": p_visits, "calls": p_calls},
            "monthly": monthly, "ranking": ranking,
            "updatedAt": datetime.now().isoformat()
        }
    except Exception as e:
        logging.error(f"Dashboard error: {e}")
        import traceback
        traceback.print_exc()
        # 500エラーを投げずに空レスポンスを返す（フロントエンドのクラッシュを防止）
        return empty_response

@app.get("/api/stats/monthly-summary")
def get_monthly_summary_stats(filename: str = DEFAULT_EXCEL_FILE, month: str = None):
    # month format is 'YY/MM'
    empty_response = {
        "totalReports": 0, "totalVisits": 0, "totalCalls": 0, 
        "priorityVisits": 0, "priorityCalls": 0,
        "totalDesignProposals": 0, "totalDesignCompleted": 0, "totalDesignRejected": 0,
        "uniqueCustomers": 0, "activeDays": 0,
        "areaBreakdown": [], "priorityCustomers": [], "designProgress": [],
        "topCustomers": [], "topCallCustomers": [], "dailyActivity": []
    }
    if not month:
        return empty_response

    try:
        logging.info(f"--- Monthly Summary Stats for {filename}, month={month} ---")
        try:
            df = get_cached_dataframe(filename, '営業日報')
        except Exception as read_err:
            logging.warning(f"MonthlySummary: Cannot read file {filename}: {read_err}")
            return empty_response

        if df is None or df.empty:
            return empty_response

        df.columns = [str(col).replace('\n', '').strip() for col in df.columns]
        date_col = next((c for c in df.columns if '日付' in c or '年月日' in c), '日付')
        action_col = next((c for c in df.columns if '行動内容' in c), '行動内容')
        priority_col = next((c for c in df.columns if '重点' in c), '重点顧客')
        customer_col = next((c for c in df.columns if '得意先CD' in c), '得意先CD')
        customer_name_col = next((c for c in df.columns if '訪問先名' in c or '得意先名' in c), '訪問先名')
        area_col = next((c for c in df.columns if 'エリア' in c), 'エリア')
        dd_code_col = next((c for c in df.columns if '直送先CD' in c), '直送先CD')
        dd_name_col = next((c for c in df.columns if '直送先名' in c), '直送先名')
        rank_col = next((c for c in df.columns if 'ランク' in c), 'ランク')
        design_exist_col = next((c for c in df.columns if 'デザイン提案有無' in c), 'デザイン提案有無')
        design_status_col = next((c for c in df.columns if 'デザイン進捗状況' in c), 'デザイン進捗状況')

        def parse_dt(x):
            if pd.isna(x): return pd.NaT
            s = str(x).strip()
            if s == '' or s == 'nan' or s == '-': return pd.NaT
            try:
                d = pd.to_datetime(s, format='%y/%m/%d')
                return d
            except: pass
            try:
                return pd.to_datetime(s, errors='coerce')
            except: return pd.NaT

        df['dt'] = df[date_col].apply(parse_dt)
        vdf = df.dropna(subset=['dt']).copy()
        if vdf.empty:
            return empty_response

        vdf['m'] = vdf['dt'].dt.strftime('%y/%m')
        mdf = vdf[vdf['m'] == month].copy()
        if mdf.empty:
            return empty_response

        mdf['is_v'] = mdf[action_col].astype(str).str.contains('訪問', na=False)
        mdf['is_c'] = mdf[action_col].astype(str).str.contains('電話', na=False)
        mdf['is_p'] = mdf[priority_col].fillna('').astype(str).apply(lambda x: x.strip() != '' and x.strip() != '-')
        
        has_design = mdf[design_exist_col].astype(str).apply(lambda x: x.strip() in ['有', 'あり']) if design_exist_col in mdf.columns else pd.Series(False, index=mdf.index)
        is_design_new = mdf[design_status_col].astype(str).str.strip() == '新規' if design_status_col in mdf.columns else pd.Series(False, index=mdf.index)
        is_design_proposal = has_design | is_design_new
        
        is_design_completed = mdf[design_status_col].astype(str).str.contains('出稿', na=False) if design_status_col in mdf.columns else pd.Series(False, index=mdf.index)
        is_design_rejected = mdf[design_status_col].astype(str).str.contains('不採用', na=False) if design_status_col in mdf.columns else pd.Series(False, index=mdf.index)

        totalVisits = int(mdf['is_v'].sum())
        totalCalls = int(mdf['is_c'].sum())
        priorityVisits = int((mdf['is_v'] & mdf['is_p']).sum())
        priorityCalls = int((mdf['is_c'] & mdf['is_p']).sum())
        totalDesignProposals = int(is_design_proposal.sum())
        totalDesignCompleted = int(is_design_completed.sum())
        totalDesignRejected = int(is_design_rejected.sum())

        uniqueCustomers = int(mdf[mdf[customer_name_col].astype(str).str.strip() != ''][customer_name_col].nunique()) if customer_name_col in mdf.columns else 0
        activeDays = int(mdf['dt'].nunique())

        areaBreakdown = []
        if area_col in mdf.columns:
            adf = mdf[mdf['is_v'] | mdf['is_c']].copy()
            adf[area_col] = adf[area_col].fillna('未設定').astype(str).str.strip()
            adf.loc[adf[area_col] == '', area_col] = '未設定'
            for an, ag in adf.groupby(area_col):
                areaBreakdown.append({
                    "area": str(an),
                    "visits": int(ag['is_v'].sum()),
                    "calls": int(ag['is_c'].sum()),
                    "priorityVisits": int((ag['is_v'] & ag['is_p']).sum()),
                    "priorityCalls": int((ag['is_c'] & ag['is_p']).sum()),
                    "designProposals": int(is_design_proposal[ag.index].sum()),
                })
            areaBreakdown.sort(key=lambda x: (1 if x['area'] == '未設定' else 0, -(x['visits'] + x['calls'])))

        priorityCustomers = []
        pdf = mdf[mdf['is_p']].copy()
        if not pdf.empty and customer_col in pdf.columns:
            for ccode, cg in pdf.groupby(customer_col):
                p_cname = cg[customer_name_col].iloc[0] if customer_name_col in cg.columns else '不明'
                p_area = cg[area_col].iloc[0] if area_col in cg.columns else ''
                p_rank = cg[rank_col].iloc[0] if rank_col in cg.columns else ''
                
                dd_list = []
                if dd_code_col in cg.columns:
                    for ddcode, ddg in cg[cg[dd_code_col].notna() & (cg[dd_code_col].astype(str).str.strip() != '')].groupby(dd_code_col):
                        dd_cname = ddg[dd_name_col].iloc[0] if dd_name_col in ddg.columns else ''
                        dd_area = ddg[area_col].iloc[0] if area_col in ddg.columns else ''
                        dd_rank = ddg[rank_col].iloc[0] if rank_col in ddg.columns else ''
                        dd_visits = int(ddg['is_v'].sum())
                        dd_calls = int(ddg['is_c'].sum())
                        
                        if dd_visits > 0 or dd_calls > 0 or int(is_design_proposal[ddg.index].sum()) > 0:
                            dd_list.append({
                                "code": str(ddcode).replace('.0', '').strip(),
                                "name": str(dd_cname),
                                "visits": dd_visits,
                                "calls": dd_calls,
                                "designProposals": int(is_design_proposal[ddg.index].sum()),
                                "lastDate": ddg['dt'].max().strftime('%Y/%m/%d') if not pd.isna(ddg['dt'].max()) else '',
                                "area": str(dd_area),
                                "rank": str(dd_rank),
                                "isPriority": True
                            })

                c_visits = int(cg['is_v'].sum())
                c_calls = int(cg['is_c'].sum())
                priorityCustomers.append({
                    "code": str(ccode),
                    "name": str(p_cname),
                    "visits": c_visits,
                    "calls": c_calls,
                    "designProposals": int(is_design_proposal[cg.index].sum()),
                    "total": c_visits + c_calls,
                    "lastDate": cg['dt'].max().strftime('%Y/%m/%d') if not pd.isna(cg['dt'].max()) else '',
                    "area": str(p_area),
                    "rank": str(p_rank),
                    "isPriority": True,
                    "directDeliveries": dd_list
                })
            priorityCustomers.sort(key=lambda x: x['total'], reverse=True)

        designProgress = []
        if design_status_col in mdf.columns:
            dsg_df = mdf[has_design].copy()
            if not dsg_df.empty:
                dsg_df[design_status_col] = dsg_df[design_status_col].fillna('未設定').astype(str).str.strip()
                dsg_df.loc[dsg_df[design_status_col] == '', design_status_col] = '未設定'
                for st, sg in dsg_df.groupby(design_status_col):
                    designProgress.append({"status": str(st), "count": len(sg)})
            designProgress.sort(key=lambda x: x['count'], reverse=True)

        def get_top_customers(action_mask):
            top = []
            tdf = mdf[action_mask & mdf[customer_name_col].notna()].copy()
            if not tdf.empty and customer_name_col in tdf.columns:
                tdf[customer_name_col] = tdf[customer_name_col].astype(str).str.strip()
                tdf = tdf[tdf[customer_name_col] != '']
                for cname, cg in tdf.groupby(customer_name_col):
                    details = []
                    if dd_name_col in cg.columns:
                        cg_dd = cg.copy()
                        cg_dd[dd_name_col] = cg_dd[dd_name_col].fillna('(直接)').astype(str).str.strip()
                        cg_dd.loc[cg_dd[dd_name_col] == '', dd_name_col] = '(直接)'
                        for ddn, ddg in cg_dd.groupby(dd_name_col):
                            details.append({"name": str(ddn), "count": len(ddg)})
                        details.sort(key=lambda x: x['count'], reverse=True)
                    top.append({"name": str(cname), "count": len(cg), "details": details})
                top.sort(key=lambda x: x['count'], reverse=True)
            return top[:10]

        topCustomers = get_top_customers(mdf['is_v'])
        topCallCustomers = get_top_customers(mdf['is_c'])

        dailyActivity = []
        mdf['date_str'] = mdf['dt'].dt.strftime('%Y/%m/%d')
        for d, dg in mdf.groupby('date_str'):
            dailyActivity.append({
                "date": str(d),
                "visits": int(dg['is_v'].sum()),
                "calls": int(dg['is_c'].sum())
            })
        dailyActivity.sort(key=lambda x: x['date'])

        return {
            "totalReports": len(mdf),
            "totalVisits": totalVisits,
            "totalCalls": totalCalls,
            "priorityVisits": priorityVisits,
            "priorityCalls": priorityCalls,
            "totalDesignProposals": totalDesignProposals,
            "totalDesignCompleted": totalDesignCompleted,
            "totalDesignRejected": totalDesignRejected,
            "uniqueCustomers": uniqueCustomers,
            "activeDays": activeDays,
            "areaBreakdown": areaBreakdown,
            "priorityCustomers": priorityCustomers,
            "designProgress": designProgress,
            "topCustomers": topCustomers,
            "topCallCustomers": topCallCustomers,
            "dailyActivity": dailyActivity
        }
    except Exception as e:
        logging.error(f"Monthly summary error: {e}")
        import traceback
        traceback.print_exc()
        return empty_response

@app.get("/api/analytics/points-table")
def get_points_table(target_months_count: int = 7):
    """
    全メンバーの日報点数表を集計する。
    """
    if not os.path.exists(EXCEL_DIR):
        raise HTTPException(status_code=500, detail="Excel directory not found")

    EXCLUDE_FILES = [
        "●20260117_2026年度用_日報【原本_2】.xlsm",
        "【電話営業資料】電話　メール話題情報.xlsx",
        "支店999_【津田_操作確認用】_2026年度用日報.xlsm",
        "本社001_【中野次長】_2026年度用日報.xlsm"
    ]

    try:
        target_files = [
            f for f in os.listdir(EXCEL_DIR) 
            if f.endswith('.xlsm') and f not in EXCLUDE_FILES and not f.startswith('~$')
        ]
        
        # 2026年度の月リスト（2月〜翌1月）
        month_keys = [
            ("26/02", "2月"), ("26/03", "3月"), ("26/04", "4月"), ("26/05", "5月"),
            ("26/06", "6月"), ("26/07", "7月"), ("26/08", "8月"), ("26/09", "9月"),
            ("26/10", "10月"), ("26/11", "11月"), ("26/12", "12月"), ("27/01", "1月")
        ]

        def extract_staff_name_py(filename: str) -> str:
            import re
            match = re.search(r'【(.+?)】', filename)
            if not match: return filename.replace('.xlsm', '')
            content = match.group(1)
            name_with_paren = re.search(r'^(.+?)（(.+?)）', content)
            if name_with_paren: return name_with_paren.group(1)
            surname = re.search(r'^([^\s\u4e00-\u9fa5]*[\u4e00-\u9fa5]+?)(?:課長|次長|部長|常務|社長|主任|係長|専務|取締役|マネージャー|リーダー|担当|氏)?$', content)
            if surname: return surname.group(1)
            return content[:4]

        # 柔軟な日付解析ヘルパー
        def parse_dt(x):
            if pd.isna(x): return pd.NaT
            s = str(x).strip()
            if s == '' or s == 'nan' or s == '-': return pd.NaT
            try:
                return pd.to_datetime(s, format='%y/%m/%d')
            except:
                pass
            try:
                return pd.to_datetime(s, format='%Y/%m/%d')
            except:
                pass
            try:
                return pd.to_datetime(s, format='%Y-%m-%d')
            except:
                pass
            try:
                return pd.to_datetime(s, errors='coerce')
            except:
                return pd.NaT

        records = []

        for filename in target_files:
            try:
                staff_name = extract_staff_name_py(filename)
                
                # 1. 重点件数（得意先_Listから個人の重点顧客の数を数える）
                priority_count = 0
                try:
                    cust_df = get_cached_dataframe(filename, '得意先_List')
                    if cust_df is not None and not cust_df.empty:
                        cust_df.columns = [str(col).replace('\n', '').strip() for col in cust_df.columns]
                        priority_col = next((c for c in cust_df.columns if '重点' in c), None)
                        rep_col = cust_df.columns[8] if len(cust_df.columns) > 8 else None
                        
                        if priority_col and rep_col:
                            is_priority = cust_df[priority_col].astype(str).str.contains('重点', na=False)
                            
                            def match_representative(rep_val, target_name):
                                if pd.isna(rep_val): return False
                                rep_val = str(rep_val).strip()
                                target_name = str(target_name).strip()
                                if rep_val == target_name: return True
                                
                                import re
                                def clean_name(n):
                                    n = re.sub(r'[\(（].*?[\)）]', '', n)
                                    n = re.sub(r'(?:次長|課長|部長|常務|専務|社長|主任|係長|取締役|マネージャー|マネ|リーダー|担当)$', '', n)
                                    return n.strip()
                                    
                                cleaned_rep = clean_name(rep_val)
                                cleaned_staff = clean_name(target_name)
                                if cleaned_rep == cleaned_staff: return True
                                
                                match = re.search(r'(.+?)[（\(](.+?)[）\)]', rep_val)
                                if match:
                                    base = match.group(1).strip()
                                    inside = match.group(2).strip()
                                    combined = base + inside
                                    if clean_name(combined) == cleaned_staff: return True
                                return False
                                
                            matches_rep = cust_df[rep_col].apply(lambda x: match_representative(x, staff_name))
                            priority_count = int(cust_df[is_priority & matches_rep].shape[0])
                        elif priority_col:
                            # フォールバック: 担当者列が取れなければ全体の重点顧客数をカウント
                            priority_count = int(cust_df[cust_df[priority_col].astype(str).str.contains('重点', na=False)].shape[0])
                except Exception as cust_err:
                    logging.warning(f"Error reading priority count for {filename}: {cust_err}")

                # 2. 日報データの読み込み
                df = get_cached_dataframe(filename, '営業日報')
                if df is None or df.empty:
                    # 空白レコードを追加
                    monthly_data = {label: {"priority_calls": 0, "general_calls": 0, "priority_visits": 0, "general_visits": 0} for _, label in month_keys}
                    records.append({
                        "staff": staff_name,
                        "priority_count": priority_count,
                        "monthly_data": monthly_data,
                        "totals": {"priority_calls": 0, "general_calls": 0, "total_calls": 0, "priority_visits": 0, "general_visits": 0, "total_visits": 0},
                        "points": 0.0,
                        "achievement_rate": 0.0,
                        "rating": 0.0,
                        "file": filename
                    })
                    continue

                # カラムクリーンアップ
                df.columns = [str(col).replace('\n', '').strip() for col in df.columns]
                
                date_col = next((c for c in df.columns if '日付' in c), '日付')
                action_col = next((c for c in df.columns if '行動内容' in c), '行動内容')
                priority_col = next((c for c in df.columns if '重点' in c), '重点顧客')

                df['dt'] = df[date_col].apply(parse_dt)
                df = df.dropna(subset=['dt'])

                # 各項目のフラグ付け
                def is_priority(val):
                    if pd.isnull(val): return False
                    s = str(val).strip()
                    return s != '' and s != '-' and '重点' in s

                df['is_priority_cust'] = df[priority_col].apply(is_priority) if priority_col in df.columns else False
                df['is_visit'] = df[action_col].apply(lambda x: True if pd.notnull(x) and '訪問' in str(x) else False)
                df['is_call'] = df[action_col].apply(lambda x: True if pd.notnull(x) and '電話' in str(x) else False)

                # 月ごとの集計
                monthly_data = {}
                tot_p_calls = 0
                tot_g_calls = 0
                tot_p_visits = 0
                tot_g_visits = 0

                for m_key, m_label in month_keys:
                    parts = m_key.split('/')
                    iso_month = f"20{parts[0]}-{parts[1]}"
                    mask1 = df['dt'].dt.strftime('%y/%m') == m_key
                    mask2 = df['dt'].dt.strftime('%Y-%m') == iso_month
                    m_df = df[mask1 | mask2]

                    if m_df.empty:
                        monthly_data[m_label] = {
                            "priority_calls": 0,
                            "general_calls": 0,
                            "priority_visits": 0,
                            "general_visits": 0
                        }
                        continue

                    p_calls = int((m_df['is_call'] & m_df['is_priority_cust']).sum())
                    g_calls = int((m_df['is_call'] & ~m_df['is_priority_cust']).sum())
                    p_visits = int((m_df['is_visit'] & m_df['is_priority_cust']).sum())
                    g_visits = int((m_df['is_visit'] & ~m_df['is_priority_cust']).sum())

                    monthly_data[m_label] = {
                        "priority_calls": p_calls,
                        "general_calls": g_calls,
                        "priority_visits": p_visits,
                        "general_visits": g_visits
                    }

                    tot_p_calls += p_calls
                    tot_g_calls += g_calls
                    tot_p_visits += p_visits
                    tot_g_visits += g_visits

                # 点数計算: 重点電話 + 一般電話/2 + 重点訪問*10 + 一般訪問*3
                points = float(tot_p_calls) + (float(tot_g_calls) / 2.0) + (float(tot_p_visits) * 10.0) + (float(tot_g_visits) * 3.0)
                
                # 目標値 (月200点 * target_months_count)
                target_score = float(200 * target_months_count)
                achievement_rate = (points / target_score) * 100.0 if target_score > 0 else 0.0
                rating = points / 140.0

                records.append({
                    "staff": staff_name,
                    "priority_count": priority_count,
                    "monthly_data": monthly_data,
                    "totals": {
                        "priority_calls": tot_p_calls,
                        "general_calls": tot_g_calls,
                        "total_calls": tot_p_calls + tot_g_calls,
                        "priority_visits": tot_p_visits,
                        "general_visits": tot_g_visits,
                        "total_visits": tot_p_visits + tot_g_visits
                    },
                    "points": round(points, 1),
                    "achievement_rate": round(achievement_rate, 1),
                    "rating": round(rating, 1),
                    "file": filename
                })

            except Exception as file_err:
                logging.warning(f"Error processing points table for {filename}: {file_err}")
                continue

        return {
            "months": [m_label for _, m_label in month_keys],
            "records": records,
            "target_months_count": target_months_count
        }

    except Exception as e:
        logging.error(f"Error generating points table: {e}")
        raise HTTPException(status_code=500, detail=str(e))

@app.get("/api/analytics/export/points-table")
def export_points_table_excel(target_months_count: int = 7):
    """
    日報点数表の美しいExcelファイルをエクスポートする。
    """
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    import io
    from fastapi.responses import StreamingResponse

    # 1. データの取得
    try:
        data = get_points_table(target_months_count=target_months_count)
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Failed to gather data: {e}")

    records = data["records"]
    months = data["months"]

    # 実績がある月のみ抽出
    active_months = []
    for m in months:
        has_act = False
        for r in records:
            act = r["monthly_data"].get(m, {})
            if (act.get("priority_calls", 0) > 0 or 
                act.get("general_calls", 0) > 0 or 
                act.get("priority_visits", 0) > 0 or 
                act.get("general_visits", 0) > 0):
                has_act = True
                break
        if has_act:
            active_months.append(m)

    if not active_months:
        active_months = ["2月", "3月", "4月", "5月"]

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "日報点数表"
    ws.views.sheetView[0].showGridLines = True

    # 印刷設定: A4横1枚に収める
    ws.page_setup.orientation = "landscape"
    ws.page_setup.paperSize = 9 # 9 = A4
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 1

    # 余白設定 (狭めにして印刷領域を最大化)
    ws.page_margins.left = 0.25
    ws.page_margins.right = 0.25
    ws.page_margins.top = 0.3
    ws.page_margins.bottom = 0.3
    ws.page_margins.header = 0.1
    ws.page_margins.footer = 0.1

    # スタイル定義
    font_title = Font(name="Meiryo UI", size=16, bold=True)
    font_header = Font(name="Meiryo UI", size=9, bold=True)
    font_sub_header = Font(name="Meiryo UI", size=8, bold=True)
    font_data = Font(name="Meiryo UI", size=9)
    font_bold = Font(name="Meiryo UI", size=9, bold=True)

    fill_blue_header = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    fill_gray_header = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    fill_sub_blue = PatternFill(start_color="F2F5F9", end_color="F2F5F9", fill_type="solid")
    fill_yellow = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
    fill_light_blue = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")
    fill_green = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")

    border_thin_side = Side(border_style="thin", color="D9D9D9")
    border_thin = Border(left=border_thin_side, right=border_thin_side, top=border_thin_side, bottom=border_thin_side)
    border_double_bottom = Border(top=Side(border_style="thin", color="000000"), bottom=Side(border_style="double", color="000000"))

    align_center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    align_left = Alignment(horizontal="left", vertical="center")
    align_right = Alignment(horizontal="right", vertical="center")

    # タイトル行
    ws.merge_cells("A1:C1")
    ws["A1"] = "日報点数表 2026"
    ws["A1"].font = font_title
    ws["A1"].alignment = align_left
    ws.row_dimensions[1].height = 30

    # ヘッダー作成
    ws.row_dimensions[3].height = 24
    ws.row_dimensions[4].height = 20

    ws.merge_cells("A3:A4")
    ws["A3"] = "営業名"
    ws["A3"].font = font_header
    ws["A3"].fill = fill_gray_header
    ws["A3"].alignment = align_center
    ws["A3"].border = border_thin
    ws["A4"].border = border_thin

    ws.merge_cells("B3:B4")
    ws["B3"] = "重点件数"
    ws["B3"].font = font_header
    ws["B3"].fill = fill_gray_header
    ws["B3"].alignment = align_center
    ws["B3"].border = border_thin
    ws["B4"].border = border_thin

    current_col = 3

    # 各月のヘッダー
    for m in active_months:
        col_letter_start = get_column_letter(current_col)
        col_letter_end = get_column_letter(current_col + 3)
        ws.merge_cells(f"{col_letter_start}3:{col_letter_end}3")
        cell_ref = f"{col_letter_start}3"
        ws[cell_ref] = m
        ws[cell_ref].font = font_header
        ws[cell_ref].fill = fill_blue_header
        ws[cell_ref].alignment = align_center

        sub_headers = ["重点電話", "電話総数", "重点訪問", "訪問件数"]
        for idx, sub in enumerate(sub_headers):
            c_ref = f"{get_column_letter(current_col + idx)}4"
            ws[c_ref] = sub
            ws[c_ref].font = font_sub_header
            ws[c_ref].fill = fill_sub_blue
            ws[c_ref].alignment = align_center
            ws[c_ref].border = border_thin
            ws[f"{get_column_letter(current_col + idx)}3"].border = border_thin
            
        current_col += 4

    # 計ヘッダー
    tot_col_start = current_col
    ws.merge_cells(f"{get_column_letter(current_col)}3:{get_column_letter(current_col + 5)}3")
    cell_ref = f"{get_column_letter(current_col)}3"
    ws[cell_ref] = "計"
    ws[cell_ref].font = font_header
    ws[cell_ref].fill = fill_gray_header
    ws[cell_ref].alignment = align_center

    sub_tots = ["重点電話総数", "電話総件数", "総電話件数", "重点訪問総数", "訪問総件数", "総訪問件数"]
    for idx, sub in enumerate(sub_tots):
        c_ref = f"{get_column_letter(current_col + idx)}4"
        ws[c_ref] = sub
        ws[c_ref].font = font_sub_header
        ws[c_ref].fill = fill_gray_header
        ws[c_ref].alignment = align_center
        ws[c_ref].border = border_thin
        ws[f"{get_column_letter(current_col + idx)}3"].border = border_thin
        
    current_col += 6

    # 各種得点ヘッダー
    points_col_idx = current_col
    ws.merge_cells(f"{get_column_letter(current_col)}3:{get_column_letter(current_col)}4")
    ws[f"{get_column_letter(current_col)}3"] = "点数"
    ws[f"{get_column_letter(current_col)}3"].font = font_header
    ws[f"{get_column_letter(current_col)}3"].fill = fill_yellow
    ws[f"{get_column_letter(current_col)}3"].alignment = align_center
    ws[f"{get_column_letter(current_col)}3"].border = border_thin
    ws[f"{get_column_letter(current_col)}4"].border = border_thin
    current_col += 1

    ach_col_idx = current_col
    ws.merge_cells(f"{get_column_letter(current_col)}3:{get_column_letter(current_col)}4")
    ws[f"{get_column_letter(current_col)}3"] = f"月200点×{target_months_count}\n({200 * target_months_count})"
    ws[f"{get_column_letter(current_col)}3"].font = font_header
    ws[f"{get_column_letter(current_col)}3"].fill = fill_light_blue
    ws[f"{get_column_letter(current_col)}3"].alignment = align_center
    ws[f"{get_column_letter(current_col)}3"].border = border_thin
    ws[f"{get_column_letter(current_col)}4"].border = border_thin
    current_col += 1

    rat_col_idx = current_col
    ws.merge_cells(f"{get_column_letter(current_col)}3:{get_column_letter(current_col)}4")
    ws[f"{get_column_letter(current_col)}3"] = "評価点"
    ws[f"{get_column_letter(current_col)}3"].font = font_header
    ws[f"{get_column_letter(current_col)}3"].fill = fill_green
    ws[f"{get_column_letter(current_col)}3"].alignment = align_center
    ws[f"{get_column_letter(current_col)}3"].border = border_thin
    ws[f"{get_column_letter(current_col)}4"].border = border_thin

    # データ行書き込み
    row_num = 5
    for r in records:
        ws.row_dimensions[row_num].height = 18

        ws[f"A{row_num}"] = r["staff"]
        ws[f"A{row_num}"].font = font_bold
        ws[f"A{row_num}"].alignment = align_left
        ws[f"A{row_num}"].border = border_thin

        ws[f"B{row_num}"] = r["priority_count"]
        ws[f"B{row_num}"].font = font_data
        ws[f"B{row_num}"].alignment = align_center
        ws[f"B{row_num}"].border = border_thin

        col_idx = 3
        for m in active_months:
            act = r["monthly_data"].get(m, {"priority_calls": 0, "general_calls": 0, "priority_visits": 0, "general_visits": 0})
            for key in ["priority_calls", "general_calls", "priority_visits", "general_visits"]:
                cell_ref = f"{get_column_letter(col_idx)}{row_num}"
                ws[cell_ref] = act.get(key, 0)
                ws[cell_ref].font = font_data
                ws[cell_ref].alignment = align_right
                ws[cell_ref].border = border_thin
                col_idx += 1

        # 各種集計用計算式（数式によるリアルタイム計算）
        # 1. 重点電話総数
        p_calls_parts = [f"{get_column_letter(3 + m_idx * 4)}{row_num}" for m_idx in range(len(active_months))]
        ws[f"{get_column_letter(tot_col_start)}{row_num}"] = f"={'+'.join(p_calls_parts)}"
        ws[f"{get_column_letter(tot_col_start)}{row_num}"].font = font_bold
        ws[f"{get_column_letter(tot_col_start)}{row_num}"].alignment = align_right
        ws[f"{get_column_letter(tot_col_start)}{row_num}"].border = border_thin

        # 2. 電話総件数
        g_calls_parts = [f"{get_column_letter(4 + m_idx * 4)}{row_num}" for m_idx in range(len(active_months))]
        ws[f"{get_column_letter(tot_col_start + 1)}{row_num}"] = f"={'+'.join(g_calls_parts)}"
        ws[f"{get_column_letter(tot_col_start + 1)}{row_num}"].font = font_bold
        ws[f"{get_column_letter(tot_col_start + 1)}{row_num}"].alignment = align_right
        ws[f"{get_column_letter(tot_col_start + 1)}{row_num}"].border = border_thin

        # 3. 総電話件数 (重点＋一般)
        p_c_col = get_column_letter(tot_col_start)
        g_c_col = get_column_letter(tot_col_start + 1)
        ws[f"{get_column_letter(tot_col_start + 2)}{row_num}"] = f"={p_c_col}{row_num}+{g_c_col}{row_num}"
        ws[f"{get_column_letter(tot_col_start + 2)}{row_num}"].font = font_bold
        ws[f"{get_column_letter(tot_col_start + 2)}{row_num}"].alignment = align_right
        ws[f"{get_column_letter(tot_col_start + 2)}{row_num}"].border = border_thin

        # 4. 重点訪問総数
        p_visits_parts = [f"{get_column_letter(5 + m_idx * 4)}{row_num}" for m_idx in range(len(active_months))]
        ws[f"{get_column_letter(tot_col_start + 3)}{row_num}"] = f"={'+'.join(p_visits_parts)}"
        ws[f"{get_column_letter(tot_col_start + 3)}{row_num}"].font = font_bold
        ws[f"{get_column_letter(tot_col_start + 3)}{row_num}"].alignment = align_right
        ws[f"{get_column_letter(tot_col_start + 3)}{row_num}"].border = border_thin

        # 5. 訪問総件数
        g_visits_parts = [f"{get_column_letter(6 + m_idx * 4)}{row_num}" for m_idx in range(len(active_months))]
        ws[f"{get_column_letter(tot_col_start + 4)}{row_num}"] = f"={'+'.join(g_visits_parts)}"
        ws[f"{get_column_letter(tot_col_start + 4)}{row_num}"].font = font_bold
        ws[f"{get_column_letter(tot_col_start + 4)}{row_num}"].alignment = align_right
        ws[f"{get_column_letter(tot_col_start + 4)}{row_num}"].border = border_thin

        # 6. 総訪問件数 (重点＋一般)
        p_v_col = get_column_letter(tot_col_start + 3)
        g_v_col = get_column_letter(tot_col_start + 4)
        ws[f"{get_column_letter(tot_col_start + 5)}{row_num}"] = f"={p_v_col}{row_num}+{g_v_col}{row_num}"
        ws[f"{get_column_letter(tot_col_start + 5)}{row_num}"].font = font_bold
        ws[f"{get_column_letter(tot_col_start + 5)}{row_num}"].alignment = align_right
        ws[f"{get_column_letter(tot_col_start + 5)}{row_num}"].border = border_thin

        #活動得点計算式: 重点電話×1 ＋ 一般電話/2 ＋ 重点訪問×10 ＋ 一般訪問×3
        points_formula = f"={p_c_col}{row_num}+({g_c_col}{row_num}/2)+({p_v_col}{row_num}*10)+({g_v_col}{row_num}*3)"
        ws[f"{get_column_letter(points_col_idx)}{row_num}"] = points_formula
        ws[f"{get_column_letter(points_col_idx)}{row_num}"].font = font_bold
        ws[f"{get_column_letter(points_col_idx)}{row_num}"].alignment = align_right
        ws[f"{get_column_letter(points_col_idx)}{row_num}"].border = border_thin
        ws[f"{get_column_letter(points_col_idx)}{row_num}"].fill = fill_yellow
        ws[f"{get_column_letter(points_col_idx)}{row_num}"].number_format = "0.0"

        # 達成率: 総合得点 ÷ (月200点×対象月数)
        target_score = 200 * target_months_count
        pts_col_let = get_column_letter(points_col_idx)
        ach_formula = f"={pts_col_let}{row_num}/{target_score}"
        ws[f"{get_column_letter(ach_col_idx)}{row_num}"] = ach_formula
        ws[f"{get_column_letter(ach_col_idx)}{row_num}"].font = font_bold
        ws[f"{get_column_letter(ach_col_idx)}{row_num}"].alignment = align_right
        ws[f"{get_column_letter(ach_col_idx)}{row_num}"].border = border_thin
        ws[f"{get_column_letter(ach_col_idx)}{row_num}"].fill = fill_light_blue
        ws[f"{get_column_letter(ach_col_idx)}{row_num}"].number_format = "0.0%"

        # 評価点: 総合得点 ÷ 140.0
        rat_formula = f"={pts_col_let}{row_num}/140.0"
        ws[f"{get_column_letter(rat_col_idx)}{row_num}"] = rat_formula
        ws[f"{get_column_letter(rat_col_idx)}{row_num}"].font = font_bold
        ws[f"{get_column_letter(rat_col_idx)}{row_num}"].alignment = align_right
        ws[f"{get_column_letter(rat_col_idx)}{row_num}"].border = border_thin
        ws[f"{get_column_letter(rat_col_idx)}{row_num}"].fill = fill_green
        ws[f"{get_column_letter(rat_col_idx)}{row_num}"].number_format = "0.0"

        row_num += 1

    # 「計」行の作成
    ws.row_dimensions[row_num].height = 20
    ws[f"A{row_num}"] = "計"
    ws[f"A{row_num}"].font = font_bold
    ws[f"A{row_num}"].alignment = align_left
    ws[f"A{row_num}"].border = border_thin

    ws[f"B{row_num}"] = f"=SUM(B5:B{row_num - 1})"
    ws[f"B{row_num}"].font = font_bold
    ws[f"B{row_num}"].alignment = align_center
    ws[f"B{row_num}"].border = border_thin

    for c_idx in range(3, rat_col_idx + 1):
        col_let = get_column_letter(c_idx)
        if c_idx == ach_col_idx:
            tot_pts_cell = f"{get_column_letter(points_col_idx)}{row_num}"
            ws[f"{col_let}{row_num}"] = f"={tot_pts_cell}/{200 * target_months_count}"
            ws[f"{col_let}{row_num}"].number_format = "0.0%"
            ws[f"{col_let}{row_num}"].fill = fill_light_blue
        elif c_idx == rat_col_idx:
            tot_pts_cell = f"{get_column_letter(points_col_idx)}{row_num}"
            ws[f"{col_let}{row_num}"] = f"={tot_pts_cell}/140.0"
            ws[f"{col_let}{row_num}"].number_format = "0.0"
            ws[f"{col_let}{row_num}"].fill = fill_green
        elif c_idx == points_col_idx:
            ws[f"{col_let}{row_num}"] = f"=SUM({col_let}5:{col_let}{row_num - 1})"
            ws[f"{col_let}{row_num}"].number_format = "0.0"
            ws[f"{col_let}{row_num}"].fill = fill_yellow
        else:
            ws[f"{col_let}{row_num}"] = f"=SUM({col_let}5:{col_let}{row_num - 1})"

        ws[f"{col_let}{row_num}"].font = font_bold
        ws[f"{col_let}{row_num}"].alignment = align_right
        ws[f"{col_let}{row_num}"].border = border_thin

    row_num += 1

    # 「平均点」行の作成
    ws.row_dimensions[row_num].height = 20
    ws[f"A{row_num}"] = "平均点"
    ws[f"A{row_num}"].font = font_bold
    ws[f"A{row_num}"].alignment = align_left
    ws[f"A{row_num}"].border = border_double_bottom

    ws[f"B{row_num}"] = f"=AVERAGE(B5:B{row_num - 2})"
    ws[f"B{row_num}"].font = font_bold
    ws[f"B{row_num}"].alignment = align_center
    ws[f"B{row_num}"].border = border_double_bottom
    ws[f"B{row_num}"].number_format = "0.0"

    for c_idx in range(3, rat_col_idx + 1):
        col_let = get_column_letter(c_idx)
        if c_idx == ach_col_idx:
            avg_pts_cell = f"{get_column_letter(points_col_idx)}{row_num}"
            ws[f"{col_let}{row_num}"] = f"={avg_pts_cell}/{200 * target_months_count}"
            ws[f"{col_let}{row_num}"].number_format = "0.0%"
            ws[f"{col_let}{row_num}"].fill = fill_light_blue
        elif c_idx == rat_col_idx:
            avg_pts_cell = f"{get_column_letter(points_col_idx)}{row_num}"
            ws[f"{col_let}{row_num}"] = f"={avg_pts_cell}/140.0"
            ws[f"{col_let}{row_num}"].number_format = "0.0"
            ws[f"{col_let}{row_num}"].fill = fill_green
        elif c_idx == points_col_idx:
            ws[f"{col_let}{row_num}"] = f"=AVERAGE({col_let}5:{col_let}{row_num - 2})"
            ws[f"{col_let}{row_num}"].number_format = "0.0"
            ws[f"{col_let}{row_num}"].fill = fill_yellow
        else:
            ws[f"{col_let}{row_num}"] = f"=AVERAGE({col_let}5:{col_let}{row_num - 2})"
            ws[f"{col_let}{row_num}"].number_format = "0.0"

        ws[f"{col_let}{row_num}"].font = font_bold
        ws[f"{col_let}{row_num}"].alignment = align_right
        ws[f"{col_let}{row_num}"].border = border_double_bottom

    # 列幅の自動調整
    ws.column_dimensions["A"].width = 15
    ws.column_dimensions["B"].width = 10
    for c_idx in range(3, points_col_idx):
        ws.column_dimensions[get_column_letter(c_idx)].width = 11
    ws.column_dimensions[get_column_letter(points_col_idx)].width = 11
    ws.column_dimensions[get_column_letter(ach_col_idx)].width = 18
    ws.column_dimensions[get_column_letter(rat_col_idx)].width = 10

    # ストリーミングレスポンスで返却
    file_stream = io.BytesIO()
    wb.save(file_stream)
    file_stream.seek(0)

    headers_dict = {
        'Content-Disposition': f'attachment; filename="DailyReportPointsTable_{target_months_count}months.xlsx"'
    }
    return StreamingResponse(file_stream, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", headers=headers_dict)

@app.get("/api/analytics/export/team-summary")
def export_team_summary_excel(month: str = None):
    """
    活動集計の美しいExcelファイルをエクスポートする。
    """
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    import io
    from fastapi.responses import StreamingResponse

    try:
        data = get_team_summary(month=month)
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Failed to gather data: {e}")

    records = data["records"]

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "活動集計"
    ws.views.sheetView[0].showGridLines = True

    # 印刷設定: A4横1枚に収める
    ws.page_setup.orientation = "landscape"
    ws.page_setup.paperSize = 9 # 9 = A4
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 1

    # 余白設定 (狭めにして印刷領域を最大化)
    ws.page_margins.left = 0.25
    ws.page_margins.right = 0.25
    ws.page_margins.top = 0.3
    ws.page_margins.bottom = 0.3
    ws.page_margins.header = 0.1
    ws.page_margins.footer = 0.1

    # スタイル定義
    font_title = Font(name="Meiryo UI", size=14, bold=True)
    font_header = Font(name="Meiryo UI", size=9, bold=True)
    font_data = Font(name="Meiryo UI", size=9)
    font_bold = Font(name="Meiryo UI", size=9, bold=True)

    fill_header = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    fill_blue_total = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")

    border_thin_side = Side(border_style="thin", color="D9D9D9")
    border_thin = Border(left=border_thin_side, right=border_thin_side, top=border_thin_side, bottom=border_thin_side)
    border_double_bottom = Border(top=Side(border_style="thin", color="000000"), bottom=Side(border_style="double", color="000000"))

    align_center = Alignment(horizontal="center", vertical="center")
    align_left = Alignment(horizontal="left", vertical="center")
    align_right = Alignment(horizontal="right", vertical="center")

    # タイトル行
    ws["A1"] = f"活動集計 ({month})" if month else "活動集計"
    ws["A1"].font = font_title
    ws["A1"].alignment = align_left
    ws.row_dimensions[1].height = 24

    # ヘッダー列
    headers = ["担当者", "エリア", "区分", "訪問件数", "電話件数", "合計", "ファイル名"]
    ws.row_dimensions[3].height = 20
    for idx, h in enumerate(headers):
        col_let = get_column_letter(idx + 1)
        ws[f"{col_let}3"] = h
        ws[f"{col_let}3"].font = font_header
        ws[f"{col_let}3"].fill = fill_header
        ws[f"{col_let}3"].alignment = align_center if idx < 3 or idx == 6 else align_right
        ws[f"{col_let}3"].border = border_thin

    # データ行
    row_num = 4
    for r in records:
        ws.row_dimensions[row_num].height = 18

        ws[f"A{row_num}"] = r["staff"]
        ws[f"A{row_num}"].font = font_bold
        ws[f"A{row_num}"].alignment = align_left
        ws[f"A{row_num}"].border = border_thin

        ws[f"B{row_num}"] = r["area"]
        ws[f"B{row_num}"].font = font_data
        ws[f"B{row_num}"].alignment = align_center
        ws[f"B{row_num}"].border = border_thin

        ws[f"C{row_num}"] = r["category"]
        ws[f"C{row_num}"].font = font_data
        ws[f"C{row_num}"].alignment = align_center
        ws[f"C{row_num}"].border = border_thin

        ws[f"D{row_num}"] = r["visits"]
        ws[f"D{row_num}"].font = font_data
        ws[f"D{row_num}"].alignment = align_right
        ws[f"D{row_num}"].border = border_thin

        ws[f"E{row_num}"] = r["calls"]
        ws[f"E{row_num}"].font = font_data
        ws[f"E{row_num}"].alignment = align_right
        ws[f"E{row_num}"].border = border_thin

        # 合計（数式）
        ws[f"F{row_num}"] = f"=D{row_num}+E{row_num}"
        ws[f"F{row_num}"].font = font_bold
        ws[f"F{row_num}"].fill = fill_blue_total
        ws[f"F{row_num}"].alignment = align_right
        ws[f"F{row_num}"].border = border_thin

        ws[f"G{row_num}"] = r["file"]
        ws[f"G{row_num}"].font = font_data
        ws[f"G{row_num}"].alignment = align_left
        ws[f"G{row_num}"].border = border_thin

        row_num += 1

    # 合計行
    ws.row_dimensions[row_num].height = 20
    ws[f"A{row_num}"] = "合計"
    ws[f"A{row_num}"].font = font_bold
    ws[f"A{row_num}"].alignment = align_left
    ws[f"A{row_num}"].border = border_double_bottom

    ws[f"B{row_num}"].border = border_double_bottom
    ws[f"C{row_num}"].border = border_double_bottom

    ws[f"D{row_num}"] = f"=SUM(D4:D{row_num - 1})"
    ws[f"D{row_num}"].font = font_bold
    ws[f"D{row_num}"].alignment = align_right
    ws[f"D{row_num}"].border = border_double_bottom

    ws[f"E{row_num}"] = f"=SUM(E4:E{row_num - 1})"
    ws[f"E{row_num}"].font = font_bold
    ws[f"E{row_num}"].alignment = align_right
    ws[f"E{row_num}"].border = border_double_bottom

    ws[f"F{row_num}"] = f"=SUM(F4:F{row_num - 1})"
    ws[f"F{row_num}"].font = font_bold
    ws[f"F{row_num}"].fill = fill_blue_total
    ws[f"F{row_num}"].alignment = align_right
    ws[f"F{row_num}"].border = border_double_bottom

    ws[f"G{row_num}"].border = border_double_bottom

    # 列幅調整
    ws.column_dimensions["A"].width = 16
    ws.column_dimensions["B"].width = 14
    ws.column_dimensions["C"].width = 10
    ws.column_dimensions["D"].width = 12
    ws.column_dimensions["E"].width = 12
    ws.column_dimensions["F"].width = 12
    ws.column_dimensions["G"].width = 40

    file_stream = io.BytesIO()
    wb.save(file_stream)
    file_stream.seek(0)

    month_fn = month.replace('/', '') if month else "total"
    headers_dict = {
        'Content-Disposition': f'attachment; filename="ActivitySummary_{month_fn}.xlsx"'
    }
    return StreamingResponse(file_stream, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", headers=headers_dict)

STATIC_DIR = os.path.join(BUNDLE_DIR, "static")
if os.path.exists(STATIC_DIR):
    if os.path.exists(os.path.join(STATIC_DIR, "_next")):
         app.mount("/_next", StaticFiles(directory=os.path.join(STATIC_DIR, "_next")), name="next_assets")
    @app.get("/")
    async def serve_index():
        p = os.path.join(STATIC_DIR, "index.html")
        return FileResponse(p) if os.path.exists(p) else {"msg": "No static"}
    @app.get("/{full_path:path}")
    async def serve_spa(full_path: str):
        fp = os.path.join(STATIC_DIR, full_path)
        if os.path.isfile(fp): return FileResponse(fp)
        si = os.path.join(STATIC_DIR, "index.html")
        return FileResponse(si) if os.path.exists(si) else {"detail": "Not Found"}

if __name__ == "__main__":
    import uvicorn, webbrowser, threading
    def ob():
        import time; time.sleep(2); webbrowser.open("http://localhost:8001")
    threading.Thread(target=ob, daemon=True).start()
    uvicorn.run(app, host="0.0.0.0", port=8001, log_level="info")
