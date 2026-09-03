import os
import sys
import tempfile
import openpyxl
import pytest
from fastapi import BackgroundTasks

sys.path.insert(0, os.path.abspath(os.path.join(os.path.dirname(__file__), '..')))

import config
import sync_queue
import routes_reports
import models

def test_enqueue_and_get_pending_tasks():
    """タスクがキューに正しく保存され、FIFOで取得・完了処理できることを検証"""
    temp_dir = tempfile.mkdtemp()
    orig_db = config.SQLITE_CACHE_DB
    test_db = os.path.join(temp_dir, "test_queue.db")
    config.SQLITE_CACHE_DB = test_db

    try:
        assert sync_queue.get_pending_task_count() == 0

        # タスクを2件登録
        t1 = sync_queue.enqueue_sync_task("create", "report.xlsm", {"商談内容": "内容1"})
        t2 = sync_queue.enqueue_sync_task("update", "report.xlsm", {"商談内容": "内容2"}, management_number=10)

        assert sync_queue.get_pending_task_count() == 2

        tasks = sync_queue.get_pending_tasks()
        assert len(tasks) == 2
        # FIFO順（t1 -> t2）
        assert tasks[0]["id"] == t1
        assert tasks[0]["task_type"] == "create"
        assert tasks[0]["payload"]["商談内容"] == "内容1"
        assert tasks[1]["id"] == t2
        assert tasks[1]["management_number"] == 10

        # 完了マーク
        sync_queue.mark_task_completed(t1)
        assert sync_queue.get_pending_task_count() == 1

        # 失敗記録
        sync_queue.mark_task_failed(t2, "Dummy network error")
        assert sync_queue.get_pending_task_count() == 1  # error状態も再試行対象としてカウント

        remaining = sync_queue.get_pending_tasks()
        assert len(remaining) == 1
        assert remaining[0]["id"] == t2
        assert remaining[0]["retry_count"] == 1

    finally:
        config.SQLITE_CACHE_DB = orig_db
        try:
            for f in os.listdir(temp_dir):
                os.remove(os.path.join(temp_dir, f))
            os.rmdir(temp_dir)
        except Exception:
            pass


def test_offline_add_report_fallback():
    """ファイルサーバー切断時、add_reportがエラーにならず安全にキューに一時退避されることを検証"""
    temp_dir = tempfile.mkdtemp()
    orig_db = config.SQLITE_CACHE_DB
    orig_excel_dir = config.EXCEL_DIR
    test_db = os.path.join(temp_dir, "test_fallback.db")
    # 存在しないディレクトリを模してオフライン切断状態を作る
    disconnected_dir = os.path.join(temp_dir, "non_existent_server")
    
    config.SQLITE_CACHE_DB = test_db
    config.EXCEL_DIR = disconnected_dir

    bg_tasks = BackgroundTasks()

    try:
        data = {
            "日付": "26/08/18",
            "行動内容": "訪問（新規）",
            "商談内容": "オフライン中に作成された商談内容",
        }
        report_in = models.ReportInput(**data)

        # 実行：ファイルサーバーが存在しないためオフライン退避される
        res = routes_reports.add_report(report_in, bg_tasks, filename="missing.xlsm")

        assert res.get("status") == "queued"
        assert res.get("offline") is True
        assert "一時退避しました" in res.get("message", "")

        # キューに保存されたか確認
        assert sync_queue.get_pending_task_count() == 1
        tasks = sync_queue.get_pending_tasks()
        assert tasks[0]["task_type"] == "create"
        assert tasks[0]["payload"]["商談内容"] == "オフライン中に作成された商談内容"

    finally:
        config.SQLITE_CACHE_DB = orig_db
        config.EXCEL_DIR = orig_excel_dir
        try:
            for f in os.listdir(temp_dir):
                os.remove(os.path.join(temp_dir, f))
            os.rmdir(temp_dir)
        except Exception:
            pass


def test_process_sync_queue_reconnect():
    """ファイルサーバー復旧時にprocess_sync_queueが未同期データを原本Excelに自動保存することを検証"""
    temp_dir = tempfile.mkdtemp()
    orig_db = config.SQLITE_CACHE_DB
    orig_excel_dir = config.EXCEL_DIR
    test_db = os.path.join(temp_dir, "test_reconnect.db")
    test_filename = "reconnect.xlsm"
    test_filepath = os.path.join(temp_dir, test_filename)

    # 1. 正常なテンプレートExcelを作成
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "営業日報"
    ws.cell(row=1, column=1, value="管理番号")
    ws.cell(row=1, column=2, value="商談内容")
    wb.save(test_filepath)
    wb.close()

    config.SQLITE_CACHE_DB = test_db
    config.EXCEL_DIR = temp_dir

    try:
        # キューに2件の操作（追加1件 + コメント返信1件）を事前に退避
        sync_queue.enqueue_sync_task(
            task_type="create",
            filename=test_filename,
            payload={"日付": "26/08/18", "商談内容": "自動同期された新規商談"}
        )
        sync_queue.enqueue_sync_task(
            task_type="reply",
            filename=test_filename,
            payload={"コメント返信欄": "オフライン中に書いた返信"},
            management_number=1
        )

        assert sync_queue.get_pending_task_count() == 2

        # 同期ワーカーを実行（ファイルサーバーが復旧した状態）
        result = sync_queue.process_sync_queue()
        assert result["processed"] == 2
        assert result["failed"] == 0
        assert result["remaining"] == 0
        assert sync_queue.get_pending_task_count() == 0

        import excel_schema
        # Excel原本を開いて実際に反映されたか確認
        check_wb = openpyxl.load_workbook(test_filepath)
        check_ws = check_wb["営業日報"]
        # 管理番号 1 の行が追加されていること
        assert check_ws.cell(row=2, column=1).value == 1
        # 日付が反映されていること
        assert check_ws.cell(row=2, column=2).value == "26/08/18"
        # 商談内容が反映されていること
        assert check_ws.cell(row=2, column=excel_schema.DailyReportColumns.BUSINESS_CONTENT).value == "自動同期された新規商談"
        # コメント返信欄が反映されていること
        assert check_ws.cell(row=2, column=excel_schema.DailyReportColumns.COMMENT_REPLY).value == "オフライン中に書いた返信"
        check_wb.close()

    finally:
        config.SQLITE_CACHE_DB = orig_db
        config.EXCEL_DIR = orig_excel_dir
        try:
            for f in os.listdir(temp_dir):
                os.remove(os.path.join(temp_dir, f))
            os.rmdir(temp_dir)
        except Exception:
            pass
