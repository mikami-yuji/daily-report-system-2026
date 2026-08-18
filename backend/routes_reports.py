from fastapi import APIRouter, HTTPException, BackgroundTasks, UploadFile, File, Request
from fastapi.responses import FileResponse, JSONResponse, StreamingResponse
import os
import shutil
import json
import logging
import re
import traceback
import math
import io
from datetime import datetime, timedelta
import pandas as pd
import openpyxl
import numpy as np
from copy import copy
from typing import Optional, List, Dict, Any

import config
import cache
import models
import excel_schema
import excel_io

router = APIRouter()

@router.get("/api/health")
def read_root() -> Dict[str, str]:
    return {"message": "Daily Report API is running", "excel_dir": config.EXCEL_DIR}



@router.get("/api/files")
def list_excel_files() -> Dict[str, Any]:
    """List all Excel files in the directory"""
    logging.debug(f"Listing files in {config.EXCEL_DIR}")
    if not os.path.exists(config.EXCEL_DIR):
         logging.error(f"Directory not found: {config.EXCEL_DIR}")
         raise HTTPException(status_code=500, detail=f"Excel Directory not found: {config.EXCEL_DIR}")
         
    try:
        files = []
        # Add timeout protection or more verbose logging? 
        # listing network drive can appear to hang.
        
        items = os.listdir(config.EXCEL_DIR)
        logging.debug(f"Found {len(items)} items in directory")
        
        for file in items:
            if file.endswith(('.xlsx', '.xlsm')):
                file_path = os.path.join(config.EXCEL_DIR, file)
                try:
                    file_size = os.path.getsize(file_path)
                    file_mtime = os.path.getmtime(file_path)
                    files.append({
                        "name": file,
                        "size": file_size,
                        "modified": datetime.fromtimestamp(file_mtime).isoformat()
                    })
                except Exception as file_err:
                    logging.warning(f"Error processing file {file}: {file_err}")
                    continue
                    
        return {"files": files, "default": config.DEFAULT_EXCEL_FILE}
    except Exception as e:
        logging.critical(f"CRITICAL ERROR in list_excel_files: {str(e)}")
        import traceback
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=f"Failed to list files: {str(e)}")


# Cache for Excel dataframes: {(filename, sheet_name): {'mtime': float, 'df': pd.DataFrame}}


@router.get("/api/customers")
def get_customers(filename: str = config.DEFAULT_EXCEL_FILE) -> List[Dict[str, Any]]:
    """Get customer list from the Excel file"""
    try:
        # Get dataframe from cache
        df = cache.get_cached_dataframe(filename, excel_schema.SHEET_CUSTOMERS)
        
        # Clean up column names
        excel_schema.clean_column_names(df)
        
        # デバッグログ: カラム名を出力
        logging.info(f"得意先_List columns: {list(df.columns)}")
        
        # 現目標カラムの存在確認
        target_col = None
        for col in df.columns:
            if '現目標' in col or '目標' in col:
                target_col = col
                logging.info(f"Found target column: '{col}'")
        
        # Rename specific columns
        # 都道府県（カラムF）→ エリア: フロントエンドが c.エリア でアクセスするため
        df = df.rename(columns={
            '得意先CD.': '得意先CD',
            '直送先CD.': '直送先CD',
            '都道府県': 'エリア',
        })
        
        # Fill NaN values with empty strings
        df = df.fillna(value='')
        
        # Convert to dict
        records = df.to_dict(orient="records")
        
        # デバッグ: 最初のレコードの現目標値を確認
        if records and len(records) > 0:
            sample = records[0]
            logging.info(f"Sample record keys: {list(sample.keys())}")
            if '現目標' in sample:
                logging.info(f"Sample 現目標 value: '{sample.get('現目標')}'")
            else:
                logging.warning("現目標 column not found in record!")
        
        # Clean the records
        import math
        cleaned_records = []
        for record in records:
            cleaned_record = {}
            for key, value in record.items():
                if isinstance(value, float):
                    if math.isnan(value) or math.isinf(value):
                        cleaned_record[key] = None
                    else:
                        cleaned_record[key] = value
                elif value == '':
                    cleaned_record[key] = None
                elif isinstance(value, str):
                    import re
                    cleaned_value = re.sub(r'_x000D_', '\n', value)
                    cleaned_value = cleaned_value.replace('\r', '')
                    cleaned_record[key] = cleaned_value
                else:
                    cleaned_record[key] = value
            cleaned_records.append(cleaned_record)

        return cleaned_records
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))




@router.get("/api/priority-customers")
def get_priority_customers(filename: str = config.DEFAULT_EXCEL_FILE) -> List[Dict[str, str]]:
    """得意先_Listからカラム H (重点顧客) が「重点」の顧客を取得。カラム I の担当者情報も含める"""
    try:
        # 得意先_Listを読み込み
        df = cache.get_cached_dataframe(filename, excel_schema.SHEET_CUSTOMERS)
        
        # カラム名をクリーンアップ
        excel_schema.clean_column_names(df)
        
        # カラム名を保存
        col_customer_cd = df.columns[0]  # 得意先CD
        
        # 得意先名カラムを特定する（'得意先名'を含むカラムを探す）
        col_customer_name = None
        for col in df.columns:
            if '得意先名' in str(col):
                col_customer_name = col
                break
        if not col_customer_name:
            col_customer_name = df.columns[2] if len(df.columns) > 2 else (df.columns[1] if len(df.columns) > 1 else None)
            
        col_priority = df.columns[7] if len(df.columns) > 7 else None  # カラムH: 重点顧客
        col_staff = df.columns[8] if len(df.columns) > 8 else None  # カラムI: 担当者
        
        # 得意先CDがある行のみ抽出（ヘッダー行や空行を除外）
        df = df.dropna(subset=[col_customer_cd])
        
        # カラム H の名前を特定
        priority_col = col_priority
        if not priority_col:
            # フォールバック: 「重点顧客」という名前のカラムを探す
            for col in df.columns:
                if '重点' in str(col):
                    priority_col = col
                    break
            if not priority_col:
                logging.warning(f"Priority column not found. Columns: {list(df.columns)}")
                return []
        
        logging.info(f"Priority column: {priority_col}, Staff column: {col_staff}")
        
        # 「重点」と記載されている行のみ抽出
        priority_df = df[df[priority_col].astype(str).str.contains('重点', na=False)]
        
        logging.info(f"Found {len(priority_df)} priority customers")
        
        # レコードを作成
        records = []
        for _, row in priority_df.iterrows():
            customer_cd = row[col_customer_cd]
            customer_name = row[col_customer_name] if col_customer_name else ''
            staff = row[col_staff] if col_staff else ''
            
            # CDをクリーンアップ
            if isinstance(customer_cd, float):
                import math
                if math.isnan(customer_cd):
                    continue
                customer_cd = str(int(customer_cd))
            else:
                customer_cd = str(customer_cd).strip()
            
            if not customer_cd:
                continue
            
            # 担当者をクリーンアップ
            if isinstance(staff, float):
                import math
                if math.isnan(staff):
                    staff = ''
                else:
                    staff = str(staff)
            else:
                staff = str(staff).strip() if staff else ''
                
            records.append({
                '得意先CD': customer_cd,
                '得意先名': str(customer_name).strip() if customer_name else '',
                '担当者': staff
            })
        
        return records
    except Exception as e:
        logging.error(f"Error in get_priority_customers: {e}")
        import traceback
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=str(e))





@router.get("/api/interviewers")
def get_interviewers_by_query(customer_code: str, filename: str = config.DEFAULT_EXCEL_FILE) -> List[str]:
    """Get list of interviewers for a specific customer"""
    excel_file = os.path.join(config.EXCEL_DIR, filename)
    if not os.path.exists(excel_file):
        raise HTTPException(status_code=404, detail=f"Excel file '{filename}' not found")
    
    try:
        # Read the '営業日報' sheet
        df = pd.read_excel(excel_file, sheet_name='営業日報', header=0)
        
        # Clean up column names
        df.columns = [str(col).replace('\n', '').strip() for col in df.columns]
        
        # Rename specific columns to match frontend expectations
        df = df.rename(columns={
            '得意先CD.': '得意先CD',
        })
        
        # Filter by customer code
        customer_reports = df[df['得意先CD'] == customer_code]
        
        # Get unique interviewers, excluding NaN and '-'
        interviewers = customer_reports['面談者'].dropna().unique().tolist()
        interviewers = [i for i in interviewers if i and str(i).strip() not in ['-', 'nan', '']]
        
        return interviewers
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))



@router.get("/api/reports/{management_number}")
def get_report_by_id(management_number: int, filename: str = config.DEFAULT_EXCEL_FILE) -> Dict[str, Any]:
    """指定された管理番号の日報を取得"""
    try:
        # Get dataframe from cache
        df = cache.get_cached_dataframe(filename, excel_schema.SHEET_DAILY_REPORT)
        
        # Clean up column names
        excel_schema.clean_column_names(df)
        
        # Rename specific columns
        df = df.rename(columns={
            '得意先CD.': '得意先CD',
            '訪問先名得意先名': '訪問先名',
            'コメント': '上長コメント',  # Excel uses 'コメント' for manager comment
            'コメント返信欄': 'コメント返信欄'  # Keep as-is for reply field
        })
        
        # Filter by management number
        report_df = df[df['管理番号'] == management_number]
        
        if report_df.empty:
            raise HTTPException(status_code=404, detail=f"Report with management number {management_number} not found")
        
        # Get the first (and should be only) record
        record = report_df.iloc[0].to_dict()
        
        # Clean the record
        import math
        cleaned_record = {}
        for key, value in record.items():
            if isinstance(value, float):
                if math.isnan(value) or math.isinf(value):
                    cleaned_record[key] = None
                elif key == '得意先CD' and not math.isnan(value):
                    cleaned_record[key] = str(int(value))
                else:
                    cleaned_record[key] = value
            elif value == '':
                cleaned_record[key] = None
            elif isinstance(value, str):
                import re
                cleaned_value = re.sub(r'_x000D_', '\n', value)
                cleaned_value = cleaned_value.replace('\r', '')
                cleaned_record[key] = cleaned_value
            else:
                cleaned_record[key] = value
        
        return cleaned_record
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))



@router.get("/api/reports")
def get_reports(filename: str = config.DEFAULT_EXCEL_FILE) -> List[Dict[str, Any]]:
    import time
    import math
    import re
    
    max_retries = 2
    last_error = None
    
    for attempt in range(max_retries + 1):
        try:
            logging.info(f"Fetching reports for {filename} from {config.EXCEL_DIR} (attempt {attempt + 1}/{max_retries + 1})")
            # キャッシュからDataFrameを取得
            df = cache.get_cached_dataframe(filename, excel_schema.SHEET_DAILY_REPORT)
            
            # 列名のクリーンアップ（改行除去・空白トリム）
            excel_schema.clean_column_names(df)
            
            # フロントエンド期待値に合わせた列名のリネーム
            df = df.rename(columns={
                '得意先CD.': '得意先CD',
                '訪問先名得意先名': '訪問先名',
                '直送先CD.': '直送先CD',
                '直送先名.': '直送先名',
                'コメント': '上長コメント',
                'コメント返信欄': 'コメント返信欄'
            })
            
            # NaN・無限大・null値を空文字に置換
            df = df.fillna(value='')
            
            # 日付列を文字列に変換（シリアライズエラー回避）
            if '日付' in df.columns:
                 df['日付'] = df['日付'].astype(str)
                 
            # システム確認用デザインNo.が空の場合、デザイン依頼No.で補完する
            if 'システム確認用デザインNo.' in df.columns and 'デザイン依頼No.' in df.columns:
                import numpy as np
                df['システム確認用デザインNo.'] = np.where(
                    (df['システム確認用デザインNo.'] == '') | df['システム確認用デザインNo.'].isna(), 
                    df['デザイン依頼No.'], 
                    df['システム確認用デザインNo.']
                )
            
            # レコード変換とJSON互換性のクリーニング
            records = df.to_dict(orient="records")
            
            cleaned_records = []
            for record in records:
                cleaned_record = {}
                for key, value in record.items():
                    if isinstance(value, float):
                        if math.isnan(value) or math.isinf(value):
                            cleaned_record[key] = None
                        # 得意先CD・直送先CDを小数なし文字列に変換
                        elif key in ['得意先CD', '直送先CD'] and not math.isnan(value):
                            cleaned_record[key] = str(int(value))
                        else:
                            cleaned_record[key] = value
                    elif value == '':
                        cleaned_record[key] = None
                    elif isinstance(value, str):
                        # Excelの改行コードアーティファクトを正規の改行に置換
                        cleaned_value = re.sub(r'_x000D_', '\n', value)
                        cleaned_value = cleaned_value.replace('\r', '')
                        cleaned_record[key] = cleaned_value
                    else:
                        cleaned_record[key] = value
                cleaned_records.append(cleaned_record)

            logging.info(f"Successfully fetched {len(cleaned_records)} reports for {filename}")
            return cleaned_records
            
        except HTTPException:
            # HTTPException（404など）はそのまま再送出
            raise
        except (PermissionError, OSError) as e:
            # ファイルロック・ネットワーク障害は一時的な可能性があるためリトライ
            last_error = e
            logging.warning(f"File access error for {filename} (attempt {attempt + 1}): {type(e).__name__}: {e}")
            if attempt < max_retries:
                time.sleep(1)
                # キャッシュをクリアして再試行
                cache.CACHE.pop((filename, '営業日報'), None)
                continue
            else:
                logging.error(f"All retries exhausted for {filename}: {e}")
                import traceback
                traceback.print_exc()
                raise HTTPException(
                    status_code=503,
                    detail=f"ファイルへのアクセスに失敗しました（{type(e).__name__}）。ファイルが他のユーザーに開かれているか、ネットワーク接続を確認してください。"
                )
        except Exception as e:
            # その他の予期しないエラー
            logging.error(f"Unexpected error fetching reports for {filename}: {type(e).__name__}: {e}")
            import traceback
            traceback.print_exc()
            raise HTTPException(
                status_code=500, 
                detail=f"日報データの読み込みに失敗しました: {type(e).__name__}: {str(e)}"
            )



@router.get("/api/interviewers/{customer_cd}")
def get_interviewers(
    customer_cd: str, 
    filename: str = config.DEFAULT_EXCEL_FILE,
    customer_name: Optional[str] = None,
    delivery_name: Optional[str] = None
) -> Dict[str, Any]:
    """Get list of interviewers for a specific customer with optional name filtering"""
    try:
        # Get dataframe from cache
        df = cache.get_cached_dataframe(filename, excel_schema.SHEET_DAILY_REPORT)
        
        # Clean up column names
        excel_schema.clean_column_names(df)
        
        # Rename specific columns to standard names
        df = df.rename(columns={
            '得意先CD.': '得意先CD',
            '直送先CD.': '直送先CD',
            '訪問先名得意先名': '訪問先名', # Case where \n was removed
            '訪問先名\n得意先名': '訪問先名', # Just in case
        })
        
        # Convert customer_cd to float for matching (Excel stores as float)
        try:
            customer_cd_float = float(customer_cd)
        except ValueError:
            # If conversion fails, try string matching
            customer_cd_float = customer_cd
        
        # Base filter: customer code
        # Handle potential float/string mismatch by trying both if needed, but usually float is correct for number-like CDs
        customer_reports = df[df['得意先CD'] == customer_cd_float]
        
        # Name filtering logic
        if delivery_name:
            # If delivery_name is provided, filter by it
            # Ensure '直送先名' column exists
            if '直送先名' in customer_reports.columns:
                customer_reports = customer_reports[customer_reports['直送先名'] == delivery_name]
        
        # If no delivery_name, or if we want to also filter by customer_name when delivery_name is NOT set
        # The requirement is "past interviewers matching ... Direct Delivery Name (if provided) OR Customer Name (if matching)"
        # If we selected a delivery destination, we used the block above.
        # If we selected a customer (without direct delivery), we should strictly filter by that customer name if possible, 
        # to avoid showing interviewers from other branches if they share the same Customer CD (rare but possible).
        # OR more importantly, if we are in "Direct Delivery" mode, we already filtered.
        # If we are NOT in "Direct Delivery" mode (delivery_name is None), we might want to filter by customer_name
        # to ensure we don't pick up rows that HAVE a direct delivery name (i.e. different destination).
        elif customer_name:
             if '訪問先名' in customer_reports.columns:
                 # Filter rows where Visit Name contains customer_name OR matches exactly
                 # Since '訪問先名' in DB might contain 'Customer Name Direct Delivery Name' now,
                 # strict matching might be tricky.
                 # But usually '訪問先名' == 'Customer Name' for standard records.
                 # Let's try exact match or contains.
                 # Also, we might want to EXCLUDE rows that have a '直送先名' if we are selecting a generic customer?
                 # User said: " 直近のログではなく、得意先名や直送先名が一致する過去に入力した面談者のみ"
                 
                 # If delivery_name is NOT provided, it effectively means we want records for the main customer.
                 # So we should probably match rows where Visit Name is exactly Customer Name, OR contains it.
                 # Let's use simple filtering: if customer_name is passed, try to filter by it.
                 # But keep in mind 訪問先名 might include 直送先名 suffix now.
                 pass

        # Updated Strategy for clarity:
        # 1. Base: Customer CD
        # 2. If delivery_name provided: matches '直送先名' == delivery_name
        # 3. If delivery_name NOT provided but customer_name provided: matches '訪問先名' == customer_name OR '直送先名' is Empty/NaN
        #    (This covers the case where we selected the main customer and want to exclude direct delivery branches)
        
        if delivery_name and '直送先名' in customer_reports.columns:
            customer_reports = customer_reports[customer_reports['直送先名'] == delivery_name]
        elif '直送先名' in customer_reports.columns:
             # If no delivery name specified, we prefer rows that ALSO have no delivery name specified
             # This avoids suggesting interviewers specific to a branch when we selected the HQ
             customer_reports = customer_reports[
                 customer_reports['直送先名'].isna() | 
                 (customer_reports['直送先名'] == '') | 
                 (customer_reports['直送先名'] == None)
             ]

        interviewers = customer_reports['面談者'].dropna().unique().tolist()
        
        # Remove empty strings and sort
        interviewers = [str(i).strip() for i in interviewers if str(i).strip() and str(i).strip() != 'nan']
        interviewers = sorted(set(interviewers))
        
        return {"customer_cd": customer_cd, "interviewers": interviewers}
    except Exception as e:
        logging.error(f"Error in get_interviewers: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))



@router.get("/api/designs/{customer_cd}")
def get_designs(
    customer_cd: str, 
    delivery_name: Optional[str] = None, 
    filename: str = config.DEFAULT_EXCEL_FILE
) -> Dict[str, Any]:
    """Get list of design requests for a specific customer (optionally filtered by delivery destination)"""
    try:
        logging.info(f"get_designs called: customer_cd={customer_cd}, delivery_name={delivery_name}")
        
        # Get dataframe from cache
        df = cache.get_cached_dataframe(filename, excel_schema.SHEET_DAILY_REPORT)
        
        # Clean up column names
        excel_schema.clean_column_names(df)
        
        # Rename specific columns
        df = df.rename(columns={
            '得意先CD.': '得意先CD',
            '直送先名.': '直送先名'
        })
        
        logging.info(f"Columns in dataframe: {list(df.columns)[:20]}...")  # Log first 20 columns
        
        # Convert customer_cd to float for matching
        try:
            customer_cd_float = float(customer_cd)
        except ValueError:
            customer_cd_float = customer_cd
        
        # Filter by customer code
        customer_reports = df[df['得意先CD'] == customer_cd_float].copy()
        
        # Fallback to name search if no records found by CD (e.g. manually typed customer name, or custom code not in list)
        if len(customer_reports) == 0 and str(customer_cd).strip():
            logging.info(f"No designs found for customer_cd={customer_cd} as code. Trying name match...")
            name_lower = str(customer_cd).lower().strip()
            cond = pd.Series(False, index=df.index)
            if '得意先名' in df.columns:
                cond = cond | df['得意先名'].astype(str).str.lower().str.contains(name_lower, na=False)
            if '直送先名' in df.columns:
                cond = cond | df['直送先名'].astype(str).str.lower().str.contains(name_lower, na=False)
            if '訪問先名' in df.columns:
                cond = cond | df['訪問先名'].astype(str).str.lower().str.contains(name_lower, na=False)
            if '訪問先名得意先名' in df.columns:
                cond = cond | df['訪問先名得意先名'].astype(str).str.lower().str.contains(name_lower, na=False)
            
            customer_reports = df[cond].copy()
            logging.info(f"After name partial match filter: {len(customer_reports)} rows")
            
        logging.info(f"After customer filter: {len(customer_reports)} rows")
        
        # Filter by delivery destination if provided
        if delivery_name and '直送先名' in customer_reports.columns:
            before_count = len(customer_reports)
            customer_reports = customer_reports[customer_reports['直送先名'] == delivery_name]
            logging.info(f"After delivery_name filter ({delivery_name}): {len(customer_reports)} rows (was {before_count})")
        else:
            logging.info(f"Skipping delivery_name filter: delivery_name={delivery_name}, has_column={'直送先名' in customer_reports.columns}")
        
        # Filter for records with design request number
        design_reports = customer_reports[customer_reports['デザイン依頼No.'].notna()]
        logging.info(f"After design no filter: {len(design_reports)} rows")
        
        # Get unique design request numbers
        unique_design_nos = design_reports['デザイン依頼No.'].unique()
        
        designs = []
        for design_no in unique_design_nos:
            # Get all records for this design number
            design_records = design_reports[design_reports['デザイン依頼No.'] == design_no]
            
            # Get the latest record (assuming lower down in Excel is newer, or we could sort by date if available)
            # Here we just take the last one in the dataframe which corresponds to the last row in Excel
            latest_record = design_records.iloc[-1]
            
            # Get the status
            status = str(latest_record['デザイン進捗状況']) if pd.notna(latest_record['デザイン進捗状況']) else ""
            
            # Skip designs with completed/rejected statuses: 出稿, 不採用(コンペ負け), 不採用(企画倒れ)
            if '出稿' in status or 'コンペ負け' in status or '企画倒れ' in status:
                continue
            
            design_info = {
                "デザイン依頼No": design_no,
                "デザイン名": str(latest_record['デザイン名']) if pd.notna(latest_record['デザイン名']) else "",
                "デザイン種別": str(latest_record['デザイン種別']) if pd.notna(latest_record['デザイン種別']) else "",
                "デザイン進捗状況": status,
                "デザイン提案有無": str(latest_record['デザイン提案有無']) if pd.notna(latest_record['デザイン提案有無']) else ""
            }
            designs.append(design_info)
            
        return {"customer_cd": customer_cd, "designs": designs}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))



@router.get("/api/customers/suggest-area")
def suggest_area(customer_name: str, filename: str = config.DEFAULT_EXCEL_FILE) -> Dict[str, str]:
    """Suggest area for a typed customer name based on historical report entries"""
    try:
        if not customer_name or not customer_name.strip():
            return {"customer_name": customer_name, "suggested_area": ""}
            
        df = cache.get_cached_dataframe(filename, excel_schema.SHEET_DAILY_REPORT)
        excel_schema.clean_column_names(df)
        
        name_lower = customer_name.lower().strip()
        cond = pd.Series(False, index=df.index)
        for col in ['得意先名', '直送先名', '訪問先名', '訪問先名得意先名']:
            if col in df.columns:
                cond = cond | df[col].astype(str).str.lower().str.contains(name_lower, na=False)
                
        matched_reports = df[cond].copy()
        if len(matched_reports) == 0:
            return {"customer_name": customer_name, "suggested_area": ""}
            
        # エリアが記載されている履歴行に絞り込む
        area_col = 'エリア'
        if area_col in matched_reports.columns:
            valid_area_reports = matched_reports[matched_reports[area_col].notna()]
            valid_area_reports = valid_area_reports[valid_area_reports[area_col].astype(str).str.strip() != '']
            valid_area_reports = valid_area_reports[~valid_area_reports[area_col].astype(str).str.lower().isin(['nan', 'none'])]
            
            if len(valid_area_reports) > 0:
                latest_row = valid_area_reports.iloc[-1]
                area = str(latest_row[area_col]).strip()
                return {"customer_name": customer_name, "suggested_area": area}
            
        return {"customer_name": customer_name, "suggested_area": ""}
    except Exception as e:
        logging.error(f"Error in suggest_area: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))




@router.post("/api/reports")
def add_report(
    report: models.ReportInput, 
    background_tasks: BackgroundTasks, 
    filename: str = config.DEFAULT_EXCEL_FILE
) -> Dict[str, Any]:
    filename = os.path.basename(filename)
    logging.info(f"DEBUG ADD_REPORT: Received payload: {report.model_dump()}")
    excel_file = os.path.join(config.EXCEL_DIR, filename)
    if not os.path.exists(excel_file):
        raise HTTPException(status_code=404, detail=f"Excel file '{filename}' not found")
    
    wb = None
    try:
        # Load workbook with openpyxl to preserve formulas and macros
        wb = openpyxl.load_workbook(excel_file, keep_vba=True)
        ws = wb['営業日報']
        
        # 得意先_Listから現目標を取得
        current_target = ""
        if report.得意先CD and '得意先_List' in wb.sheetnames:
            customer_ws = wb['得意先_List']
            customer_cd = str(report.得意先CD).strip()
            direct_delivery_cd = str(report.直送先CD).strip() if report.直送先CD else ""
            
            # 得意先_Listの構造: A=得意先CD, B=直送先CD, ..., J=現目標
            for row in range(2, customer_ws.max_row + 1):
                cell_customer_cd = customer_ws.cell(row=row, column=1).value
                cell_dd_cd = customer_ws.cell(row=row, column=2).value
                
                if cell_customer_cd is not None:
                    # 得意先CDを文字列に変換（floatの場合は整数に）
                    if isinstance(cell_customer_cd, float):
                        cell_customer_cd = str(int(cell_customer_cd))
                    else:
                        cell_customer_cd = str(cell_customer_cd).strip()
                    
                    # 直送先CDも同様に処理
                    if cell_dd_cd is not None:
                        if isinstance(cell_dd_cd, float):
                            cell_dd_cd = str(int(cell_dd_cd))
                        else:
                            cell_dd_cd = str(cell_dd_cd).strip()
                    else:
                        cell_dd_cd = ""
                    
                    # マッチング条件: 得意先CDが一致 AND (直送先CDが一致 OR 両方空)
                    if cell_customer_cd == customer_cd:
                        if direct_delivery_cd:
                            # 直送先が指定されている場合は直送先CDも一致する必要がある
                            if cell_dd_cd == direct_delivery_cd:
                                target_value = customer_ws.cell(row=row, column=10).value  # J列=現目標
                                if target_value:
                                    current_target = str(target_value).strip()
                                break
                        else:
                            # 直送先が指定されていない場合は、直送先CDが空の行を探す
                            if not cell_dd_cd:
                                target_value = customer_ws.cell(row=row, column=10).value  # J列=現目標
                                if target_value:
                                    current_target = str(target_value).strip()
                                break
            
            logging.debug(f"Found current_target for {customer_cd}: {current_target}")
        
        # Find the maximum management number and its row by scanning all rows
        max_mgmt_num = 0
        max_mgmt_row = 1  # Default to header row if no data found
        for row in range(2, ws.max_row + 1):  # Start from row 2 (skip header)
            mgmt_num = ws.cell(row=row, column=1).value
            try:
                if mgmt_num is not None:
                    val = int(mgmt_num)
                    if val > max_mgmt_num:
                        max_mgmt_num = val
                        max_mgmt_row = row
            except (ValueError, TypeError):
                continue
        
        logging.debug(f"Max Mgmt Num: {max_mgmt_num}, Max Mgmt Row: {max_mgmt_row}")

        # Increment to get new management number
        new_mgmt_num = max_mgmt_num + 1
        
        # Insert at the row immediately after the last management number
        next_row = max_mgmt_row + 1
        logging.debug(f"Writing to Row: {next_row}")
        
        # Prepare the data to write
        # Adjust column indices based on actual Excel structure (251113_2026-_-_008.xlsm)
        # Also copy styles from the previous row (max_mgmt_row) to the new row (next_row)
        
        from copy import copy

        def copy_style(source_cell, target_cell):
            if source_cell.has_style:
                target_cell.font = copy(source_cell.font)
                target_cell.border = copy(source_cell.border)
                target_cell.fill = copy(source_cell.fill)
                target_cell.number_format = copy(source_cell.number_format)
                target_cell.protection = copy(source_cell.protection)
                target_cell.alignment = copy(source_cell.alignment)

        # Prepare data to write using excel_schema
        columns_to_write = excel_schema.get_new_report_column_data(report, new_mgmt_num, current_target)

        for col_idx, value in columns_to_write.items():
            target_cell = ws.cell(row=next_row, column=col_idx)
            target_cell.value = value
            
            # Copy style from the row above (max_mgmt_row)
            if max_mgmt_row >= 2:
                source_cell = ws.cell(row=max_mgmt_row, column=col_idx)
                copy_style(source_cell, target_cell)

        # Save workbook safely with retry
        excel_io.safe_save_workbook_with_retry(wb, excel_file)
        
        # Clear cache
        cache_key = (filename, '営業日報')
        if cache_key in cache.CACHE:
            del cache.CACHE[cache_key]
        
        return {
            "message": "Report added successfully", 
            "management_number": new_mgmt_num,
            "file_path": os.path.abspath(excel_file)
        }
    except HTTPException:
        raise
    except Exception as e:
        logging.error(f"Error in add_report: {e}", exc_info=True)
        raise HTTPException(status_code=500, detail="Internal server error")
    finally:
        if wb:
            try:
                wb.close()
            except Exception:
                pass

# コメント更新専用エンドポイント


@router.patch("/api/reports/{management_number}/reply")
def update_report_reply(management_number: int, reply: models.ReplyInput, background_tasks: BackgroundTasks, filename: str = config.DEFAULT_EXCEL_FILE):
    """コメント返信欄のみを更新（安全な保存）"""
    import tempfile
    import shutil
    
    filename = os.path.basename(filename)
    logging.debug(f"update_report_reply: management_number={management_number}, reply={reply.コメント返信欄}")
    wb = None
    try:
        excel_file = os.path.join(config.EXCEL_DIR, filename)
        logging.debug(f"excel_file: {excel_file}")
        
        wb = openpyxl.load_workbook(excel_file, keep_vba=True)
        logging.debug("workbook loaded")
        if '営業日報' not in wb.sheetnames:
            raise HTTPException(status_code=404, detail="Sheet '営業日報' not found")
        
        ws = wb['営業日報']
        logging.debug(f"worksheet max_row: {ws.max_row}")
        
        # Find the row
        target_row = None
        for row in range(2, ws.max_row + 1):
            val = ws.cell(row=row, column=1).value
            if val is not None:
                try:
                    if int(float(str(val))) == int(management_number):
                        target_row = row
                        break
                except (ValueError, TypeError):
                    if val == management_number:
                        target_row = row
                        break
        
        logging.debug(f"target_row: {target_row}")
        if not target_row:
            wb.close()
            raise HTTPException(status_code=404, detail=f"Report {management_number} not found")
        
        # Set reply column from excel_schema
        ws.cell(row=target_row, column=excel_schema.DailyReportColumns.COMMENT_REPLY, value=reply.コメント返信欄)
        
        # 安全な保存（リトライ・バックアップ付き）
        excel_io.safe_save_workbook_with_retry(wb, excel_file)
        
        # Clear cache
        cache_key = (filename, '営業日報')
        if cache_key in cache.CACHE:
            del cache.CACHE[cache_key]
        
        return {"success": True, "management_number": management_number}
    except HTTPException:
        raise
    except Exception as e:
        logging.error(f"Error in update_report_reply: {e}", exc_info=True)
        raise HTTPException(status_code=500, detail="Internal server error")
    finally:
        if wb:
            try:
                wb.close()
            except Exception:
                pass



@router.patch("/api/reports/{management_number}/comment")
def update_report_comment(management_number: int, comment: models.CommentInput, background_tasks: BackgroundTasks, filename: str = config.DEFAULT_EXCEL_FILE):
    """上長コメントとコメント返信欄を個別に更新（安全な保存）"""
    import tempfile
    import shutil
    
    filename = os.path.basename(filename)
    logging.debug(f"update_report_comment: management_number={management_number}, 上長コメント={comment.上長コメント}, コメント返信欄={comment.コメント返信欄}")
    wb = None
    try:
        excel_file = os.path.join(config.EXCEL_DIR, filename)
        
        wb = openpyxl.load_workbook(excel_file, keep_vba=True)
        if '営業日報' not in wb.sheetnames:
            raise HTTPException(status_code=404, detail="Sheet '営業日報' not found")
        
        ws = wb['営業日報']
        
        # Find the row
        target_row = None
        for row in range(2, ws.max_row + 1):
            val = ws.cell(row=row, column=1).value
            if val is not None:
                try:
                    if int(float(str(val))) == int(management_number):
                        target_row = row
                        break
                except (ValueError, TypeError):
                    if val == management_number:
                        target_row = row
                        break
        
        if not target_row:
            wb.close()
            raise HTTPException(status_code=404, detail=f"Report {management_number} not found")
        
        # --- Optimistic Locking Check ---
        if comment.original_values:
            logging.debug(f"Performing conflict check for Comment {management_number}")
            check_fields = {
                excel_schema.DailyReportColumns.BOSS_COMMENT: '上長コメント',
                excel_schema.DailyReportColumns.COMMENT_REPLY: 'コメント返信欄'
            }
            conflicts = []
            for col_idx, field_name in check_fields.items():
                current_val = ws.cell(row=target_row, column=col_idx).value
                current_str = str(current_val) if current_val is not None else ""
                original_val = comment.original_values.get(field_name, "")
                original_str = str(original_val) if original_val is not None else ""
                
                # Normalize newlines for comparison
                current_str = current_str.replace('\r\n', '\n').replace('\r', '\n').strip()
                original_str = original_str.replace('\r\n', '\n').replace('\r', '\n').strip()
                
                if current_str != original_str:
                    logging.warning(f"CONFLICT: Field '{field_name}' changed. Current: '{current_str}' vs Original: '{original_str}'")
                    conflicts.append(field_name)
            
            if conflicts:
                conflict_msg = ", ".join(conflicts)
                raise HTTPException(
                    status_code=409, 
                    detail=f"他の方がコメントを編集しました（{conflict_msg}）。最新の情報を読み込んでからやり直してください。"
                )
        # --------------------------------
        
        # Update only provided fields using excel_schema
        if comment.上長コメント is not None:
            ws.cell(row=target_row, column=excel_schema.COMMENT_COLUMN_MAPPING['上長コメント'], value=comment.上長コメント)
        if comment.コメント返信欄 is not None:
            ws.cell(row=target_row, column=excel_schema.COMMENT_COLUMN_MAPPING['コメント返信欄'], value=comment.コメント返信欄)
        
        # 安全な保存（リトライ・バックアップ付き）
        excel_io.safe_save_workbook_with_retry(wb, excel_file)
        
        # Clear cache
        cache_key = (filename, '営業日報')
        if cache_key in cache.CACHE:
            del cache.CACHE[cache_key]
        
        return {"success": True, "management_number": management_number}
    except HTTPException:
        raise
    except Exception as e:
        logging.error(f"Error in update_report_comment: {e}", exc_info=True)
        raise HTTPException(status_code=500, detail="Internal server error")
    finally:
        if wb:
            try:
                wb.close()
            except Exception:
                pass



@router.patch("/api/reports/{management_number}/approval")
def update_report_approval(management_number: int, approval: models.ApprovalInput, background_tasks: BackgroundTasks, filename: str = config.DEFAULT_EXCEL_FILE):
    """承認チェック（上長、山澄常務、岡本常務、中野次長、既読チェック）を個別に更新"""
    import tempfile
    import shutil
    
    filename = os.path.basename(filename)
    logging.debug(f"update_report_approval: management_number={management_number}")
    wb = None
    try:
        excel_file = os.path.join(config.EXCEL_DIR, filename)
        
        wb = openpyxl.load_workbook(excel_file, keep_vba=True)
        if '営業日報' not in wb.sheetnames:
            raise HTTPException(status_code=404, detail="Sheet '営業日報' not found")
        
        ws = wb['営業日報']
        
        # Find the row
        target_row = None
        for row in range(2, ws.max_row + 1):
            val = ws.cell(row=row, column=1).value
            if val is not None:
                try:
                    if int(float(str(val))) == int(management_number):
                        target_row = row
                        break
                except (ValueError, TypeError):
                    if val == management_number:
                        target_row = row
                        break
        
        if not target_row:
            raise HTTPException(status_code=404, detail=f"Report {management_number} not found")
        
        # --- Optimistic Locking Check ---
        if approval.original_values:
            logging.debug(f"Performing conflict check for Approval {management_number}")
            conflicts = []
            for field_name, col_idx in excel_schema.APPROVAL_COLUMN_MAPPING.items():
                if getattr(approval, field_name) is not None:
                    current_val = ws.cell(row=target_row, column=col_idx).value
                    current_str = str(current_val) if current_val is not None else ""
                    original_val = approval.original_values.get(field_name, "")
                    original_str = str(original_val) if original_val is not None else ""
                    
                    def norm_val(v):
                        v_str = str(v).strip()
                        if v_str in ['ü', '✓', '済']:
                            return '✓'
                        return ''
                    
                    if norm_val(current_str) != norm_val(original_str):
                        logging.warning(f"CONFLICT: Approval Field '{field_name}' changed. Current: '{current_str}' vs Original: '{original_str}'")
                        conflicts.append(field_name)
            
            if conflicts:
                conflict_msg = ", ".join(conflicts)
                raise HTTPException(
                    status_code=409, 
                    detail=f"他の方が承認ステータスを更新しました（{conflict_msg}）。最新の情報を読み込んでからやり直してください。"
                )
        # --------------------------------
        
        # Update only provided fields using excel_schema
        for field_name, col_idx in excel_schema.APPROVAL_COLUMN_MAPPING.items():
            val = getattr(approval, field_name, None)
            if val is not None:
                ws.cell(row=target_row, column=col_idx, value=val)
        
        # 安全な保存（リトライ・バックアップ付き）
        excel_io.safe_save_workbook_with_retry(wb, excel_file)
        
        # Clear cache
        cache_key = (filename, '営業日報')
        if cache_key in cache.CACHE:
            del cache.CACHE[cache_key]
        
        return {"success": True, "management_number": management_number}
    except HTTPException:
        raise
    except Exception as e:
        logging.error(f"Error in update_report_approval: {e}", exc_info=True)
        raise HTTPException(status_code=500, detail="Internal server error")
    finally:
        if wb:
            try:
                wb.close()
            except Exception:
                pass



@router.post("/api/reports/{management_number}")
def update_report(management_number: int, report: models.ReportInput, background_tasks: BackgroundTasks, filename: str = config.DEFAULT_EXCEL_FILE):
    """既存の日報を更新（全項目対応）"""
    filename = os.path.basename(filename)
    logging.info(f"update_report called: management_number={management_number}, original_values={report.original_values}")
    wb = None
    try:
        excel_file = os.path.join(config.EXCEL_DIR, filename)
        
        # Load the workbook
        wb = openpyxl.load_workbook(excel_file, keep_vba=True)
        if '営業日報' not in wb.sheetnames:
            raise HTTPException(status_code=404, detail="Sheet '営業日報' not found")
             
        ws = wb['営業日報']
        
        # Find the row with the matching management number
        target_row = None
        for row in range(2, ws.max_row + 1):
            val = ws.cell(row=row, column=1).value
            if val is not None:
                try:
                    if int(float(str(val))) == int(management_number):
                        target_row = row
                        break
                except (ValueError, TypeError):
                    if val == management_number:
                        target_row = row
                        break
        
        if not target_row:
            raise HTTPException(status_code=404, detail=f"Report with management number {management_number} not found")

        # --- Optimistic Locking Check ---
        if report.original_values:
            logging.debug(f"Performing conflict check for Report {management_number}")
            
            # Fields to check for conflicts (critical text fields)
            check_fields = {
                excel_schema.DailyReportColumns.BOSS_COMMENT: '上長コメント',
                excel_schema.DailyReportColumns.COMMENT_REPLY: 'コメント返信欄',
                excel_schema.DailyReportColumns.BUSINESS_CONTENT: '商談内容'
            }
            
            conflicts = []
            for col_idx, field_name in check_fields.items():
                current_val = ws.cell(row=target_row, column=col_idx).value
                current_str = str(current_val) if current_val is not None else ""
                
                original_val = report.original_values.get(field_name, "")
                original_str = str(original_val) if original_val is not None else ""
                
                # Normalize newlines for comparison
                current_str = current_str.replace('\r\n', '\n').replace('\r', '\n').strip()
                original_str = original_str.replace('\r\n', '\n').replace('\r', '\n').strip()
                
                if current_str != original_str:
                    logging.warning(f"CONFLICT: Field '{field_name}' changed. Current: '{current_str}' vs Original: '{original_str}'")
                    conflicts.append(field_name)
            
            if conflicts:
                conflict_msg = ", ".join(conflicts)
                raise HTTPException(
                    status_code=409, 
                    detail=f"他の方が編集しました（{conflict_msg}）。最新の情報を読み込んでからやり直してください。"
                )
        # --------------------------------
        
        # Update all fields using excel_schema
        columns_to_write = excel_schema.get_update_report_column_data(report)

        for col_idx, value in columns_to_write.items():
            ws.cell(row=target_row, column=col_idx, value=value)
        
        # Save workbook safely with retry
        excel_io.safe_save_workbook_with_retry(wb, excel_file)
        
        # Clear cache for this file
        cache_key = (filename, '営業日報')
        if cache_key in cache.CACHE:
            del cache.CACHE[cache_key]
        
        return {"message": "Report updated successfully", "management_number": management_number}
    except HTTPException:
        raise
    except Exception as e:
        logging.error(f"Error in update_report: {e}", exc_info=True)
        raise HTTPException(status_code=500, detail="Internal server error")
    finally:
        if wb:
            try:
                wb.close()
            except Exception:
                pass



@router.delete("/api/reports/{management_number}")
def delete_report(management_number: int, filename: str = config.DEFAULT_EXCEL_FILE):
    """指定された管理番号の日報を削除"""
    filename = os.path.basename(filename)
    wb = None
    try:
        excel_file = os.path.join(config.EXCEL_DIR, filename)
        
        # Load the workbook
        wb = openpyxl.load_workbook(excel_file, keep_vba=True)
        if '営業日報' not in wb.sheetnames:
            raise HTTPException(status_code=404, detail="Sheet '営業日報' not found")
             
        ws = wb['営業日報']
        
        # Find the row with the matching management number
        target_row = None
        for row in range(2, ws.max_row + 1):
            val = ws.cell(row=row, column=1).value
            if val is not None:
                try:
                    if int(float(str(val))) == int(management_number):
                        target_row = row
                        break
                except (ValueError, TypeError):
                    if val == management_number:
                        target_row = row
                        break
        
        if not target_row:
            raise HTTPException(status_code=404, detail=f"Report with management number {management_number} not found")
        
        # Delete the row
        ws.delete_rows(target_row, 1)
        
        # Save workbook safely with retry
        excel_io.safe_save_workbook_with_retry(wb, excel_file)
        
        # Clear cache for this file
        cache_key = (filename, '営業日報')
        if cache_key in cache.CACHE:
            del cache.CACHE[cache_key]
        
        return {"message": "Report deleted successfully", "management_number": management_number}
    except HTTPException:
        raise
    except Exception as e:
        logging.error(f"Error in delete_report: {e}", exc_info=True)
        raise HTTPException(status_code=500, detail="Internal server error")
    finally:
        if wb:
            try:
                wb.close()
            except Exception:
                pass




@router.post("/api/upload")
async def upload_file(file: UploadFile = File(...)):
    """Upload an Excel file to the backend directory"""
    try:
        # Validate file extension
        if not file.filename.endswith(('.xlsx', '.xlsm')):
            raise HTTPException(status_code=400, detail="Only .xlsx and .xlsm files are allowed")
        
        # Save the uploaded file
        filename = file.filename.replace('/', '\\')
        filename = os.path.basename(filename)
        file_path = os.path.join(config.EXCEL_DIR, filename)
        
        with open(file_path, "wb") as buffer:
            shutil.copyfileobj(file.file, buffer)
        
        return {
            "message": "File uploaded successfully",
            "filename": file.filename,
            "path": file_path
        }
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


# Set up logging


