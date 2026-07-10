from fastapi import APIRouter, HTTPException, BackgroundTasks, UploadFile, File, Request
from fastapi.responses import FileResponse, JSONResponse, StreamingResponse
import os
import shutil
import json
import logging
import re
import traceback
import math
import io
from datetime import datetime, timedelta
import pandas as pd
import openpyxl
import numpy as np
from copy import copy
from typing import Optional, List, Dict, Any

import config
import cache
import models

router = APIRouter()

@router.get("/api/analytics/team-summary")
def get_team_summary(month: str = None):
    """
    複数ファイルから全メンバーの活動状況を集計する。
    month: "YY/MM" 形式 (例: "26/03")。
    """
    logging.info(f"Team Summary triggered for month: {month}")
    if not os.path.exists(config.EXCEL_DIR):
        raise HTTPException(status_code=500, detail="Excel directory not found")

    # 除外リスト
    EXCLUDE_FILES = [
        "●20260117_2026年度用_日報【原本_2】.xlsm",
        "【電話営業資料】電話　メール話題情報.xlsx",
        "支店999_【津田_操作確認用】_2026年度用日報.xlsm",
        "本社001_【中野次長】_2026年度用日報.xlsm"
    ]

    try:
        # ファイル一覧の取得
        target_files = [
            f for f in os.listdir(config.EXCEL_DIR) 
            if f.endswith('.xlsm') and f not in EXCLUDE_FILES and not f.startswith('~$')
        ]
        logging.info(f"Scanning {len(target_files)} files for aggregation.")

        def extract_staff_name_py(filename: str) -> str:
            import re
            match = re.search(r'【(.+?)】', filename)
            if not match: return "不明"
            content = match.group(1)
            name_with_paren = re.search(r'^(.+?)（(.+?)）', content)
            if name_with_paren: return name_with_paren.group(1) + name_with_paren.group(2)
            surname = re.search(r'^([^\s\u4e00-\u9fa5]*[\u4e00-\u9fa5]+?)(?:課長|次長|部長|常務|社長|主任|係長|専務|取締役|マネージャー|リーダー|担当|氏)?$', content)
            if surname: return surname.group(1)
            return content[:4]

        summary_results = []

        # 月次ターゲットの正規化 (YY/MM -> 20YY-MM)
        target_month_iso = ""
        if month and '/' in month:
            parts = month.split('/')
            target_month_iso = f"20{parts[0]}-{parts[1]}"
            logging.info(f"Normalized target month to: {target_month_iso}")

        # 柔軟な日付解析ヘルパー
        def parse_dt(x):
            if pd.isna(x): return pd.NaT
            s = str(x).strip()
            if s == '' or s == 'nan' or s == '-': return pd.NaT
            try:
                return pd.to_datetime(s, format='%y/%m/%d')
            except:
                pass
            try:
                return pd.to_datetime(s, format='%Y/%m/%d')
            except:
                pass
            try:
                return pd.to_datetime(s, format='%Y-%m-%d')
            except:
                pass
            try:
                return pd.to_datetime(s, errors='coerce')
            except:
                return pd.NaT

        for filename in target_files:
            try:
                staff_name = extract_staff_name_py(filename)
                df = cache.get_cached_dataframe(filename, '営業日報')
                if df is None or df.empty:
                    logging.debug(f"File {filename}: Empty dataframe.")
                    continue

                # カラムクリーンアップ
                df.columns = [str(col).replace('\n', '').strip() for col in df.columns]
                
                date_col = next((c for c in df.columns if '日付' in c), '日付')
                action_col = next((c for c in df.columns if '行動内容' in c), '行動内容')
                priority_col = next((c for c in df.columns if '重点' in c), '重点顧客')
                area_col = next((c for c in df.columns if 'エリア' in c), 'エリア')

                # 日付パースとフィルタ
                df['dt'] = df[date_col].apply(parse_dt)
                df = df.dropna(subset=['dt'])

                if month:
                    # '26/03' などの文字列と比較するマスク
                    mask1 = df['dt'].dt.strftime('%y/%m') == month
                    mask2 = df['dt'].dt.strftime('%Y-%m') == target_month_iso if target_month_iso else pd.Series([False]*len(df))
                    df = df[mask1 | mask2]

                if df.empty:
                    logging.debug(f"File {filename}: No records found after filtering for {month}.")
                    continue

                # 集計
                df['エリア'] = df[area_col].apply(lambda x: str(x).strip() if pd.notnull(x) and str(x).strip() != '' else '未設定')
                
                def get_category(row):
                    val = row.get(priority_col, '')
                    if pd.notnull(val) and str(val).strip() != '' and str(val).strip() != '-':
                        return '重点'
                    return '一般'
                
                df['区分'] = df.apply(get_category, axis=1)
                df['is_visit'] = df[action_col].apply(lambda x: 1 if pd.notnull(x) and '訪問' in str(x) else 0)
                df['is_call'] = df[action_col].apply(lambda x: 1 if pd.notnull(x) and '電話' in str(x) else 0)

                grouped = df.groupby(['エリア', '区分']).agg({
                    'is_visit': 'sum',
                    'is_call': 'sum'
                }).reset_index()

                for _, row in grouped.iterrows():
                    summary_results.append({
                        "staff": staff_name,
                        "area": row['エリア'],
                        "category": row['区分'],
                        "visits": int(row['is_visit']),
                        "calls": int(row['is_call']),
                        "file": filename
                    })
                
                logging.debug(f"File {filename}: Successfully aggregated {len(grouped)} area/category pairs.")

            except Exception as file_err:
                logging.warning(f"Error aggregating file {filename}: {file_err}")
                continue

        logging.info(f"Aggregation complete. Total records: {len(summary_results)}.")
        return {"records": summary_results, "count": len(target_files)}

    except Exception as e:
        logging.error(f"Error in get_team_summary: {e}")
        import traceback
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=str(e))



@router.get("/api/stats/dashboard")
def get_dashboard_stats(filename: str = config.DEFAULT_EXCEL_FILE):
    """ダッシュボード統計をバックエンドで集計して返す"""
    # 空レスポンスのテンプレート（エラー時のフォールバック用）
    empty_response = {
        "summary": {"totalReports": 0, "thisMonth": 0, "visits": 0, "calls": 0},
        "priority": {"uniqueCustomers": 0, "visits": 0, "calls": 0},
        "monthly": [], "ranking": [], "updatedAt": datetime.now().isoformat()
    }
    try:
        logging.info(f"--- Dashboard Stats for {filename} ---")
        # Excelファイルが読めない場合（他のユーザーが開いている等）は空レスポンスを返す
        try:
            df = cache.get_cached_dataframe(filename, '営業日報')
        except Exception as read_err:
            logging.warning(f"Dashboard: Cannot read file {filename}: {read_err}")
            return empty_response

        if df is None or df.empty:
            logging.info(f"Dashboard: Empty dataframe for {filename}")
            return empty_response

        logging.info(f"Dashboard: Raw rows={len(df)}")

        # カラム名のクリーンアップ
        df.columns = [str(col).replace('\n', '').strip() for col in df.columns]
        date_col = next((c for c in df.columns if '日付' in c or '年月日' in c), '日付')
        action_col = next((c for c in df.columns if '行動内容' in c), '行動内容')
        priority_col = next((c for c in df.columns if '重点' in c), '重点顧客')
        customer_col = next((c for c in df.columns if '得意先CD' in c), '得意先CD')
        customer_name_col = next((c for c in df.columns if '訪問先名' in c or '得意先名' in c), '訪問先名')
        area_col = next((c for c in df.columns if 'エリア' in c), 'エリア')

        logging.info(f"Dashboard: Cols detected - date={date_col}, action={action_col}")

        # 日付パース（YY/MM/DD テキスト形式に対応）
        def parse_dt(x):
            if pd.isna(x): return pd.NaT
            s = str(x).strip()
            if s == '' or s == 'nan' or s == '-': return pd.NaT
            # 形式1: YY/MM/DD (26/05/13)
            try:
                d = pd.to_datetime(s, format='%y/%m/%d')
                return d
            except:
                pass
            # 形式2: 汎用パース
            try:
                return pd.to_datetime(s, errors='coerce')
            except:
                return pd.NaT

        df['dt'] = df[date_col].apply(parse_dt)
        vdf = df.dropna(subset=['dt']).copy()

        logging.info(f"Dashboard: Valid date rows={len(vdf)}")

        if vdf.empty:
            logging.warning(f"Dashboard: No valid dates. Sample: {df[date_col].head(3).tolist()}")
            return empty_response

        # 基本統計
        total = len(vdf)
        now = datetime.now()
        this_month_mask = (vdf['dt'].dt.year == now.year) & (vdf['dt'].dt.month == now.month)
        tm_cnt = int(this_month_mask.sum())

        # 訪問・電話フラグ
        is_v = vdf[action_col].astype(str).str.contains('訪問', na=False)
        is_c = vdf[action_col].astype(str).str.contains('電話', na=False)
        visits_total = int(is_v.sum())
        calls_total = int(is_c.sum())

        # 重点顧客
        is_p = vdf[priority_col].fillna('').astype(str).apply(lambda x: x.strip() != '' and x.strip() != '-')
        pdf = vdf[is_p]
        p_unique = int(pdf[customer_col].nunique()) if customer_col in pdf.columns else 0
        p_visits = int((pdf[action_col].astype(str).str.contains('訪問', na=False)).sum())
        p_calls = int((pdf[action_col].astype(str).str.contains('電話', na=False)).sum())

        # 月別推移
        vdf['m'] = vdf['dt'].dt.strftime('%y/%m')
        monthly = []
        for m, g in vdf.groupby('m'):
            mv = int(g[action_col].astype(str).str.contains('訪問', na=False).sum())
            mc = int(g[action_col].astype(str).str.contains('電話', na=False).sum())
            g_p = g[priority_col].fillna('').astype(str).apply(lambda x: x.strip() != '' and x.strip() != '-')
            pv = int((g[action_col].astype(str).str.contains('訪問', na=False) & g_p).sum())
            pc = int((g[action_col].astype(str).str.contains('電話', na=False) & g_p).sum())
            # エリア別
            area_stats = []
            if area_col in g.columns:
                ga = g.copy()
                ga[area_col] = ga[area_col].fillna('未設定').astype(str).str.strip()
                ga.loc[ga[area_col] == '', area_col] = '未設定'
                for an, ag in ga.groupby(area_col):
                    av = int(ag[action_col].astype(str).str.contains('訪問', na=False).sum())
                    ac = int(ag[action_col].astype(str).str.contains('電話', na=False).sum())
                    ap = ag[priority_col].fillna('').astype(str).apply(lambda x: x.strip() != '' and x.strip() != '-')
                    area_stats.append({
                        "area": str(an), "visits": av, "calls": ac,
                        "priorityVisits": int((ag[action_col].astype(str).str.contains('訪問', na=False) & ap).sum()),
                        "priorityCalls": int((ag[action_col].astype(str).str.contains('電話', na=False) & ap).sum())
                    })
            monthly.append({
                "month": m, "visits": mv, "calls": mc,
                "priorityVisits": pv, "priorityCalls": pc,
                "areaBreakdown": area_stats
            })
        monthly.sort(key=lambda x: x['month'], reverse=True)

        # 得意先ランキング (Top 10)
        ranking = []
        rank_base = vdf[is_v | is_c].copy()
        if not rank_base.empty and customer_name_col in rank_base.columns:
            rank_base[customer_name_col] = rank_base[customer_name_col].fillna('名称不明').astype(str)
            rank_base['_rv'] = rank_base[action_col].astype(str).str.contains('訪問', na=False)
            rank_base['_rc'] = rank_base[action_col].astype(str).str.contains('電話', na=False)
            rdf = rank_base.groupby(customer_name_col).agg(v=('_rv', 'sum'), c=('_rc', 'sum'))
            rdf['t'] = rdf['v'] + rdf['c']
            rdf = rdf.sort_values('t', ascending=False).head(10)
            for name, row in rdf.iterrows():
                ranking.append({"name": str(name), "visits": int(row['v']), "calls": int(row['c']), "total": int(row['t'])})

        logging.info(f"Dashboard: Total={total}, ThisMonth={tm_cnt}, Visits={visits_total}, Calls={calls_total}")
        return {
            "summary": {"totalReports": total, "thisMonth": tm_cnt, "visits": visits_total, "calls": calls_total},
            "priority": {"uniqueCustomers": p_unique, "visits": p_visits, "calls": p_calls},
            "monthly": monthly, "ranking": ranking,
            "updatedAt": datetime.now().isoformat()
        }
    except Exception as e:
        logging.error(f"Dashboard error: {e}")
        import traceback
        traceback.print_exc()
        # 500エラーを投げずに空レスポンスを返す（フロントエンドのクラッシュを防止）
        return empty_response



@router.get("/api/stats/monthly-summary")
def get_monthly_summary_stats(filename: str = config.DEFAULT_EXCEL_FILE, month: str = None):
    # month format is 'YY/MM'
    empty_response = {
        "totalReports": 0, "totalVisits": 0, "totalCalls": 0, 
        "priorityVisits": 0, "priorityCalls": 0,
        "totalDesignProposals": 0, "totalDesignCompleted": 0, "totalDesignRejected": 0,
        "uniqueCustomers": 0, "activeDays": 0,
        "areaBreakdown": [], "priorityCustomers": [], "designProgress": [],
        "topCustomers": [], "topCallCustomers": [], "dailyActivity": []
    }
    if not month:
        return empty_response

    try:
        logging.info(f"--- Monthly Summary Stats for {filename}, month={month} ---")
        try:
            df = cache.get_cached_dataframe(filename, '営業日報')
        except Exception as read_err:
            logging.warning(f"MonthlySummary: Cannot read file {filename}: {read_err}")
            return empty_response

        if df is None or df.empty:
            return empty_response

        df.columns = [str(col).replace('\n', '').strip() for col in df.columns]
        date_col = next((c for c in df.columns if '日付' in c or '年月日' in c), '日付')
        action_col = next((c for c in df.columns if '行動内容' in c), '行動内容')
        priority_col = next((c for c in df.columns if '重点' in c), '重点顧客')
        customer_col = next((c for c in df.columns if '得意先CD' in c), '得意先CD')
        customer_name_col = next((c for c in df.columns if '訪問先名' in c or '得意先名' in c), '訪問先名')
        area_col = next((c for c in df.columns if 'エリア' in c), 'エリア')
        dd_code_col = next((c for c in df.columns if '直送先CD' in c), '直送先CD')
        dd_name_col = next((c for c in df.columns if '直送先名' in c), '直送先名')
        rank_col = next((c for c in df.columns if 'ランク' in c), 'ランク')
        design_exist_col = next((c for c in df.columns if 'デザイン提案有無' in c), 'デザイン提案有無')
        design_status_col = next((c for c in df.columns if 'デザイン進捗状況' in c), 'デザイン進捗状況')

        def parse_dt(x):
            if pd.isna(x): return pd.NaT
            s = str(x).strip()
            if s == '' or s == 'nan' or s == '-': return pd.NaT
            try:
                d = pd.to_datetime(s, format='%y/%m/%d')
                return d
            except: pass
            try:
                return pd.to_datetime(s, errors='coerce')
            except: return pd.NaT

        df['dt'] = df[date_col].apply(parse_dt)
        vdf = df.dropna(subset=['dt']).copy()
        if vdf.empty:
            return empty_response

        vdf['m'] = vdf['dt'].dt.strftime('%y/%m')
        mdf = vdf[vdf['m'] == month].copy()
        if mdf.empty:
            return empty_response

        mdf['is_v'] = mdf[action_col].astype(str).str.contains('訪問', na=False)
        mdf['is_c'] = mdf[action_col].astype(str).str.contains('電話', na=False)
        mdf['is_p'] = mdf[priority_col].fillna('').astype(str).apply(lambda x: x.strip() != '' and x.strip() != '-')
        
        has_design = mdf[design_exist_col].astype(str).apply(lambda x: x.strip() in ['有', 'あり']) if design_exist_col in mdf.columns else pd.Series(False, index=mdf.index)
        is_design_new = mdf[design_status_col].astype(str).str.strip() == '新規' if design_status_col in mdf.columns else pd.Series(False, index=mdf.index)
        is_design_proposal = has_design | is_design_new
        
        is_design_completed = mdf[design_status_col].astype(str).str.contains('出稿', na=False) if design_status_col in mdf.columns else pd.Series(False, index=mdf.index)
        is_design_rejected = mdf[design_status_col].astype(str).str.contains('不採用', na=False) if design_status_col in mdf.columns else pd.Series(False, index=mdf.index)

        totalVisits = int(mdf['is_v'].sum())
        totalCalls = int(mdf['is_c'].sum())
        priorityVisits = int((mdf['is_v'] & mdf['is_p']).sum())
        priorityCalls = int((mdf['is_c'] & mdf['is_p']).sum())
        totalDesignProposals = int(is_design_proposal.sum())
        totalDesignCompleted = int(is_design_completed.sum())
        totalDesignRejected = int(is_design_rejected.sum())

        uniqueCustomers = int(mdf[mdf[customer_name_col].astype(str).str.strip() != ''][customer_name_col].nunique()) if customer_name_col in mdf.columns else 0
        activeDays = int(mdf['dt'].nunique())

        areaBreakdown = []
        if area_col in mdf.columns:
            adf = mdf[mdf['is_v'] | mdf['is_c']].copy()
            adf[area_col] = adf[area_col].fillna('未設定').astype(str).str.strip()
            adf.loc[adf[area_col] == '', area_col] = '未設定'
            for an, ag in adf.groupby(area_col):
                areaBreakdown.append({
                    "area": str(an),
                    "visits": int(ag['is_v'].sum()),
                    "calls": int(ag['is_c'].sum()),
                    "priorityVisits": int((ag['is_v'] & ag['is_p']).sum()),
                    "priorityCalls": int((ag['is_c'] & ag['is_p']).sum()),
                    "designProposals": int(is_design_proposal[ag.index].sum()),
                })
            areaBreakdown.sort(key=lambda x: (1 if x['area'] == '未設定' else 0, -(x['visits'] + x['calls'])))

        priorityCustomers = []
        pdf = mdf[mdf['is_p']].copy()
        if not pdf.empty and customer_col in pdf.columns:
            for ccode, cg in pdf.groupby(customer_col):
                p_cname = cg[customer_name_col].iloc[0] if customer_name_col in cg.columns else '不明'
                p_area = cg[area_col].iloc[0] if area_col in cg.columns else ''
                p_rank = cg[rank_col].iloc[0] if rank_col in cg.columns else ''
                
                dd_list = []
                if dd_code_col in cg.columns:
                    for ddcode, ddg in cg[cg[dd_code_col].notna() & (cg[dd_code_col].astype(str).str.strip() != '')].groupby(dd_code_col):
                        dd_cname = ddg[dd_name_col].iloc[0] if dd_name_col in ddg.columns else ''
                        dd_area = ddg[area_col].iloc[0] if area_col in ddg.columns else ''
                        dd_rank = ddg[rank_col].iloc[0] if rank_col in ddg.columns else ''
                        dd_visits = int(ddg['is_v'].sum())
                        dd_calls = int(ddg['is_c'].sum())
                        
                        if dd_visits > 0 or dd_calls > 0 or int(is_design_proposal[ddg.index].sum()) > 0:
                            dd_list.append({
                                "code": str(ddcode).replace('.0', '').strip(),
                                "name": str(dd_cname),
                                "visits": dd_visits,
                                "calls": dd_calls,
                                "designProposals": int(is_design_proposal[ddg.index].sum()),
                                "lastDate": ddg['dt'].max().strftime('%Y/%m/%d') if not pd.isna(ddg['dt'].max()) else '',
                                "area": str(dd_area),
                                "rank": str(dd_rank),
                                "isPriority": True
                            })

                c_visits = int(cg['is_v'].sum())
                c_calls = int(cg['is_c'].sum())
                priorityCustomers.append({
                    "code": str(ccode),
                    "name": str(p_cname),
                    "visits": c_visits,
                    "calls": c_calls,
                    "designProposals": int(is_design_proposal[cg.index].sum()),
                    "total": c_visits + c_calls,
                    "lastDate": cg['dt'].max().strftime('%Y/%m/%d') if not pd.isna(cg['dt'].max()) else '',
                    "area": str(p_area),
                    "rank": str(p_rank),
                    "isPriority": True,
                    "directDeliveries": dd_list
                })
            priorityCustomers.sort(key=lambda x: x['total'], reverse=True)

        designProgress = []
        if design_status_col in mdf.columns:
            dsg_df = mdf[has_design].copy()
            if not dsg_df.empty:
                dsg_df[design_status_col] = dsg_df[design_status_col].fillna('未設定').astype(str).str.strip()
                dsg_df.loc[dsg_df[design_status_col] == '', design_status_col] = '未設定'
                for st, sg in dsg_df.groupby(design_status_col):
                    designProgress.append({"status": str(st), "count": len(sg)})
            designProgress.sort(key=lambda x: x['count'], reverse=True)

        def get_top_customers(action_mask):
            top = []
            tdf = mdf[action_mask & mdf[customer_name_col].notna()].copy()
            if not tdf.empty and customer_name_col in tdf.columns:
                tdf[customer_name_col] = tdf[customer_name_col].astype(str).str.strip()
                tdf = tdf[tdf[customer_name_col] != '']
                for cname, cg in tdf.groupby(customer_name_col):
                    details = []
                    if dd_name_col in cg.columns:
                        cg_dd = cg.copy()
                        cg_dd[dd_name_col] = cg_dd[dd_name_col].fillna('(直接)').astype(str).str.strip()
                        cg_dd.loc[cg_dd[dd_name_col] == '', dd_name_col] = '(直接)'
                        for ddn, ddg in cg_dd.groupby(dd_name_col):
                            details.append({"name": str(ddn), "count": len(ddg)})
                        details.sort(key=lambda x: x['count'], reverse=True)
                    top.append({"name": str(cname), "count": len(cg), "details": details})
                top.sort(key=lambda x: x['count'], reverse=True)
            return top[:10]

        topCustomers = get_top_customers(mdf['is_v'])
        topCallCustomers = get_top_customers(mdf['is_c'])

        dailyActivity = []
        mdf['date_str'] = mdf['dt'].dt.strftime('%Y/%m/%d')
        for d, dg in mdf.groupby('date_str'):
            day_acts = []
            for _, row in dg.iterrows():
                def clean_val(val):
                    if pd.isna(val) or str(val).strip() == '' or str(val).strip().lower() == 'nan' or str(val).strip() == '-':
                        return None
                    return str(val).strip()
                
                cname = clean_val(row.get(customer_name_col))
                action = clean_val(row.get(action_col))
                
                dd_name = None
                if dd_name_col in row:
                    dd_name = clean_val(row.get(dd_name_col))
                
                is_p = False
                if priority_col in row:
                    p_val = clean_val(row.get(priority_col))
                    is_p = p_val is not None
                
                biz_content = None
                for col in ['商談内容', '詳細', '商談詳細', '内容']:
                    if col in row:
                        biz_content = clean_val(row.get(col))
                        if biz_content:
                            break
                            
                design_no = None
                if 'デザイン依頼No.' in row:
                    design_no = clean_val(row.get('デザイン依頼No.'))
                elif 'デザイン依頼No' in row:
                    design_no = clean_val(row.get('デザイン依頼No'))
                
                if design_no:
                    design_no = design_no.replace('.0', '').strip()
                
                design_name = None
                if 'デザイン名' in row:
                    design_name = clean_val(row.get('デザイン名'))
                
                design_status = None
                if design_status_col in row:
                    design_status = clean_val(row.get(design_status_col))
                
                area = None
                if area_col in row:
                    area = clean_val(row.get(area_col))
                
                day_acts.append({
                    "customer_name": cname,
                    "action": action,
                    "dd_name": dd_name,
                    "is_priority": is_p,
                    "business_content": biz_content,
                    "design_no": design_no,
                    "design_name": design_name,
                    "design_status": design_status,
                    "area": area
                })

            dailyActivity.append({
                "date": str(d),
                "visits": int(dg['is_v'].sum()),
                "calls": int(dg['is_c'].sum()),
                "activities": day_acts
            })
        dailyActivity.sort(key=lambda x: x['date'])

        return {
            "totalReports": len(mdf),
            "totalVisits": totalVisits,
            "totalCalls": totalCalls,
            "priorityVisits": priorityVisits,
            "priorityCalls": priorityCalls,
            "totalDesignProposals": totalDesignProposals,
            "totalDesignCompleted": totalDesignCompleted,
            "totalDesignRejected": totalDesignRejected,
            "uniqueCustomers": uniqueCustomers,
            "activeDays": activeDays,
            "areaBreakdown": areaBreakdown,
            "priorityCustomers": priorityCustomers,
            "designProgress": designProgress,
            "topCustomers": topCustomers,
            "topCallCustomers": topCallCustomers,
            "dailyActivity": dailyActivity
        }
    except Exception as e:
        logging.error(f"Monthly summary error: {e}")
        import traceback
        traceback.print_exc()
        return empty_response



@router.get("/api/analytics/points-table")
def get_points_table(target_months_count: int = 7):
    """
    全メンバーの日報点数表を集計する。
    """
    if not os.path.exists(config.EXCEL_DIR):
        raise HTTPException(status_code=500, detail="Excel directory not found")

    EXCLUDE_FILES = [
        "●20260117_2026年度用_日報【原本_2】.xlsm",
        "【電話営業資料】電話　メール話題情報.xlsx",
        "支店999_【津田_操作確認用】_2026年度用日報.xlsm",
        "本社001_【中野次長】_2026年度用日報.xlsm"
    ]

    try:
        target_files = [
            f for f in os.listdir(config.EXCEL_DIR) 
            if f.endswith('.xlsm') and f not in EXCLUDE_FILES and not f.startswith('~$')
        ]
        
        # 2026年度の月リスト（2月〜翌1月）
        month_keys = [
            ("26/02", "2月"), ("26/03", "3月"), ("26/04", "4月"), ("26/05", "5月"),
            ("26/06", "6月"), ("26/07", "7月"), ("26/08", "8月"), ("26/09", "9月"),
            ("26/10", "10月"), ("26/11", "11月"), ("26/12", "12月"), ("27/01", "1月")
        ]

        def extract_staff_name_py(filename: str) -> str:
            import re
            match = re.search(r'【(.+?)】', filename)
            if not match: return filename.replace('.xlsm', '')
            content = match.group(1)
            name_with_paren = re.search(r'^(.+?)（(.+?)）', content)
            if name_with_paren: return name_with_paren.group(1) + name_with_paren.group(2)
            surname = re.search(r'^([^\s\u4e00-\u9fa5]*[\u4e00-\u9fa5]+?)(?:課長|次長|部長|常務|社長|主任|係長|専務|取締役|マネージャー|リーダー|担当|氏)?$', content)
            if surname: return surname.group(1)
            return content[:4]

        # 柔軟な日付解析ヘルパー
        def parse_dt(x):
            if pd.isna(x): return pd.NaT
            s = str(x).strip()
            if s == '' or s == 'nan' or s == '-': return pd.NaT
            try:
                return pd.to_datetime(s, format='%y/%m/%d')
            except:
                pass
            try:
                return pd.to_datetime(s, format='%Y/%m/%d')
            except:
                pass
            try:
                return pd.to_datetime(s, format='%Y-%m-%d')
            except:
                pass
            try:
                return pd.to_datetime(s, errors='coerce')
            except:
                return pd.NaT

        records = []

        for filename in target_files:
            try:
                staff_name = extract_staff_name_py(filename)
                
                # 1. 重点件数（得意先_Listから個人の重点顧客の数を数える）
                priority_count = 0
                try:
                    cust_df = cache.get_cached_dataframe(filename, '得意先_List')
                    if cust_df is not None and not cust_df.empty:
                        cust_df.columns = [str(col).replace('\n', '').strip() for col in cust_df.columns]
                        priority_col = next((c for c in cust_df.columns if '重点' in c), None)
                        rep_col = cust_df.columns[8] if len(cust_df.columns) > 8 else None
                        
                        if priority_col and rep_col:
                            is_priority = cust_df[priority_col].astype(str).str.contains('重点', na=False)
                            
                            def match_representative(rep_val, target_name):
                                if pd.isna(rep_val): return False
                                rep_val = str(rep_val).strip()
                                target_name = str(target_name).strip()
                                if rep_val == target_name: return True
                                
                                import re
                                def clean_name(n):
                                    n = re.sub(r'[\(（].*?[\)）]', '', n)
                                    n = re.sub(r'(?:次長|課長|部長|常務|専務|社長|主任|係長|取締役|マネージャー|マネ|リーダー|担当)$', '', n)
                                    return n.strip()
                                    
                                cleaned_rep = clean_name(rep_val)
                                cleaned_staff = clean_name(target_name)
                                if cleaned_rep == cleaned_staff: return True
                                
                                match = re.search(r'(.+?)[（\(](.+?)[）\)]', rep_val)
                                if match:
                                    base = match.group(1).strip()
                                    inside = match.group(2).strip()
                                    combined = base + inside
                                    if clean_name(combined) == cleaned_staff: return True
                                return False
                                
                            matches_rep = cust_df[rep_col].apply(lambda x: match_representative(x, staff_name))
                            priority_count = int(cust_df[is_priority & matches_rep].shape[0])
                        elif priority_col:
                            # フォールバック: 担当者列が取れなければ全体の重点顧客数をカウント
                            priority_count = int(cust_df[cust_df[priority_col].astype(str).str.contains('重点', na=False)].shape[0])
                except Exception as cust_err:
                    logging.warning(f"Error reading priority count for {filename}: {cust_err}")

                # 2. 日報データの読み込み
                df = cache.get_cached_dataframe(filename, '営業日報')
                if df is None or df.empty:
                    # 空白レコードを追加
                    monthly_data = {label: {"priority_calls": 0, "general_calls": 0, "priority_visits": 0, "general_visits": 0} for _, label in month_keys}
                    records.append({
                        "staff": staff_name,
                        "priority_count": priority_count,
                        "monthly_data": monthly_data,
                        "totals": {"priority_calls": 0, "general_calls": 0, "total_calls": 0, "priority_visits": 0, "general_visits": 0, "total_visits": 0},
                        "points": 0.0,
                        "achievement_rate": 0.0,
                        "rating": 0.0,
                        "file": filename
                    })
                    continue

                # カラムクリーンアップ
                df.columns = [str(col).replace('\n', '').strip() for col in df.columns]
                
                date_col = next((c for c in df.columns if '日付' in c), '日付')
                action_col = next((c for c in df.columns if '行動内容' in c), '行動内容')
                priority_col = next((c for c in df.columns if '重点' in c), '重点顧客')

                df['dt'] = df[date_col].apply(parse_dt)
                df = df.dropna(subset=['dt'])

                # 各項目のフラグ付け
                def is_priority(val):
                    if pd.isnull(val): return False
                    s = str(val).strip()
                    return s != '' and s != '-' and '重点' in s

                df['is_priority_cust'] = df[priority_col].apply(is_priority) if priority_col in df.columns else False
                df['is_visit'] = df[action_col].apply(lambda x: True if pd.notnull(x) and '訪問' in str(x) else False)
                df['is_call'] = df[action_col].apply(lambda x: True if pd.notnull(x) and '電話' in str(x) else False)

                # 月ごとの集計
                monthly_data = {}
                tot_p_calls = 0
                tot_g_calls = 0
                tot_p_visits = 0
                tot_g_visits = 0

                for m_key, m_label in month_keys:
                    parts = m_key.split('/')
                    iso_month = f"20{parts[0]}-{parts[1]}"
                    mask1 = df['dt'].dt.strftime('%y/%m') == m_key
                    mask2 = df['dt'].dt.strftime('%Y-%m') == iso_month
                    m_df = df[mask1 | mask2]

                    if m_df.empty:
                        monthly_data[m_label] = {
                            "priority_calls": 0,
                            "general_calls": 0,
                            "priority_visits": 0,
                            "general_visits": 0
                        }
                        continue

                    p_calls = int((m_df['is_call'] & m_df['is_priority_cust']).sum())
                    g_calls = int((m_df['is_call'] & ~m_df['is_priority_cust']).sum())
                    p_visits = int((m_df['is_visit'] & m_df['is_priority_cust']).sum())
                    g_visits = int((m_df['is_visit'] & ~m_df['is_priority_cust']).sum())

                    monthly_data[m_label] = {
                        "priority_calls": p_calls,
                        "general_calls": g_calls,
                        "priority_visits": p_visits,
                        "general_visits": g_visits
                    }

                    tot_p_calls += p_calls
                    tot_g_calls += g_calls
                    tot_p_visits += p_visits
                    tot_g_visits += g_visits

                # 点数計算: 重点電話 + 一般電話/2 + 重点訪問*10 + 一般訪問*3
                points = float(tot_p_calls) + (float(tot_g_calls) / 2.0) + (float(tot_p_visits) * 10.0) + (float(tot_g_visits) * 3.0)
                
                # 目標値 (月200点 * target_months_count)
                target_score = float(200 * target_months_count)
                achievement_rate = (points / target_score) * 100.0 if target_score > 0 else 0.0
                rating = points / 140.0

                records.append({
                    "staff": staff_name,
                    "priority_count": priority_count,
                    "monthly_data": monthly_data,
                    "totals": {
                        "priority_calls": tot_p_calls,
                        "general_calls": tot_g_calls,
                        "total_calls": tot_p_calls + tot_g_calls,
                        "priority_visits": tot_p_visits,
                        "general_visits": tot_g_visits,
                        "total_visits": tot_p_visits + tot_g_visits
                    },
                    "points": round(points, 1),
                    "achievement_rate": round(achievement_rate, 1),
                    "rating": round(rating, 1),
                    "file": filename
                })

            except Exception as file_err:
                logging.warning(f"Error processing points table for {filename}: {file_err}")
                continue

        return {
            "months": [m_label for _, m_label in month_keys],
            "records": records,
            "target_months_count": target_months_count
        }

    except Exception as e:
        logging.error(f"Error generating points table: {e}")
        raise HTTPException(status_code=500, detail=str(e))



@router.get("/api/analytics/export/points-table")
def export_points_table_excel(target_months_count: int = 7):
    """
    日報点数表の美しいExcelファイルをエクスポートする。
    """
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    import io
    from fastapi.responses import StreamingResponse

    # 1. データの取得
    try:
        data = get_points_table(target_months_count=target_months_count)
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Failed to gather data: {e}")

    records = data["records"]
    months = data["months"]

    # 実績がある月のみ抽出
    active_months = []
    for m in months:
        has_act = False
        for r in records:
            act = r["monthly_data"].get(m, {})
            if (act.get("priority_calls", 0) > 0 or 
                act.get("general_calls", 0) > 0 or 
                act.get("priority_visits", 0) > 0 or 
                act.get("general_visits", 0) > 0):
                has_act = True
                break
        if has_act:
            active_months.append(m)

    if not active_months:
        active_months = ["2月", "3月", "4月", "5月"]

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "日報点数表"
    ws.views.sheetView[0].showGridLines = True

    # 印刷設定: A4横1枚に収める
    ws.page_setup.orientation = "landscape"
    ws.page_setup.paperSize = 9 # 9 = A4
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 1

    # 余白設定 (狭めにして印刷領域を最大化)
    ws.page_margins.left = 0.25
    ws.page_margins.right = 0.25
    ws.page_margins.top = 0.3
    ws.page_margins.bottom = 0.3
    ws.page_margins.header = 0.1
    ws.page_margins.footer = 0.1

    # スタイル定義
    font_title = Font(name="Meiryo UI", size=16, bold=True)
    font_header = Font(name="Meiryo UI", size=9, bold=True)
    font_sub_header = Font(name="Meiryo UI", size=8, bold=True)
    font_data = Font(name="Meiryo UI", size=9)
    font_bold = Font(name="Meiryo UI", size=9, bold=True)

    fill_blue_header = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    fill_gray_header = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    fill_sub_blue = PatternFill(start_color="F2F5F9", end_color="F2F5F9", fill_type="solid")
    fill_yellow = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
    fill_light_blue = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")
    fill_green = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")

    border_thin_side = Side(border_style="thin", color="D9D9D9")
    border_thin = Border(left=border_thin_side, right=border_thin_side, top=border_thin_side, bottom=border_thin_side)
    border_double_bottom = Border(top=Side(border_style="thin", color="000000"), bottom=Side(border_style="double", color="000000"))

    align_center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    align_left = Alignment(horizontal="left", vertical="center")
    align_right = Alignment(horizontal="right", vertical="center")

    # タイトル行
    ws.merge_cells("A1:C1")
    ws["A1"] = "日報点数表 2026"
    ws["A1"].font = font_title
    ws["A1"].alignment = align_left
    ws.row_dimensions[1].height = 30

    # ヘッダー作成
    ws.row_dimensions[3].height = 24
    ws.row_dimensions[4].height = 20

    ws.merge_cells("A3:A4")
    ws["A3"] = "営業名"
    ws["A3"].font = font_header
    ws["A3"].fill = fill_gray_header
    ws["A3"].alignment = align_center
    ws["A3"].border = border_thin
    ws["A4"].border = border_thin

    ws.merge_cells("B3:B4")
    ws["B3"] = "重点件数"
    ws["B3"].font = font_header
    ws["B3"].fill = fill_gray_header
    ws["B3"].alignment = align_center
    ws["B3"].border = border_thin
    ws["B4"].border = border_thin

    current_col = 3

    # 各月のヘッダー
    for m in active_months:
        col_letter_start = get_column_letter(current_col)
        col_letter_end = get_column_letter(current_col + 3)
        ws.merge_cells(f"{col_letter_start}3:{col_letter_end}3")
        cell_ref = f"{col_letter_start}3"
        ws[cell_ref] = m
        ws[cell_ref].font = font_header
        ws[cell_ref].fill = fill_blue_header
        ws[cell_ref].alignment = align_center

        sub_headers = ["重点電話", "一般電話", "重点訪問", "一般訪問"]
        for idx, sub in enumerate(sub_headers):
            c_ref = f"{get_column_letter(current_col + idx)}4"
            ws[c_ref] = sub
            ws[c_ref].font = font_sub_header
            ws[c_ref].fill = fill_sub_blue
            ws[c_ref].alignment = align_center
            ws[c_ref].border = border_thin
            ws[f"{get_column_letter(current_col + idx)}3"].border = border_thin
            
        current_col += 4

    # 計ヘッダー
    tot_col_start = current_col
    ws.merge_cells(f"{get_column_letter(current_col)}3:{get_column_letter(current_col + 5)}3")
    cell_ref = f"{get_column_letter(current_col)}3"
    ws[cell_ref] = "計"
    ws[cell_ref].font = font_header
    ws[cell_ref].fill = fill_gray_header
    ws[cell_ref].alignment = align_center

    sub_tots = ["重点電話総数", "一般電話総数", "総電話件数", "重点訪問総数", "一般訪問総数", "総訪問件数"]
    for idx, sub in enumerate(sub_tots):
        c_ref = f"{get_column_letter(current_col + idx)}4"
        ws[c_ref] = sub
        ws[c_ref].font = font_sub_header
        ws[c_ref].fill = fill_gray_header
        ws[c_ref].alignment = align_center
        ws[c_ref].border = border_thin
        ws[f"{get_column_letter(current_col + idx)}3"].border = border_thin
        
    current_col += 6

    # 各種得点ヘッダー
    points_col_idx = current_col
    ws.merge_cells(f"{get_column_letter(current_col)}3:{get_column_letter(current_col)}4")
    ws[f"{get_column_letter(current_col)}3"] = "点数"
    ws[f"{get_column_letter(current_col)}3"].font = font_header
    ws[f"{get_column_letter(current_col)}3"].fill = fill_yellow
    ws[f"{get_column_letter(current_col)}3"].alignment = align_center
    ws[f"{get_column_letter(current_col)}3"].border = border_thin
    ws[f"{get_column_letter(current_col)}4"].border = border_thin
    current_col += 1

    ach_col_idx = current_col
    ws.merge_cells(f"{get_column_letter(current_col)}3:{get_column_letter(current_col)}4")
    ws[f"{get_column_letter(current_col)}3"] = f"月200点×{target_months_count}\n({200 * target_months_count})"
    ws[f"{get_column_letter(current_col)}3"].font = font_header
    ws[f"{get_column_letter(current_col)}3"].fill = fill_light_blue
    ws[f"{get_column_letter(current_col)}3"].alignment = align_center
    ws[f"{get_column_letter(current_col)}3"].border = border_thin
    ws[f"{get_column_letter(current_col)}4"].border = border_thin
    current_col += 1

    rat_col_idx = current_col
    ws.merge_cells(f"{get_column_letter(current_col)}3:{get_column_letter(current_col)}4")
    ws[f"{get_column_letter(current_col)}3"] = "評価点"
    ws[f"{get_column_letter(current_col)}3"].font = font_header
    ws[f"{get_column_letter(current_col)}3"].fill = fill_green
    ws[f"{get_column_letter(current_col)}3"].alignment = align_center
    ws[f"{get_column_letter(current_col)}3"].border = border_thin
    ws[f"{get_column_letter(current_col)}4"].border = border_thin

    # データ行書き込み
    row_num = 5
    for r in records:
        ws.row_dimensions[row_num].height = 18

        ws[f"A{row_num}"] = r["staff"]
        ws[f"A{row_num}"].font = font_bold
        ws[f"A{row_num}"].alignment = align_left
        ws[f"A{row_num}"].border = border_thin

        ws[f"B{row_num}"] = r["priority_count"]
        ws[f"B{row_num}"].font = font_data
        ws[f"B{row_num}"].alignment = align_center
        ws[f"B{row_num}"].border = border_thin

        col_idx = 3
        for m in active_months:
            act = r["monthly_data"].get(m, {"priority_calls": 0, "general_calls": 0, "priority_visits": 0, "general_visits": 0})
            for key in ["priority_calls", "general_calls", "priority_visits", "general_visits"]:
                cell_ref = f"{get_column_letter(col_idx)}{row_num}"
                ws[cell_ref] = act.get(key, 0)
                ws[cell_ref].font = font_data
                ws[cell_ref].alignment = align_right
                ws[cell_ref].border = border_thin
                col_idx += 1

        # 各種集計用計算式（数式によるリアルタイム計算）
        # 1. 重点電話総数
        p_calls_parts = [f"{get_column_letter(3 + m_idx * 4)}{row_num}" for m_idx in range(len(active_months))]
        ws[f"{get_column_letter(tot_col_start)}{row_num}"] = f"={'+'.join(p_calls_parts)}"
        ws[f"{get_column_letter(tot_col_start)}{row_num}"].font = font_bold
        ws[f"{get_column_letter(tot_col_start)}{row_num}"].alignment = align_right
        ws[f"{get_column_letter(tot_col_start)}{row_num}"].border = border_thin

        # 2. 電話総件数
        g_calls_parts = [f"{get_column_letter(4 + m_idx * 4)}{row_num}" for m_idx in range(len(active_months))]
        ws[f"{get_column_letter(tot_col_start + 1)}{row_num}"] = f"={'+'.join(g_calls_parts)}"
        ws[f"{get_column_letter(tot_col_start + 1)}{row_num}"].font = font_bold
        ws[f"{get_column_letter(tot_col_start + 1)}{row_num}"].alignment = align_right
        ws[f"{get_column_letter(tot_col_start + 1)}{row_num}"].border = border_thin

        # 3. 総電話件数 (重点＋一般)
        p_c_col = get_column_letter(tot_col_start)
        g_c_col = get_column_letter(tot_col_start + 1)
        ws[f"{get_column_letter(tot_col_start + 2)}{row_num}"] = f"={p_c_col}{row_num}+{g_c_col}{row_num}"
        ws[f"{get_column_letter(tot_col_start + 2)}{row_num}"].font = font_bold
        ws[f"{get_column_letter(tot_col_start + 2)}{row_num}"].alignment = align_right
        ws[f"{get_column_letter(tot_col_start + 2)}{row_num}"].border = border_thin

        # 4. 重点訪問総数
        p_visits_parts = [f"{get_column_letter(5 + m_idx * 4)}{row_num}" for m_idx in range(len(active_months))]
        ws[f"{get_column_letter(tot_col_start + 3)}{row_num}"] = f"={'+'.join(p_visits_parts)}"
        ws[f"{get_column_letter(tot_col_start + 3)}{row_num}"].font = font_bold
        ws[f"{get_column_letter(tot_col_start + 3)}{row_num}"].alignment = align_right
        ws[f"{get_column_letter(tot_col_start + 3)}{row_num}"].border = border_thin

        # 5. 訪問総件数
        g_visits_parts = [f"{get_column_letter(6 + m_idx * 4)}{row_num}" for m_idx in range(len(active_months))]
        ws[f"{get_column_letter(tot_col_start + 4)}{row_num}"] = f"={'+'.join(g_visits_parts)}"
        ws[f"{get_column_letter(tot_col_start + 4)}{row_num}"].font = font_bold
        ws[f"{get_column_letter(tot_col_start + 4)}{row_num}"].alignment = align_right
        ws[f"{get_column_letter(tot_col_start + 4)}{row_num}"].border = border_thin

        # 6. 総訪問件数 (重点＋一般)
        p_v_col = get_column_letter(tot_col_start + 3)
        g_v_col = get_column_letter(tot_col_start + 4)
        ws[f"{get_column_letter(tot_col_start + 5)}{row_num}"] = f"={p_v_col}{row_num}+{g_v_col}{row_num}"
        ws[f"{get_column_letter(tot_col_start + 5)}{row_num}"].font = font_bold
        ws[f"{get_column_letter(tot_col_start + 5)}{row_num}"].alignment = align_right
        ws[f"{get_column_letter(tot_col_start + 5)}{row_num}"].border = border_thin

        #活動得点計算式: 重点電話×1 ＋ 一般電話/2 ＋ 重点訪問×10 ＋ 一般訪問×3
        points_formula = f"={p_c_col}{row_num}+({g_c_col}{row_num}/2)+({p_v_col}{row_num}*10)+({g_v_col}{row_num}*3)"
        ws[f"{get_column_letter(points_col_idx)}{row_num}"] = points_formula
        ws[f"{get_column_letter(points_col_idx)}{row_num}"].font = font_bold
        ws[f"{get_column_letter(points_col_idx)}{row_num}"].alignment = align_right
        ws[f"{get_column_letter(points_col_idx)}{row_num}"].border = border_thin
        ws[f"{get_column_letter(points_col_idx)}{row_num}"].fill = fill_yellow
        ws[f"{get_column_letter(points_col_idx)}{row_num}"].number_format = "0.0"

        # 達成率: 総合得点 ÷ (月200点×対象月数)
        target_score = 200 * target_months_count
        pts_col_let = get_column_letter(points_col_idx)
        ach_formula = f"={pts_col_let}{row_num}/{target_score}"
        ws[f"{get_column_letter(ach_col_idx)}{row_num}"] = ach_formula
        ws[f"{get_column_letter(ach_col_idx)}{row_num}"].font = font_bold
        ws[f"{get_column_letter(ach_col_idx)}{row_num}"].alignment = align_right
        ws[f"{get_column_letter(ach_col_idx)}{row_num}"].border = border_thin
        ws[f"{get_column_letter(ach_col_idx)}{row_num}"].fill = fill_light_blue
        ws[f"{get_column_letter(ach_col_idx)}{row_num}"].number_format = "0.0%"

        # 評価点: 総合得点 ÷ 140.0
        rat_formula = f"={pts_col_let}{row_num}/140.0"
        ws[f"{get_column_letter(rat_col_idx)}{row_num}"] = rat_formula
        ws[f"{get_column_letter(rat_col_idx)}{row_num}"].font = font_bold
        ws[f"{get_column_letter(rat_col_idx)}{row_num}"].alignment = align_right
        ws[f"{get_column_letter(rat_col_idx)}{row_num}"].border = border_thin
        ws[f"{get_column_letter(rat_col_idx)}{row_num}"].fill = fill_green
        ws[f"{get_column_letter(rat_col_idx)}{row_num}"].number_format = "0.0"

        row_num += 1

    # 「計」行の作成
    ws.row_dimensions[row_num].height = 20
    ws[f"A{row_num}"] = "計"
    ws[f"A{row_num}"].font = font_bold
    ws[f"A{row_num}"].alignment = align_left
    ws[f"A{row_num}"].border = border_thin

    ws[f"B{row_num}"] = f"=SUM(B5:B{row_num - 1})"
    ws[f"B{row_num}"].font = font_bold
    ws[f"B{row_num}"].alignment = align_center
    ws[f"B{row_num}"].border = border_thin

    for c_idx in range(3, rat_col_idx + 1):
        col_let = get_column_letter(c_idx)
        if c_idx == ach_col_idx:
            tot_pts_cell = f"{get_column_letter(points_col_idx)}{row_num}"
            ws[f"{col_let}{row_num}"] = f"={tot_pts_cell}/{200 * target_months_count}"
            ws[f"{col_let}{row_num}"].number_format = "0.0%"
            ws[f"{col_let}{row_num}"].fill = fill_light_blue
        elif c_idx == rat_col_idx:
            tot_pts_cell = f"{get_column_letter(points_col_idx)}{row_num}"
            ws[f"{col_let}{row_num}"] = f"={tot_pts_cell}/140.0"
            ws[f"{col_let}{row_num}"].number_format = "0.0"
            ws[f"{col_let}{row_num}"].fill = fill_green
        elif c_idx == points_col_idx:
            ws[f"{col_let}{row_num}"] = f"=SUM({col_let}5:{col_let}{row_num - 1})"
            ws[f"{col_let}{row_num}"].number_format = "0.0"
            ws[f"{col_let}{row_num}"].fill = fill_yellow
        else:
            ws[f"{col_let}{row_num}"] = f"=SUM({col_let}5:{col_let}{row_num - 1})"

        ws[f"{col_let}{row_num}"].font = font_bold
        ws[f"{col_let}{row_num}"].alignment = align_right
        ws[f"{col_let}{row_num}"].border = border_thin

    row_num += 1

    # 「平均点」行の作成
    ws.row_dimensions[row_num].height = 20
    ws[f"A{row_num}"] = "平均点"
    ws[f"A{row_num}"].font = font_bold
    ws[f"A{row_num}"].alignment = align_left
    ws[f"A{row_num}"].border = border_double_bottom

    ws[f"B{row_num}"] = f"=AVERAGE(B5:B{row_num - 2})"
    ws[f"B{row_num}"].font = font_bold
    ws[f"B{row_num}"].alignment = align_center
    ws[f"B{row_num}"].border = border_double_bottom
    ws[f"B{row_num}"].number_format = "0.0"

    for c_idx in range(3, rat_col_idx + 1):
        col_let = get_column_letter(c_idx)
        if c_idx == ach_col_idx:
            avg_pts_cell = f"{get_column_letter(points_col_idx)}{row_num}"
            ws[f"{col_let}{row_num}"] = f"={avg_pts_cell}/{200 * target_months_count}"
            ws[f"{col_let}{row_num}"].number_format = "0.0%"
            ws[f"{col_let}{row_num}"].fill = fill_light_blue
        elif c_idx == rat_col_idx:
            avg_pts_cell = f"{get_column_letter(points_col_idx)}{row_num}"
            ws[f"{col_let}{row_num}"] = f"={avg_pts_cell}/140.0"
            ws[f"{col_let}{row_num}"].number_format = "0.0"
            ws[f"{col_let}{row_num}"].fill = fill_green
        elif c_idx == points_col_idx:
            ws[f"{col_let}{row_num}"] = f"=AVERAGE({col_let}5:{col_let}{row_num - 2})"
            ws[f"{col_let}{row_num}"].number_format = "0.0"
            ws[f"{col_let}{row_num}"].fill = fill_yellow
        else:
            ws[f"{col_let}{row_num}"] = f"=AVERAGE({col_let}5:{col_let}{row_num - 2})"
            ws[f"{col_let}{row_num}"].number_format = "0.0"

        ws[f"{col_let}{row_num}"].font = font_bold
        ws[f"{col_let}{row_num}"].alignment = align_right
        ws[f"{col_let}{row_num}"].border = border_double_bottom

    # 列幅の自動調整
    ws.column_dimensions["A"].width = 15
    ws.column_dimensions["B"].width = 10
    for c_idx in range(3, points_col_idx):
        ws.column_dimensions[get_column_letter(c_idx)].width = 11
    ws.column_dimensions[get_column_letter(points_col_idx)].width = 11
    ws.column_dimensions[get_column_letter(ach_col_idx)].width = 18
    ws.column_dimensions[get_column_letter(rat_col_idx)].width = 10

    # ストリーミングレスポンスで返却
    file_stream = io.BytesIO()
    wb.save(file_stream)
    file_stream.seek(0)

    headers_dict = {
        'Content-Disposition': f'attachment; filename="DailyReportPointsTable_{target_months_count}months.xlsx"'
    }
    return StreamingResponse(file_stream, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", headers=headers_dict)



@router.get("/api/analytics/export/team-summary")
def export_team_summary_excel(month: str = None):
    """
    活動集計の美しいExcelファイルをエクスポートする。
    """
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    import io
    from fastapi.responses import StreamingResponse

    try:
        data = get_team_summary(month=month)
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Failed to gather data: {e}")

    records = data["records"]

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "活動集計"
    ws.views.sheetView[0].showGridLines = True

    # 印刷設定: A4横1枚に収める
    ws.page_setup.orientation = "landscape"
    ws.page_setup.paperSize = 9 # 9 = A4
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 1

    # 余白設定 (狭めにして印刷領域を最大化)
    ws.page_margins.left = 0.25
    ws.page_margins.right = 0.25
    ws.page_margins.top = 0.3
    ws.page_margins.bottom = 0.3
    ws.page_margins.header = 0.1
    ws.page_margins.footer = 0.1

    # スタイル定義
    font_title = Font(name="Meiryo UI", size=14, bold=True)
    font_header = Font(name="Meiryo UI", size=9, bold=True)
    font_data = Font(name="Meiryo UI", size=9)
    font_bold = Font(name="Meiryo UI", size=9, bold=True)

    fill_header = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    fill_blue_total = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")

    border_thin_side = Side(border_style="thin", color="D9D9D9")
    border_thin = Border(left=border_thin_side, right=border_thin_side, top=border_thin_side, bottom=border_thin_side)
    border_double_bottom = Border(top=Side(border_style="thin", color="000000"), bottom=Side(border_style="double", color="000000"))

    align_center = Alignment(horizontal="center", vertical="center")
    align_left = Alignment(horizontal="left", vertical="center")
    align_right = Alignment(horizontal="right", vertical="center")

    # タイトル行
    ws["A1"] = f"活動集計 ({month})" if month else "活動集計"
    ws["A1"].font = font_title
    ws["A1"].alignment = align_left
    ws.row_dimensions[1].height = 24

    # ヘッダー列
    headers = ["担当者", "エリア", "区分", "訪問件数", "電話件数", "合計", "ファイル名"]
    ws.row_dimensions[3].height = 20
    for idx, h in enumerate(headers):
        col_let = get_column_letter(idx + 1)
        ws[f"{col_let}3"] = h
        ws[f"{col_let}3"].font = font_header
        ws[f"{col_let}3"].fill = fill_header
        ws[f"{col_let}3"].alignment = align_center if idx < 3 or idx == 6 else align_right
        ws[f"{col_let}3"].border = border_thin

    # データ行
    row_num = 4
    for r in records:
        ws.row_dimensions[row_num].height = 18

        ws[f"A{row_num}"] = r["staff"]
        ws[f"A{row_num}"].font = font_bold
        ws[f"A{row_num}"].alignment = align_left
        ws[f"A{row_num}"].border = border_thin

        ws[f"B{row_num}"] = r["area"]
        ws[f"B{row_num}"].font = font_data
        ws[f"B{row_num}"].alignment = align_center
        ws[f"B{row_num}"].border = border_thin

        ws[f"C{row_num}"] = r["category"]
        ws[f"C{row_num}"].font = font_data
        ws[f"C{row_num}"].alignment = align_center
        ws[f"C{row_num}"].border = border_thin

        ws[f"D{row_num}"] = r["visits"]
        ws[f"D{row_num}"].font = font_data
        ws[f"D{row_num}"].alignment = align_right
        ws[f"D{row_num}"].border = border_thin

        ws[f"E{row_num}"] = r["calls"]
        ws[f"E{row_num}"].font = font_data
        ws[f"E{row_num}"].alignment = align_right
        ws[f"E{row_num}"].border = border_thin

        # 合計（数式）
        ws[f"F{row_num}"] = f"=D{row_num}+E{row_num}"
        ws[f"F{row_num}"].font = font_bold
        ws[f"F{row_num}"].fill = fill_blue_total
        ws[f"F{row_num}"].alignment = align_right
        ws[f"F{row_num}"].border = border_thin

        ws[f"G{row_num}"] = r["file"]
        ws[f"G{row_num}"].font = font_data
        ws[f"G{row_num}"].alignment = align_left
        ws[f"G{row_num}"].border = border_thin

        row_num += 1

    # 合計行
    ws.row_dimensions[row_num].height = 20
    ws[f"A{row_num}"] = "合計"
    ws[f"A{row_num}"].font = font_bold
    ws[f"A{row_num}"].alignment = align_left
    ws[f"A{row_num}"].border = border_double_bottom

    ws[f"B{row_num}"].border = border_double_bottom
    ws[f"C{row_num}"].border = border_double_bottom

    ws[f"D{row_num}"] = f"=SUM(D4:D{row_num - 1})"
    ws[f"D{row_num}"].font = font_bold
    ws[f"D{row_num}"].alignment = align_right
    ws[f"D{row_num}"].border = border_double_bottom

    ws[f"E{row_num}"] = f"=SUM(E4:E{row_num - 1})"
    ws[f"E{row_num}"].font = font_bold
    ws[f"E{row_num}"].alignment = align_right
    ws[f"E{row_num}"].border = border_double_bottom

    ws[f"F{row_num}"] = f"=SUM(F4:F{row_num - 1})"
    ws[f"F{row_num}"].font = font_bold
    ws[f"F{row_num}"].fill = fill_blue_total
    ws[f"F{row_num}"].alignment = align_right
    ws[f"F{row_num}"].border = border_double_bottom

    ws[f"G{row_num}"].border = border_double_bottom

    # 列幅調整
    ws.column_dimensions["A"].width = 16
    ws.column_dimensions["B"].width = 14
    ws.column_dimensions["C"].width = 10
    ws.column_dimensions["D"].width = 12
    ws.column_dimensions["E"].width = 12
    ws.column_dimensions["F"].width = 12
    ws.column_dimensions["G"].width = 40

    file_stream = io.BytesIO()
    wb.save(file_stream)
    file_stream.seek(0)

    month_fn = month.replace('/', '') if month else "total"
    headers_dict = {
        'Content-Disposition': f'attachment; filename="ActivitySummary_{month_fn}.xlsx"'
    }
    return StreamingResponse(file_stream, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", headers=headers_dict)



