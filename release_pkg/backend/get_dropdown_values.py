import openpyxl

excel_file = '../daily_report_template.xlsm'

wb = openpyxl.load_workbook(excel_file, keep_vba=True)

# Check defined names (named ranges)
print("Defined Names in workbook:")
for name_obj in wb.defined_names:
    print(f"\nName: {name_obj}")
    name_value = wb.defined_names[name_obj]
    print(f"  Destinations: {name_value.destinations}")
    
    # Try to get the actual values if it's a dropdown list
    if name_obj in ['行動内容', '滞在時間', 'ランク', '重点flag', '提案有無', '提案進捗', 'デザイン種別', '都道府県']:
        try:
            for sheet_name, cell_range in name_value.destinations:
                print(f"  Sheet: {sheet_name}, Range: {cell_range}")
                ws = wb[sheet_name]
                
                # Parse the range
                cell_range = cell_range.replace('$', '')
                
                print(f"  Values from range {cell_range}:")
                
                # Try to read the range
                if ':' in cell_range:
                    # It's a range
                    cells = ws[cell_range]
                    values = []
                    for row in cells:
                        if isinstance(row, tuple):
                            for cell in row:
                                if cell.value:
                                    values.append(cell.value)
                        else:
                            if row.value:
                                values.append(row.value)
                    print(f"    {values}")
                else:
                    # Single cell
                    cell = ws[cell_range]
                    print(f"    {cell.value}")
        except Exception as e:
            print(f"  Error reading values: {e}")

wb.close()
