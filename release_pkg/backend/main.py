from fastapi import FastAPI, HTTPException, UploadFile, File, BackgroundTasks, Request
from fastapi.staticfiles import StaticFiles
from fastapi.responses import FileResponse, JSONResponse
from fastapi.middleware.cors import CORSMiddleware
from fastapi.exceptions import RequestValidationError
from pydantic import BaseModel, Field, field_validator
import pandas as pd
import openpyxl
from datetime import datetime, timedelta
import os
import shutil
import json
import pickle
import hashlib
from typing import Optional, List, Dict, Any


import logging
import sys
import os

def get_base_path():
    if getattr(sys, 'frozen', False):
        return os.path.dirname(sys.executable)
    return os.path.dirname(os.path.abspath(__file__))

BASE_DIR = get_base_path()

# Ensure multipart libraries are bundled by PyInstaller
try:
    import python_multipart
    import multipart
except ImportError:
    pass

# Setup logging
logging.basicConfig(
    filename='server_debug.log',
    level=logging.DEBUG,
    format='%(asctime)s - %(levelname)s - %(message)s'
)
logging.info("Server starting up...")
logging.info(f"DEBUG PATHS: BASE_DIR={BASE_DIR}")
STATIC_DIR = os.path.join(BASE_DIR, "static")
logging.info(f"DEBUG PATHS: STATIC_DIR={STATIC_DIR}")
logging.info(f"DEBUG PATHS: Static Exists={os.path.exists(STATIC_DIR)}")

logging.info(f"Loaded python_multipart: {python_multipart.__file__ if 'python_multipart' in locals() else 'Not found'}")

app = FastAPI()

# Enable CORS for frontend communication
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],  # In production, specify the frontend URL
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

# バリデーションエラーの詳細をログ出力
@app.exception_handler(RequestValidationError)
async def validation_exception_handler(request: Request, exc: RequestValidationError):
    logging.error(f"Validation error on {request.url.path}: {exc.errors()}")
    logging.error(f"Request body: {exc.body}")
    return JSONResponse(
        status_code=422,
        content={"detail": exc.errors(), "body": str(exc.body)[:500]}
    )

from fastapi import Request

@app.middleware("http")
async def strip_api_prefix(request: Request, call_next):
    if request.url.path.startswith("/api/"):
        request.scope["path"] = request.url.path[4:]
    response = await call_next(request)
    return response


# Load configuration
def load_config():
    config_path = os.path.join(BASE_DIR, 'config.json')
    default_path = r'\\Asahipack02\社内書類ｎｅｗ\01：部署別　営業部\02：営業日報\2025年度'
    
    if os.path.exists(config_path):
        try:
            with open(config_path, 'r', encoding='utf-8') as f:
                config = json.load(f)
                path = config.get('excel_dir', default_path)
                logging.info(f"Successfully loaded config. Excel Path: {path}")
                return path
        except Exception as e:
            logging.warning(f"Failed to load config.json: {e}")
            logging.info(f"Using default path: {default_path}")
            return default_path
    
    logging.info(f"config.json not found. Using default path: {default_path}")
    return default_path

    logging.info(f"config.json not found. Using default path: {default_path}")
    return default_path

EXCEL_DIR = load_config()
logging.info(f"STARTUP: Working with EXCEL_DIR: {EXCEL_DIR}")

# --- Global Sales Data Storage ---
DATA_DIR = os.path.join(BASE_DIR, 'data')
SALES_CSV_PATH = os.path.join(DATA_DIR, 'sales_data.csv')
os.makedirs(DATA_DIR, exist_ok=True)

# Global DataFrame to hold sales data
global_sales_df = None

def load_sales_data():
    """Loads sales data from CSV into global DataFrame."""
    global global_sales_df
    if not os.path.exists(SALES_CSV_PATH):
        logging.info("No existing sales data found.")
        return

    try:
        logging.info("Loading sales data from disk...")
        # Try cp932 first
        try:
            df = pd.read_csv(SALES_CSV_PATH, encoding='cp932')
        except:
            df = pd.read_csv(SALES_CSV_PATH, encoding='utf-8')
        
        # Clean columns
        df.columns = [str(col).strip() for col in df.columns]
        
        # Ensure customer code exists
        if '得意先コード' in df.columns:
            # Normalize customer code (remove decimals, convert to string)
            df['得意先コード'] = df['得意先コード'].astype(str).str.split('.').str[0]
            global_sales_df = df
            logging.info(f"Sales data loaded successfully. {len(df)} rows.")
        else:
            logging.error("Sales CSV missing '得意先コード' column.")
    except Exception as e:
        logging.error(f"Failed to load sales data: {e}")

# Load on startup
load_sales_data()

# Find a default Excel file dynamically
DEFAULT_EXCEL_FILE = "daily_report_template.xlsm" # Fallback
if os.path.exists(EXCEL_DIR):
    files = [f for f in os.listdir(EXCEL_DIR) if f.endswith('.xlsm') and not f.startswith('~$')]
    if files:
        DEFAULT_EXCEL_FILE = files[0]
        logging.info(f"Set default Excel file to: {DEFAULT_EXCEL_FILE}")
    else:
        logging.warning("No .xlsm files found in directory. Using fallback default.")


class ReportInput(BaseModel):
    model_config = {"populate_by_name": True, "extra": "ignore"}  # 余分なフィールドを無視
    
    日付: str = ""
    行動内容: str = ""
    エリア: str = ""
    得意先CD: str = ""
    直送先CD: str = ""
    訪問先名: str = ""
    直送先名: str = ""
    重点顧客: str = ""
    ランク: str = ""
    面談者: str = ""
    滞在時間: str = ""
    商談内容: str = ""
    提案物: str = ""
    次回プラン: str = ""
    競合他社情報: str = ""
    デザイン提案有無: str = ""
    デザイン種別: str = ""
    デザイン名: str = ""
    デザイン進捗状況: str = ""
    デザイン依頼No: str = Field("", alias="デザイン依頼No.")
    上長コメント: str = ""
    コメント返信欄: str = ""
    上長: str = ""
    山澄常務: str = ""
    岡本常務: str = ""
    中野次長: str = ""
    既読チェック: str = ""
    original_values: Optional[dict] = None # For optimistic locking

    @field_validator('得意先CD', '直送先CD', mode='before')
    @classmethod
    def convert_to_string(cls, v):
        if v is None:
            return ""
        return str(v)

@app.get("/health")
def read_root():
    return {"message": "Daily Report API is running", "excel_dir": EXCEL_DIR}

@app.get("/files")
def list_excel_files():
    """List all Excel files in the directory"""
    logging.debug(f"Listing files in {EXCEL_DIR}")
    if not os.path.exists(EXCEL_DIR):
         logging.error(f"Directory not found: {EXCEL_DIR}")
         raise HTTPException(status_code=500, detail=f"Excel Directory not found: {EXCEL_DIR}")
         
    try:
        files = []
        # Add timeout protection or more verbose logging? 
        # listing network drive can appear to hang.
        
        items = os.listdir(EXCEL_DIR)
        logging.debug(f"Found {len(items)} items in directory")
        
        for file in items:
            if file.endswith(('.xlsx', '.xlsm')):
                file_path = os.path.join(EXCEL_DIR, file)
                try:
                    file_size = os.path.getsize(file_path)
                    file_mtime = os.path.getmtime(file_path)
                    files.append({
                        "name": file,
                        "size": file_size,
                        "modified": datetime.fromtimestamp(file_mtime).isoformat()
                    })
                except Exception as file_err:
                    logging.warning(f"Error processing file {file}: {file_err}")
                    continue
                    
        return {"files": files, "default": DEFAULT_EXCEL_FILE}
    except Exception as e:
        logging.critical(f"CRITICAL ERROR in list_excel_files: {str(e)}")
        import traceback
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=f"Failed to list files: {str(e)}")


# Cache for Excel dataframes: {(filename, sheet_name): {'mtime': float, 'df': pd.DataFrame}}
CACHE = {}

def create_backup(file_path):
    try:
        backup_dir = os.path.join(os.path.dirname(file_path), 'backup')
        if not os.path.exists(backup_dir):
            os.makedirs(backup_dir)
        
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        filename = os.path.basename(file_path)
        name, ext = os.path.splitext(filename)
        backup_filename = f"{name}_{timestamp}{ext}"
        backup_path = os.path.join(backup_dir, backup_filename)
        
        shutil.copy2(file_path, backup_path)
        logging.info(f"Backup created: {backup_path}")
    except Exception as e:
        logging.warning(f"Failed to create backup: {e}")


def get_cached_dataframe(filename: str, sheet_name: str) -> pd.DataFrame:
    """
    Get dataframe from cache or read from Excel file if modified or not in cache.
    """
    excel_file = os.path.join(EXCEL_DIR, filename)
    
    if not os.path.exists(excel_file):
        logging.error(f"File not found: {excel_file}")
        raise HTTPException(status_code=404, detail=f"Excel file '{filename}' not found at {excel_file}")
    
    current_mtime = os.path.getmtime(excel_file)
    cache_key = (filename, sheet_name)
    
    # --- In-Memory Cache Check ---
    if cache_key in CACHE:
        cached_data = CACHE[cache_key]
        if cached_data['mtime'] == current_mtime:
            return cached_data['df'].copy() # Return copy to prevent mutation of cached data

    # --- Disk Cache Check ---
    # Create cache directory if needed
    CACHE_DIR = os.path.join(os.path.dirname(os.path.abspath(__file__)), ".cache")
    if not os.path.exists(CACHE_DIR):
        os.makedirs(CACHE_DIR)
        
    # Create unique cache filename based on file path and sheet
    cache_id = hashlib.md5(f"{filename}_{sheet_name}".encode('utf-8')).hexdigest()
    cache_path = os.path.join(CACHE_DIR, f"{cache_id}.pkl")
    
    try:
        if os.path.exists(cache_path):
            with open(cache_path, 'rb') as f:
                disk_cache = pickle.load(f)
            
            # Use disk cache if timestamp matches
            if disk_cache.get('mtime') == current_mtime:
                logging.debug(f"Loaded {filename} ({sheet_name}) from disk cache")
                df = disk_cache['df']
                # Update in-memory cache
                CACHE[cache_key] = {'mtime': current_mtime, 'df': df}
                return df.copy()
    except Exception as e:
        logging.warning(f"Failed to load from disk cache: {e}")

    # --- Read from Excel (Expensive Operation) ---
    try:
        logging.debug(f"Reading Excel {excel_file}, sheet={sheet_name}")
        df = pd.read_excel(excel_file, sheet_name=sheet_name, header=0)
        
        # Update in-memory cache
        CACHE[cache_key] = {'mtime': current_mtime, 'df': df}
        
        # Update disk cache
        try:
            with open(cache_path, 'wb') as f:
                pickle.dump({'mtime': current_mtime, 'df': df}, f)
            logging.debug(f"Saved {filename} ({sheet_name}) to disk cache")
        except Exception as e:
            logging.warning(f"Failed to save disk cache: {e}")
            
        return df.copy()
    except Exception as e:
        logging.error(f"Reading Excel failed: {e}")
        import traceback
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=f"Error reading Excel file: {str(e)}")


@app.get("/customers")
def get_customers(filename: str = DEFAULT_EXCEL_FILE):
    """Get customer list from the Excel file"""
    try:
        # Get dataframe from cache
        df = get_cached_dataframe(filename, '得意先_List')
        
        # Clean up column names
        df.columns = [str(col).replace('\n', '').strip() for col in df.columns]
        
        # Rename specific columns
        df = df.rename(columns={
            '得意先CD.': '得意先CD',
            '直送先CD.': '直送先CD',
        })
        
        # Fill NaN values with empty strings
        df = df.fillna(value='')
        
        # Convert to dict
        records = df.to_dict(orient="records")
        
        # Clean the records
        import math
        cleaned_records = []
        for record in records:
            cleaned_record = {}
            for key, value in record.items():
                if isinstance(value, float):
                    if math.isnan(value) or math.isinf(value):
                        cleaned_record[key] = None
                    else:
                        cleaned_record[key] = value
                elif value == '':
                    cleaned_record[key] = None
                elif isinstance(value, str):
                    import re
                    cleaned_value = re.sub(r'_x000D_', '\n', value)
                    cleaned_value = cleaned_value.replace('\r', '')
                    cleaned_record[key] = cleaned_value
                else:
                    cleaned_record[key] = value
            cleaned_records.append(cleaned_record)

        return cleaned_records
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.get("/priority-customers")
def get_priority_customers(filename: str = DEFAULT_EXCEL_FILE):
    """得意先_Listからカラム H (重点顧客) が「重点」の顧客を取得。カラム I の担当者情報も含める"""
    try:
        # 得意先_Listを読み込み
        df = get_cached_dataframe(filename, '得意先_List')
        
        # カラム名をクリーンアップ
        df.columns = [str(col).replace('\n', '').strip() for col in df.columns]
        
        # カラム名を保存
        col_customer_cd = df.columns[0]  # 得意先CD
        col_customer_name = df.columns[1] if len(df.columns) > 1 else None  # 得意先名
        col_priority = df.columns[7] if len(df.columns) > 7 else None  # カラムH: 重点顧客
        col_staff = df.columns[8] if len(df.columns) > 8 else None  # カラムI: 担当者
        
        # 得意先CDがある行のみ抽出（ヘッダー行や空行を除外）
        df = df.dropna(subset=[col_customer_cd])
        
        # カラム H の名前を特定
        priority_col = col_priority
        if not priority_col:
            # フォールバック: 「重点顧客」という名前のカラムを探す
            for col in df.columns:
                if '重点' in str(col):
                    priority_col = col
                    break
            if not priority_col:
                logging.warning(f"Priority column not found. Columns: {list(df.columns)}")
                return []
        
        logging.info(f"Priority column: {priority_col}, Staff column: {col_staff}")
        
        # 「重点」と記載されている行のみ抽出
        priority_df = df[df[priority_col].astype(str).str.contains('重点', na=False)]
        
        logging.info(f"Found {len(priority_df)} priority customers")
        
        # レコードを作成
        records = []
        for _, row in priority_df.iterrows():
            customer_cd = row[col_customer_cd]
            customer_name = row[col_customer_name] if col_customer_name else ''
            staff = row[col_staff] if col_staff else ''
            
            # CDをクリーンアップ
            if isinstance(customer_cd, float):
                import math
                if math.isnan(customer_cd):
                    continue
                customer_cd = str(int(customer_cd))
            else:
                customer_cd = str(customer_cd).strip()
            
            if not customer_cd:
                continue
            
            # 担当者をクリーンアップ
            if isinstance(staff, float):
                import math
                if math.isnan(staff):
                    staff = ''
                else:
                    staff = str(staff)
            else:
                staff = str(staff).strip() if staff else ''
                
            records.append({
                '得意先CD': customer_cd,
                '得意先名': str(customer_name).strip() if customer_name else '',
                '担当者': staff
            })
        
        return records
    except Exception as e:
        logging.error(f"Error in get_priority_customers: {e}")
        import traceback
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=str(e))



@app.get("/interviewers")
def get_interviewers(customer_code: str, filename: str = DEFAULT_EXCEL_FILE):
    """Get list of interviewers for a specific customer"""
    excel_file = os.path.join(EXCEL_DIR, filename)
    if not os.path.exists(excel_file):
        raise HTTPException(status_code=404, detail=f"Excel file '{filename}' not found")
    
    try:
        # Read the '営業日報' sheet
        df = pd.read_excel(excel_file, sheet_name='営業日報', header=0)
        
        # Clean up column names
        df.columns = [str(col).replace('\n', '').strip() for col in df.columns]
        
        # Rename specific columns to match frontend expectations
        df = df.rename(columns={
            '得意先CD.': '得意先CD',
        })
        
        # Filter by customer code
        customer_reports = df[df['得意先CD'] == customer_code]
        
        # Get unique interviewers, excluding NaN and '-'
        interviewers = customer_reports['面談者'].dropna().unique().tolist()
        interviewers = [i for i in interviewers if i and str(i).strip() not in ['-', 'nan', '']]
        
        return interviewers
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@app.get("/reports/{management_number}")
def get_report_by_id(management_number: int, filename: str = DEFAULT_EXCEL_FILE):
    """指定された管理番号の日報を取得"""
    try:
        # Get dataframe from cache
        df = get_cached_dataframe(filename, '営業日報')
        
        # Clean up column names
        df.columns = [str(col).replace('\n', '') for col in df.columns]
        
        # Rename specific columns
        df = df.rename(columns={
            '得意先CD.': '得意先CD',
            '訪問先名得意先名': '訪問先名',
            'コメント': '上長コメント',  # Excel uses 'コメント' for manager comment
            'コメント返信欄': 'コメント返信欄'  # Keep as-is for reply field
        })
        
        # Filter by management number
        report_df = df[df['管理番号'] == management_number]
        
        if report_df.empty:
            raise HTTPException(status_code=404, detail=f"Report with management number {management_number} not found")
        
        # Get the first (and should be only) record
        record = report_df.iloc[0].to_dict()
        
        # Clean the record
        import math
        cleaned_record = {}
        for key, value in record.items():
            if isinstance(value, float):
                if math.isnan(value) or math.isinf(value):
                    cleaned_record[key] = None
                elif key == '得意先CD' and not math.isnan(value):
                    cleaned_record[key] = str(int(value))
                else:
                    cleaned_record[key] = value
            elif value == '':
                cleaned_record[key] = None
            elif isinstance(value, str):
                import re
                cleaned_value = re.sub(r'_x000D_', '\n', value)
                cleaned_value = cleaned_value.replace('\r', '')
                cleaned_record[key] = cleaned_value
            else:
                cleaned_record[key] = value
        
        return cleaned_record
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@app.get("/reports")
def get_reports(filename: str = DEFAULT_EXCEL_FILE):
    try:
        logging.debug(f"Fetching reports for {filename} from {EXCEL_DIR}")
        # Get dataframe from cache
        df = get_cached_dataframe(filename, '営業日報')
        
        # Clean up column names (remove newlines and strip)
        df.columns = [str(col).replace('\n', '').strip() for col in df.columns]
        
        # Rename specific columns to match frontend expectations
        df = df.rename(columns={
            '得意先CD.': '得意先CD',
            '訪問先名得意先名': '訪問先名',
            '直送先CD.': '直送先CD',
            '直送先名.': '直送先名',
            'コメント': '上長コメント',  # Excel uses 'コメント' for manager comment
            'コメント返信欄': 'コメント返信欄'  # Keep as-is for reply field
        })
        
        # Replace all NaN, infinity, and null values with None
        # Use fillna to replace NaN with None
        df = df.fillna(value='')
        
        # Convert dates to string to avoid serialization issues
        if '日付' in df.columns:
             df['日付'] = df['日付'].astype(str)
        
        # Convert to dict and manually clean any remaining problematic values
        records = df.to_dict(orient="records")
        
        # Clean the records to ensure JSON compatibility
        import math
        cleaned_records = []
        for record in records:
            cleaned_record = {}
            for key, value in record.items():
                if isinstance(value, float):
                    if math.isnan(value) or math.isinf(value):
                        cleaned_record[key] = None
                    # Convert customer code to string without decimal
                    elif key in ['得意先CD', '直送先CD'] and not math.isnan(value):
                        cleaned_record[key] = str(int(value))
                    else:
                        cleaned_record[key] = value
                elif value == '':
                    cleaned_record[key] = None
                elif isinstance(value, str):
                    # Replace Excel's carriage return artifacts with proper newlines
                    import re
                    cleaned_value = re.sub(r'_x000D_', '\n', value)
                    cleaned_value = cleaned_value.replace('\r', '')
                    cleaned_record[key] = cleaned_value
                else:
                    cleaned_record[key] = value
            cleaned_records.append(cleaned_record)

        return cleaned_records
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@app.get("/interviewers/{customer_cd}")
def get_interviewers(
    customer_cd: str, 
    filename: str = DEFAULT_EXCEL_FILE,
    customer_name: Optional[str] = None,
    delivery_name: Optional[str] = None
):
    """Get list of interviewers for a specific customer with optional name filtering"""
    try:
        # Get dataframe from cache
        df = get_cached_dataframe(filename, '営業日報')
        
        # Clean up column names
        df.columns = [str(col).replace('\n', '') for col in df.columns]
        
        # Rename specific columns to standard names
        df = df.rename(columns={
            '得意先CD.': '得意先CD',
            '直送先CD.': '直送先CD',
            '訪問先名得意先名': '訪問先名', # Case where \n was removed
            '訪問先名\n得意先名': '訪問先名', # Just in case
        })
        
        # Convert customer_cd to float for matching (Excel stores as float)
        try:
            customer_cd_float = float(customer_cd)
        except ValueError:
            # If conversion fails, try string matching
            customer_cd_float = customer_cd
        
        # Base filter: customer code
        # Handle potential float/string mismatch by trying both if needed, but usually float is correct for number-like CDs
        customer_reports = df[df['得意先CD'] == customer_cd_float]
        
        # Name filtering logic
        if delivery_name:
            # If delivery_name is provided, filter by it
            # Ensure '直送先名' column exists
            if '直送先名' in customer_reports.columns:
                customer_reports = customer_reports[customer_reports['直送先名'] == delivery_name]
        
        # If no delivery_name, or if we want to also filter by customer_name when delivery_name is NOT set
        # The requirement is "past interviewers matching ... Direct Delivery Name (if provided) OR Customer Name (if matching)"
        # If we selected a delivery destination, we used the block above.
        # If we selected a customer (without direct delivery), we should strictly filter by that customer name if possible, 
        # to avoid showing interviewers from other branches if they share the same Customer CD (rare but possible).
        # OR more importantly, if we are in "Direct Delivery" mode, we already filtered.
        # If we are NOT in "Direct Delivery" mode (delivery_name is None), we might want to filter by customer_name
        # to ensure we don't pick up rows that HAVE a direct delivery name (i.e. different destination).
        elif customer_name:
             if '訪問先名' in customer_reports.columns:
                 # Filter rows where Visit Name contains customer_name OR matches exactly
                 # Since '訪問先名' in DB might contain 'Customer Name Direct Delivery Name' now,
                 # strict matching might be tricky.
                 # But usually '訪問先名' == 'Customer Name' for standard records.
                 # Let's try exact match or contains.
                 # Also, we might want to EXCLUDE rows that have a '直送先名' if we are selecting a generic customer?
                 # User said: " 直近のログではなく、得意先名や直送先名が一致する過去に入力した面談者のみ"
                 
                 # If delivery_name is NOT provided, it effectively means we want records for the main customer.
                 # So we should probably match rows where Visit Name is exactly Customer Name, OR contains it.
                 # Let's use simple filtering: if customer_name is passed, try to filter by it.
                 # But keep in mind 訪問先名 might include 直送先名 suffix now.
                 pass

        # Updated Strategy for clarity:
        # 1. Base: Customer CD
        # 2. If delivery_name provided: matches '直送先名' == delivery_name
        # 3. If delivery_name NOT provided but customer_name provided: matches '訪問先名' == customer_name OR '直送先名' is Empty/NaN
        #    (This covers the case where we selected the main customer and want to exclude direct delivery branches)
        
        if delivery_name and '直送先名' in customer_reports.columns:
            customer_reports = customer_reports[customer_reports['直送先名'] == delivery_name]
        elif '直送先名' in customer_reports.columns:
             # If no delivery name specified, we prefer rows that ALSO have no delivery name specified
             # This avoids suggesting interviewers specific to a branch when we selected the HQ
             customer_reports = customer_reports[
                 customer_reports['直送先名'].isna() | 
                 (customer_reports['直送先名'] == '') | 
                 (customer_reports['直送先名'] == None)
             ]

        interviewers = customer_reports['面談者'].dropna().unique().tolist()
        
        # Remove empty strings and sort
        interviewers = [str(i).strip() for i in interviewers if str(i).strip() and str(i).strip() != 'nan']
        interviewers = sorted(set(interviewers))
        
        return {"customer_cd": customer_cd, "interviewers": interviewers}
    except Exception as e:
        logging.error(f"Error in get_interviewers: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))

@app.get("/designs/{customer_cd}")
def get_designs(customer_cd: str, delivery_name: Optional[str] = None, filename: str = DEFAULT_EXCEL_FILE):
    """Get list of design requests for a specific customer (optionally filtered by delivery destination)"""
    try:
        logging.info(f"get_designs called: customer_cd={customer_cd}, delivery_name={delivery_name}")
        
        # Get dataframe from cache
        df = get_cached_dataframe(filename, '営業日報')
        
        # Clean up column names
        df.columns = [str(col).replace('\n', '') for col in df.columns]
        
        # Rename specific columns
        df = df.rename(columns={
            '得意先CD.': '得意先CD',
            '直送先名.': '直送先名'
        })
        
        logging.info(f"Columns in dataframe: {list(df.columns)[:20]}...")  # Log first 20 columns
        
        # Convert customer_cd to float for matching
        try:
            customer_cd_float = float(customer_cd)
        except ValueError:
            customer_cd_float = customer_cd
        
        # Filter by customer code
        customer_reports = df[df['得意先CD'] == customer_cd_float].copy()
        logging.info(f"After customer filter: {len(customer_reports)} rows")
        
        # Filter by delivery destination if provided
        if delivery_name and '直送先名' in customer_reports.columns:
            before_count = len(customer_reports)
            customer_reports = customer_reports[customer_reports['直送先名'] == delivery_name]
            logging.info(f"After delivery_name filter ({delivery_name}): {len(customer_reports)} rows (was {before_count})")
        else:
            logging.info(f"Skipping delivery_name filter: delivery_name={delivery_name}, has_column={'直送先名' in customer_reports.columns}")
        
        # Filter for records with design request number
        design_reports = customer_reports[customer_reports['デザイン依頼No.'].notna()]
        logging.info(f"After design no filter: {len(design_reports)} rows")
        
        # Get unique design request numbers
        unique_design_nos = design_reports['デザイン依頼No.'].unique()
        
        designs = []
        for design_no in unique_design_nos:
            # Get all records for this design number
            design_records = design_reports[design_reports['デザイン依頼No.'] == design_no]
            
            # Get the latest record (assuming lower down in Excel is newer, or we could sort by date if available)
            # Here we just take the last one in the dataframe which corresponds to the last row in Excel
            latest_record = design_records.iloc[-1]
            
            # Get the status
            status = str(latest_record['デザイン進捗状況']) if pd.notna(latest_record['デザイン進捗状況']) else ""
            
            # Skip designs with completed/rejected statuses: 出稿, 不採用(コンペ負け), 不採用(企画倒れ)
            if '出稿' in status or 'コンペ負け' in status or '企画倒れ' in status:
                continue
            
            design_info = {
                "デザイン依頼No": design_no,
                "デザイン名": str(latest_record['デザイン名']) if pd.notna(latest_record['デザイン名']) else "",
                "デザイン種別": str(latest_record['デザイン種別']) if pd.notna(latest_record['デザイン種別']) else "",
                "デザイン進捗状況": status,
                "デザイン提案有無": str(latest_record['デザイン提案有無']) if pd.notna(latest_record['デザイン提案有無']) else ""
            }
            designs.append(design_info)
            
        return {"customer_cd": customer_cd, "designs": designs}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.post("/reports")
def add_report(report: ReportInput, background_tasks: BackgroundTasks, filename: str = DEFAULT_EXCEL_FILE):
    excel_file = os.path.join(EXCEL_DIR, filename)
    if not os.path.exists(excel_file):
        raise HTTPException(status_code=404, detail=f"Excel file '{filename}' not found")
    
    try:
        # Load workbook with openpyxl to preserve formulas and macros
        wb = openpyxl.load_workbook(excel_file, keep_vba=True)
        ws = wb['営業日報']
        
        # Find the maximum management number and its row by scanning all rows
        max_mgmt_num = 0
        max_mgmt_row = 1  # Default to header row if no data found
        for row in range(2, ws.max_row + 1):  # Start from row 2 (skip header)
            mgmt_num = ws.cell(row=row, column=1).value
            try:
                if mgmt_num is not None:
                    val = int(mgmt_num)
                    if val > max_mgmt_num:
                        max_mgmt_num = val
                        max_mgmt_row = row
            except (ValueError, TypeError):
                continue
        
        logging.debug(f"Max Mgmt Num: {max_mgmt_num}, Max Mgmt Row: {max_mgmt_row}")

        # Increment to get new management number
        new_mgmt_num = max_mgmt_num + 1
        
        # Insert at the row immediately after the last management number
        next_row = max_mgmt_row + 1
        logging.debug(f"Writing to Row: {next_row}")
        
        # Prepare the data to write
        # Adjust column indices based on actual Excel structure (251113_2026-_-_008.xlsm)
        # Also copy styles from the previous row (max_mgmt_row) to the new row (next_row)
        
        from copy import copy

        def copy_style(source_cell, target_cell):
            if source_cell.has_style:
                target_cell.font = copy(source_cell.font)
                target_cell.border = copy(source_cell.border)
                target_cell.fill = copy(source_cell.fill)
                target_cell.number_format = copy(source_cell.number_format)
                target_cell.protection = copy(source_cell.protection)
                target_cell.alignment = copy(source_cell.alignment)

        # Define the columns to write to and their values
        columns_to_write = {
            1: new_mgmt_num, # 管理番号
            2: report.日付, # 日付
            3: report.行動内容, # 行動内容
            4: report.エリア, # エリア
            5: report.得意先CD, # 得意先CD.
            6: report.直送先CD, # 直送先CD.
            7: report.訪問先名, # 訪問先名\n得意先名
            8: report.直送先名, # 直送先名
            9: report.重点顧客, # 重点顧客
            10: report.ランク, # ランク
            11: report.面談者, # 面談者 (Corrected from 12)
            12: report.滞在時間, # 滞在\n時間 (Corrected from 13)
            13: report.デザイン提案有無, # デザイン提案有無 (Corrected from 14)
            14: report.デザイン種別, # デザイン種別 (Corrected from 15)
            15: report.デザイン名, # デザイン名 (Corrected from 16)
            16: report.デザイン進捗状況, # デザイン進捗状況 (Corrected from 17)
            17: report.デザイン依頼No, # デザイン依頼No. (Corrected from 18)
            18: report.商談内容, # 商談内容 (Corrected from 19)
            19: report.提案物, # 提案物 (Corrected from 20)
            20: report.次回プラン, # 次回プラン (Corrected from 21)
            21: report.競合他社情報, # 競合他社情報 (New)
            22: report.上長コメント, # コメント (Column 22)
            23: report.コメント返信欄  # コメント返信欄 (Column 23)
        }

        for col_idx, value in columns_to_write.items():
            target_cell = ws.cell(row=next_row, column=col_idx)
            target_cell.value = value
            
            # Copy style from the row above (max_mgmt_row)
            # Ensure we are copying from a valid row
            if max_mgmt_row >= 2:
                source_cell = ws.cell(row=max_mgmt_row, column=col_idx)
                copy_style(source_cell, target_cell)

        
        # Save the workbook (Critical path - blocking)
        wb.save(excel_file)
        wb.close()

        # Create backup in background
        background_tasks.add_task(create_backup, excel_file)
        
        
        # Clear cache
        cache_key = (filename, '営業日報')
        if cache_key in CACHE:
            del CACHE[cache_key]
        
        return {
            "message": "Report added successfully", 
            "management_number": new_mgmt_num,
            "file_path": os.path.abspath(excel_file)
        }
    except PermissionError:
        raise HTTPException(
            status_code=409,
            detail="ファイルが開かれているため保存できません。Excelファイルを閉じてから再度実行してください。"
        )
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

# コメント更新専用エンドポイント（楽観的ロックなし）
class CommentInput(BaseModel):
    上長コメント: Optional[str] = None
    コメント返信欄: Optional[str] = None

# 承認チェック更新専用
class ApprovalInput(BaseModel):
    上長: Optional[str] = None
    山澄常務: Optional[str] = None
    岡本常務: Optional[str] = None
    中野次長: Optional[str] = None
    既読チェック: Optional[str] = None

# 後方互換性のため
class ReplyInput(BaseModel):
    コメント返信欄: str

@app.patch("/reports/{management_number}/reply")
def update_report_reply(management_number: int, reply: ReplyInput, background_tasks: BackgroundTasks, filename: str = DEFAULT_EXCEL_FILE):
    """コメント返信欄のみを更新（安全な保存）"""
    import tempfile
    import shutil
    
    logging.debug(f"update_report_reply: management_number={management_number}, reply={reply.コメント返信欄}")
    try:
        excel_file = os.path.join(EXCEL_DIR, filename)
        logging.debug(f"excel_file: {excel_file}")
        
        wb = openpyxl.load_workbook(excel_file, keep_vba=True)
        logging.debug("workbook loaded")
        if '営業日報' not in wb.sheetnames:
            raise HTTPException(status_code=404, detail="Sheet '営業日報' not found")
        
        ws = wb['営業日報']
        logging.debug(f"worksheet max_row: {ws.max_row}")
        
        # Find the row
        target_row = None
        for row in range(2, ws.max_row + 1):
            if ws.cell(row=row, column=1).value == management_number:
                target_row = row
                break
        
        logging.debug(f"target_row: {target_row}")
        if not target_row:
            wb.close()
            raise HTTPException(status_code=404, detail=f"Report {management_number} not found")
        
        # Column 23 = コメント返信欄
        ws.cell(row=target_row, column=23, value=reply.コメント返信欄)
        logging.debug("cell set, saving to temp file...")
        
        # 安全な保存: 一時ファイルに保存してから置き換え
        temp_dir = tempfile.gettempdir()
        temp_file = os.path.join(temp_dir, f"temp_{filename}")
        
        try:
            # 一時ファイルに保存
            wb.save(temp_file)
            wb.close()
            logging.debug(f"saved to temp: {temp_file}")
            
            # 一時ファイルが正常か確認（読み込みテスト）
            test_wb = openpyxl.load_workbook(temp_file, read_only=True)
            test_wb.close()
            logging.debug("temp file verified")
            
            # 元のファイルを一時ファイルで置き換え
            shutil.copy2(temp_file, excel_file)
            logging.debug("replaced original file")
            
        finally:
            # 一時ファイルを削除
            if os.path.exists(temp_file):
                try:
                    os.remove(temp_file)
                except:
                    pass
        
        # Clear cache
        cache_key = (filename, '営業日報')
        if cache_key in CACHE:
            del CACHE[cache_key]
        
        return {"success": True, "management_number": management_number}
    except HTTPException:
        raise
    except Exception as e:
        import traceback
        logging.error(f"update_report_reply: {e}")
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=str(e))

@app.patch("/reports/{management_number}/comment")
def update_report_comment(management_number: int, comment: CommentInput, background_tasks: BackgroundTasks, filename: str = DEFAULT_EXCEL_FILE):
    """上長コメントとコメント返信欄を個別に更新（安全な保存）"""
    import tempfile
    import shutil
    
    logging.debug(f"update_report_comment: management_number={management_number}, 上長コメント={comment.上長コメント}, コメント返信欄={comment.コメント返信欄}")
    try:
        excel_file = os.path.join(EXCEL_DIR, filename)
        
        wb = openpyxl.load_workbook(excel_file, keep_vba=True)
        if '営業日報' not in wb.sheetnames:
            raise HTTPException(status_code=404, detail="Sheet '営業日報' not found")
        
        ws = wb['営業日報']
        
        # Find the row
        target_row = None
        for row in range(2, ws.max_row + 1):
            if ws.cell(row=row, column=1).value == management_number:
                target_row = row
                break
        
        if not target_row:
            wb.close()
            raise HTTPException(status_code=404, detail=f"Report {management_number} not found")
        
        # Update only provided fields
        # Column 22 = 上長コメント, Column 23 = コメント返信欄
        if comment.上長コメント is not None:
            ws.cell(row=target_row, column=22, value=comment.上長コメント)
        if comment.コメント返信欄 is not None:
            ws.cell(row=target_row, column=23, value=comment.コメント返信欄)
        
        # 安全な保存: 一時ファイルに保存してから置き換え
        temp_dir = tempfile.gettempdir()
        temp_file = os.path.join(temp_dir, f"temp_{filename}")
        
        try:
            wb.save(temp_file)
            wb.close()
            
            # 一時ファイルが正常か確認
            test_wb = openpyxl.load_workbook(temp_file, read_only=True)
            test_wb.close()
            
            # 元のファイルを置き換え
            shutil.copy2(temp_file, excel_file)
        finally:
            if os.path.exists(temp_file):
                try:
                    os.remove(temp_file)
                except:
                    pass
        
        # Clear cache
        cache_key = (filename, '営業日報')
        if cache_key in CACHE:
            del CACHE[cache_key]
        
        # Create backup in background
        background_tasks.add_task(create_backup, excel_file)
        
        return {"success": True, "management_number": management_number}
    except HTTPException:
        raise
    except Exception as e:
        import traceback
        logging.error(f"update_report_comment: {e}")
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=str(e))

@app.patch("/reports/{management_number}/approval")
def update_report_approval(management_number: int, approval: ApprovalInput, background_tasks: BackgroundTasks, filename: str = DEFAULT_EXCEL_FILE):
    """承認チェック（上長、山澄常務、岡本常務、中野次長、既読チェック）を個別に更新"""
    import tempfile
    import shutil
    
    logging.debug(f"update_report_approval: management_number={management_number}")
    try:
        excel_file = os.path.join(EXCEL_DIR, filename)
        
        wb = openpyxl.load_workbook(excel_file, keep_vba=True)
        if '営業日報' not in wb.sheetnames:
            raise HTTPException(status_code=404, detail="Sheet '営業日報' not found")
        
        ws = wb['営業日報']
        
        # Find the row
        target_row = None
        for row in range(2, ws.max_row + 1):
            if ws.cell(row=row, column=1).value == management_number:
                target_row = row
                break
        
        if not target_row:
            wb.close()
            raise HTTPException(status_code=404, detail=f"Report {management_number} not found")
        
        # Update only provided fields
        # Column mapping: 24=上長, 25=山澄常務, 26=岡本常務, 27=中野次長, 28=既読チェック
        column_mapping = {
            '上長': 24,
            '山澄常務': 25,
            '岡本常務': 26,
            '中野次長': 27,
            '既読チェック': 28
        }
        
        if approval.上長 is not None:
            ws.cell(row=target_row, column=column_mapping['上長'], value=approval.上長)
        if approval.山澄常務 is not None:
            ws.cell(row=target_row, column=column_mapping['山澄常務'], value=approval.山澄常務)
        if approval.岡本常務 is not None:
            ws.cell(row=target_row, column=column_mapping['岡本常務'], value=approval.岡本常務)
        if approval.中野次長 is not None:
            ws.cell(row=target_row, column=column_mapping['中野次長'], value=approval.中野次長)
        if approval.既読チェック is not None:
            ws.cell(row=target_row, column=column_mapping['既読チェック'], value=approval.既読チェック)
        
        # 安全な保存
        temp_dir = tempfile.gettempdir()
        temp_file = os.path.join(temp_dir, f"temp_{filename}")
        
        try:
            wb.save(temp_file)
            wb.close()
            
            test_wb = openpyxl.load_workbook(temp_file, read_only=True)
            test_wb.close()
            
            shutil.copy2(temp_file, excel_file)
        finally:
            if os.path.exists(temp_file):
                try:
                    os.remove(temp_file)
                except:
                    pass
        
        # Clear cache
        cache_key = (filename, '営業日報')
        if cache_key in CACHE:
            del CACHE[cache_key]
        
        # Create backup in background
        background_tasks.add_task(create_backup, excel_file)
        
        return {"success": True, "management_number": management_number}
    except HTTPException:
        raise
    except Exception as e:
        import traceback
        logging.error(f"update_report_approval: {e}")
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=str(e))

@app.post("/reports/{management_number}")
def update_report(management_number: int, report: ReportInput, background_tasks: BackgroundTasks, filename: str = DEFAULT_EXCEL_FILE):
    """既存の日報を更新（全項目対応）"""
    logging.info(f"update_report called: management_number={management_number}, original_values={report.original_values}")
    try:
        excel_file = os.path.join(EXCEL_DIR, filename)
        
        # Load the workbook
        wb = openpyxl.load_workbook(excel_file, keep_vba=True)
        if '営業日報' not in wb.sheetnames:
            raise HTTPException(status_code=404, detail="Sheet '営業日報' not found")
             
        ws = wb['営業日報']
        
        # Find the row with the matching management number
        target_row = None
        for row in range(2, ws.max_row + 1):
            cell_value = ws.cell(row=row, column=1).value
            if cell_value == management_number:
                target_row = row
                break
        
        if not target_row:
            raise HTTPException(status_code=404, detail=f"Report with management number {management_number} not found")

        # --- Optimistic Locking Check ---
        if report.original_values:
            logging.debug(f"Performing conflict check for Report {management_number}")
            
            # Fields to check for conflicts (critical text fields)
            check_fields = {
                22: '上長コメント',
                23: 'コメント返信欄',
                18: '商談内容'
            }
            
            conflicts = []
            for col_idx, field_name in check_fields.items():
                current_val = ws.cell(row=target_row, column=col_idx).value
                current_str = str(current_val) if current_val is not None else ""
                
                original_val = report.original_values.get(field_name, "")
                original_str = str(original_val) if original_val is not None else ""
                
                # Normalize newlines for comparison
                current_str = current_str.replace('\r\n', '\n').replace('\r', '\n').strip()
                original_str = original_str.replace('\r\n', '\n').replace('\r', '\n').strip()
                
                if current_str != original_str:
                    logging.warning(f"CONFLICT: Field '{field_name}' changed. Current: '{current_str}' vs Original: '{original_str}'")
                    conflicts.append(field_name)
            
            if conflicts:
                conflict_msg = ", ".join(conflicts)
                raise HTTPException(
                    status_code=409, 
                    detail=f"他の方が編集しました（{conflict_msg}）。最新の情報を読み込んでからやり直してください。"
                )
        # --------------------------------
        
        # Update all fields (same column mapping as add_report)
        columns_to_write = {
            2: report.日付,
            3: report.行動内容,
            4: report.エリア,
            5: report.得意先CD,
            6: report.直送先CD,
            7: report.訪問先名,
            8: report.直送先名,
            9: report.重点顧客,
            10: report.ランク,
            11: report.面談者,
            12: report.滞在時間,
            13: report.デザイン提案有無,
            14: report.デザイン種別,
            15: report.デザイン名,
            16: report.デザイン進捗状況,
            17: report.デザイン依頼No,
            18: report.商談内容,
            19: report.提案物,
            20: report.次回プラン,
            21: report.競合他社情報,
            22: report.上長コメント, # コメント (Column 22)
            23: report.コメント返信欄  # コメント返信欄 (Column 23)
        }

        for col_idx, value in columns_to_write.items():
            ws.cell(row=target_row, column=col_idx, value=value)
        
        # Save the workbook (Critical path - blocking)
        wb.save(excel_file)
        wb.close()
        
        # Create backup in background
        background_tasks.add_task(create_backup, excel_file)
        
        # Clear cache for this file
        cache_key = (filename, '営業日報')
        if cache_key in CACHE:
            del CACHE[cache_key]
        
        return {"message": "Report updated successfully", "management_number": management_number}
    except HTTPException:
        raise
    except PermissionError:
        raise HTTPException(
            status_code=409,
            detail="ファイルが開かれているため保存できません。Excelファイルを閉じてから再度実行してください。"
        )
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@app.delete("/reports/{management_number}")
def delete_report(management_number: int, filename: str = DEFAULT_EXCEL_FILE):
    """指定された管理番号の日報を削除"""
    try:
        excel_file = os.path.join(EXCEL_DIR, filename)
        
        # Load the workbook
        wb = openpyxl.load_workbook(excel_file, keep_vba=True)
        if '営業日報' not in wb.sheetnames:
            raise HTTPException(status_code=404, detail="Sheet '営業日報' not found")
             
        ws = wb['営業日報']
        
        # Find the row with the matching management number
        target_row = None
        for row in range(2, ws.max_row + 1):
            cell_value = ws.cell(row=row, column=1).value
            if cell_value == management_number:
                target_row = row
                break
        
        if not target_row:
            raise HTTPException(status_code=404, detail=f"Report with management number {management_number} not found")
        
        # Delete the row
        ws.delete_rows(target_row, 1)
        
        # Save the workbook
        wb.save(excel_file)
        wb.close()
        
        # Clear cache for this file
        cache_key = (filename, '営業日報')
        if cache_key in CACHE:
            del CACHE[cache_key]
        
        return {"message": "Report deleted successfully", "management_number": management_number}
    except HTTPException:
        raise
    except PermissionError:
        raise HTTPException(
            status_code=409,
            detail="ファイルが開かれているため保存できません。Excelファイルを閉じてから再度実行してください。"
        )
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.post("/upload")
async def upload_file(file: UploadFile = File(...)):
    """Upload an Excel file to the backend directory"""
    try:
        # Validate file extension
        if not file.filename.endswith(('.xlsx', '.xlsm')):
            raise HTTPException(status_code=400, detail="Only .xlsx and .xlsm files are allowed")
        
        # Save the uploaded file
        file_path = os.path.join(EXCEL_DIR, file.filename)
        
        with open(file_path, "wb") as buffer:
            shutil.copyfileobj(file.file, buffer)
        
        return {
            "message": "File uploaded successfully",
            "filename": file.filename,
            "path": file_path
        }
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


# Set up logging
import logging

logging.basicConfig(
    filename='debug.log',
    level=logging.DEBUG,
    format='%(asctime)s - %(levelname)s - %(message)s',
    encoding='utf-8' # Ensure we can log Japanese characters
)

@app.get("/images/list")
def get_design_images(filename: str):
    """
    Get list of images from the matching folder in Design Data directory.
    Target directory: \\Asahipack02\\社内書類ｎｅｗ\\01：部署別　営業部\\03：デザインデータ
    Logic: Extract name from filename '...【Name】.xlsm' -> Search folder containing 'Name'
    """
    DESIGN_DIR = r"\\Asahipack02\社内書類ｎｅｗ\01：部署別　営業部\03：デザインデータ"
    
    logging.info(f"--- get_design_images called with filename: {filename} ---")
    
    # 0. 手動マッピングチェック
    # 特定のファイル名を特定のフォルダ名または検索語句にマッピング
    FOLDER_MAPPING = {
        # 正確なファイル名 -> 正確なターゲットフォルダ名（または検索する部分文字列）
        "本社006　2025年度用日報【木村（拓）MGR】.xlsm": "大阪本社　05：木村（拓）",
    }
    
    matched_dir = None

    if filename in FOLDER_MAPPING:
        mapped_target = FOLDER_MAPPING[filename]
        logging.info(f"Manual mapping found for {filename}: {mapped_target}")
        
        # マッピングされたディレクトリが存在するか確認
        path_check = os.path.join(DESIGN_DIR, mapped_target)
        if os.path.isdir(path_check):
            matched_dir = mapped_target
            logging.info(f"Mapped directory verified: {matched_dir}")
        else:
             logging.warning(f"Mapped directory not found: {path_check}")
             # 通常の検索にフォールバックするか失敗させるか？ここではフォールバックします。
    
    if not matched_dir:
        try:
            # Extract name from filename (e.g., 本社009　2025年度用日報【沖本】.xlsm -> 沖本)
            import re
            # Helper for normalization
            def normalize_text(text):
                # Convert full-width parens and space to half-width
                text = text.replace('（', '(').replace('）', ')').replace('　', ' ')
                # Strip whitespace
                return text.strip()

            match = re.search(r'【(.*?)】', filename)
            if not match:
                logging.warning("Regex match failed for filename")
                # Fallback extraction?
                target_name = os.path.splitext(os.path.basename(filename))[0]
                logging.info(f"Fallback extracted name: {target_name}")
            else:
                target_name = match.group(1)
                logging.info(f"Regex extracted name: {target_name}")

            normalized_target = normalize_text(target_name)
            logging.info(f"Normalized target: {normalized_target}")
            
            # 接尾辞を削除
            # MGR/Mgr をリストに追加
            stripped_target = re.sub(r'(MGR|Mgr|次長|課長|部長|係長|主任|担当|顧問|専務|常務|社長)$', '', normalized_target, flags=re.IGNORECASE)
            logging.info(f"Stripped target: {stripped_target}")

            logging.debug(f"Searching for folder containing '{target_name}' (Norm: {normalized_target}) in {DESIGN_DIR}")
            
            if not os.path.exists(DESIGN_DIR):
                 logging.error(f"Design directory not found: {DESIGN_DIR}")
                 return {"message": "Design directory not found", "images": []}

            # Find matching directory
            
            try:
                dir_list = os.listdir(DESIGN_DIR)
                # logging.debug(f"Directory listing (first 5): {dir_list[:5]}")
            except Exception as e:
                logging.error(f"Failed to list directory: {e}")
                return {"message": f"Failed to access design dir: {e}", "images": []}

            # 1. Try exact match (normalized)
            for item in dir_list:
                if not os.path.isdir(os.path.join(DESIGN_DIR, item)):
                    continue
                    
                norm_item = normalize_text(item)
                if normalized_target in norm_item:
                    matched_dir = item
                    logging.info(f"Match found (Normalized): {item}")
                    break
            
            # 2. If no match, try suffix stripping (e.g. 山下(和)次長 -> 山下(和))
            if not matched_dir:
                if stripped_target != normalized_target:
                     logging.info("Retrying with stripped name...")
                     for item in dir_list:
                        if not os.path.isdir(os.path.join(DESIGN_DIR, item)):
                            continue
                        norm_item = normalize_text(item)
                        if stripped_target in norm_item:
                            matched_dir = item
                            logging.info(f"Match found (Stripped): {item}")
                            break
            
            if not matched_dir:
                logging.warning(f"No folder found for target: {normalized_target} / {stripped_target}")
                return {"message": f"No folder found for '{target_name}'", "images": []}
                
        except Exception as e:
             logging.error(f"Error during folder search logic: {e}")
             raise HTTPException(status_code=500, detail=str(e))
            
    try:
        target_path = os.path.join(DESIGN_DIR, matched_dir)
        logging.info(f"Target path: {target_path}")
        
        # List images (recursively or just top level? Starting with top level + shallow)
        # Extensions to look for
        valid_extensions = ('.jpg', '.jpeg', '.png', '.gif', '.bmp', '.webp', '.pdf')
        
        image_files = []
        for root, dirs, files in os.walk(target_path):
            for file in files:
                if file.lower().endswith(valid_extensions):
                    # Create a relative path from DESIGN_DIR for the client to request
                    # e.g., "大阪本社　09：沖本\image.jpg"
                    full_path = os.path.join(root, file)
                    rel_path = os.path.relpath(full_path, DESIGN_DIR)
                    try:
                        mtime = os.path.getmtime(full_path)
                    except:
                        mtime = 0
                    
                    image_files.append({
                        "name": file,
                        "path": rel_path, # Path identifier to send back to serve endpoint
                        "folder": matched_dir,
                        "mtime": mtime
                    })
            if len(image_files) > 100:
                break
        
        # Sort by mtime descending (newest first)
        image_files.sort(key=lambda x: x['mtime'], reverse=True)
        
        return {"images": image_files, "folder": matched_dir}

    except Exception as e:
        logging.exception("Error in get_design_images")
        logging.error(f"Error listing images: {e}")
        raise HTTPException(status_code=500, detail=str(e))

@app.get("/images/content")
def serve_design_image(path: str):
    """
    Serve the image file content.
    path: Relative path from DESIGN_DIR (e.g., "大阪本社　09：沖本\image.jpg")
    """
    DESIGN_DIR = r"\\Asahipack02\社内書類ｎｅｗ\01：部署別　営業部\03：デザインデータ"
    
    try:
        # Security check: Prevent directory traversal
        # relpath needs to be safe. 
        # Since we construct it ourselves in /images/list, it should be fine, but good to check.
        safe_path = os.path.normpath(os.path.join(DESIGN_DIR, path))
        
        if not safe_path.startswith(DESIGN_DIR):
            raise HTTPException(status_code=403, detail="Access denied")
            
        if not os.path.exists(safe_path):
            raise HTTPException(status_code=404, detail="Image not found")
            
        from fastapi.responses import FileResponse
        return FileResponse(safe_path)
    except Exception as e:
         raise HTTPException(status_code=500, detail=str(e))

@app.get("/images/search")
def search_design_images(query: str, filename: Optional[str] = None):
    """
    Search for images matching the query (Design No) in the Design Data directory.
    If filename is provided (e.g. '見上.xlsm'), it tries to find a matching user folder first (e.g. '08：見上').
    Recursively searches subfolders.
    """
    DESIGN_DIR = r"\\Asahipack02\社内書類ｎｅｗ\01：部署別　営業部\03：デザインデータ"
    
    logging.info(f"--- search_design_images called. Query: {query}, Filename: {filename} ---")

    if not query or len(query.strip()) < 2:
        return {"message": "Query too short", "images": []}
        
    # Helper for normalization
    def normalize_text(text):
        # Convert full-width parens and space to half-width
        text = text.replace('（', '(').replace('）', ')').replace('　', ' ')
        # Strip whitespace
        return text.strip()

    try:
        if not os.path.exists(DESIGN_DIR):
             return {"message": "Design directory not found", "images": []}

        search_roots = [DESIGN_DIR]
        
        # Optimize: Try to find specific user folder based on filename
        found_folder = None
        if filename:
            try:
                # Extract name part
                # 1. Try to extract from 【】
                import re
                match = re.search(r'【(.*?)】', filename)
                if match:
                    name_part = match.group(1)
                else:
                    # 拡張子削除にフォールバック
                    name_part = os.path.splitext(os.path.basename(filename))[0]
                
                # 名前部分を正規化
                normalized_name = normalize_text(name_part)
                # 接尾辞（サフィックス）を削除したバージョンも用意
                stripped_name = re.sub(r'(MGR|Mgr|次長|課長|部長|係長|主任|担当|顧問|専務|常務|社長)$', '', normalized_name, flags=re.IGNORECASE)
                
                logging.info(f"Search optimization - Extracted: {name_part}, Norm: {normalized_name}, Stripped: {stripped_name}")

                # Use scandir for better performance on network drive for top-level listing
                with os.scandir(DESIGN_DIR) as it:
                    for entry in it:
                        if entry.is_dir():
                            norm_entry_name = normalize_text(entry.name)
                            # 抽出された名前（正規化済み）がフォルダ名（正規化済み）に含まれているか確認
                            if normalized_name in norm_entry_name:
                                found_folder = entry.path
                                logging.info(f"Optimization - Found folder (Norm): {entry.name}")
                                break
                            # 見つからない場合、接尾辞なしの名前を試す
                            if stripped_name != normalized_name and stripped_name in norm_entry_name:
                                found_folder = entry.path
                                logging.info(f"Optimization - Found folder (Stripped): {entry.name}")
                                break

                if found_folder:
                    logging.debug(f"Search target set to: {found_folder}")
                    search_roots = [found_folder]
            except Exception as e:
                logging.error(f"Failed to optimize search folder: {e}")
                logging.warning(f"Failed to optimize search folder: {e}")
                
        
        if found_folder:
            search_roots = [found_folder]
        else:
            if filename:
                 # print(f"WARN: Could not find user folder for {filename}. Aborting full scan to prevent timeout.")
                 logging.info("User folder not found from filename")
                 return {"message": "User folder not found from filename", "images": []}
            
            # If no filename provided, we search everything? That's dangerous too.
            search_roots = [DESIGN_DIR]

        image_files = []
        valid_extensions = ('.jpg', '.jpeg', '.png', '.gif', '.bmp', '.webp', '.pdf') 
        
        count = 0
        MAX_RESULTS = 50
        
        # Helper for safer walking on finicky network drives
        def safe_walk(directory, query_lower, extensions, max_depth=3, current_depth=0, parent_matches_query=False):
            results = []
            try:
                # Use listdir instead of scandir/walk to avoid hanging
                items = os.listdir(directory)
            except Exception as e:
                logging.warning(f"Failed to listdir {directory}: {e}")
                return results

            dirs_to_visit = []

            for name in items:
                full_path = os.path.join(directory, name)
                name_lower = name.lower()
                
                # Check file match
                # IF parent folder matched, we take ALL images.
                # IF not, we only take images matching query.
                is_file_match = False
                if parent_matches_query and name_lower.endswith(extensions):
                    is_file_match = True
                elif query_lower in name_lower and name_lower.endswith(extensions):
                    is_file_match = True
                
                if is_file_match:
                    try:
                        if os.path.isfile(full_path):
                            try:
                                rel_path = os.path.relpath(full_path, DESIGN_DIR)
                                folder_name = os.path.basename(directory)
                                try:
                                    mtime = os.path.getmtime(full_path)
                                except:
                                    mtime = 0
                                results.append({
                                    "name": name,
                                    "path": rel_path,
                                    "folder": folder_name,
                                    "mtime": mtime
                                })
                            except ValueError:
                                pass
                    except Exception:
                        pass
                
                # Identify directories for recursion
                if current_depth < max_depth:
                    next_parent_matches = parent_matches_query
                    
                    # Logic:
                    # 1. If parent already matched, we continue down (inheriting True).
                    # 2. If parent didn't match, verify if THIS folder matches.
                    if not next_parent_matches:
                        if query_lower in name_lower:
                            next_parent_matches = True
                    
                    # Recursion Filter:
                    # If we are NOT in a matching tree yet, we MUST skip unrelated folders 
                    # to avoid massive scan (timeout).
                    if not next_parent_matches:
                         # Skip unrelated folders
                         # BUT we must handle "Generic" folders like "Data", "Images", "Design", etc.
                         # AND we must skip "Specific" folders that don't match (e.g. "12345-1").
                         
                         # Heuristic:
                         # 1. If folder name looks like a Design ID (5+ digits, maybe hyphens), assume it's Specific.
                         #    If query is NOT in it, SKIP.
                         # 2. Otherwise, assume it's Generic (e.g. "Data", "2025", "01_Sales").
                         #    ENTER.
                         
                         is_specific_id = False
                         # Simple regex for design ID: typically 5-8 digits, optionally followed by hyphens/more digits
                         # e.g. 117675, 117675-1, 101219-106074-3
                         # But NOT "2025" (Year) or "01" (Section).
                         # Let's say "Specific" if it has 5 or more consecutive digits.
                         import re
                         if re.search(r'\d{5,}', name):
                             is_specific_id = True
                         
                         # Refinement: What if the query is SMALL? e.g. "555"?
                         # The query is usually a Design No (5-6 digits).
                         # If query is matches regex, it's specific.
                         
                         if is_specific_id and query_lower not in name_lower:
                             # It looks like a DIFFERENT specific ID. Skip.
                             continue
                         
                         # If it is generic (not super long digits) OR it matched query, we enter.
                         pass

                    if '.' in name and not name.startswith('.'):
                        continue 
                        
                    try:
                        if os.path.isdir(full_path):
                            dirs_to_visit.append((full_path, next_parent_matches))
                    except Exception:
                        pass
            
            # Recurse
            for subdir_path, matches_status in dirs_to_visit:
                 sub_results = safe_walk(subdir_path, query_lower, extensions, max_depth, current_depth + 1, matches_status)
                 results.extend(sub_results)
                 if len(results) >= 50: 
                     break
            
            return results

        # Main search loop
        try:
            for search_root in search_roots:
                logging.debug(f"Searching root: {search_root}")
                # Use custom walker
                # Initial parent_matches logic:
                # If the search_root ITSELF matches query (e.g. we targeted a specific user folder matched by filename, but query is DesignNo),
                # expected behavior: search for DesignNo INSIDE user folder.
                # So initial parent_matches = False usually.
                found_images = safe_walk(search_root, query.lower(), valid_extensions)
                image_files.extend(found_images)
                if len(image_files) >= MAX_RESULTS:
                    image_files = image_files[:MAX_RESULTS]
                    break
        except Exception as e:
            logging.error(f"Search loop failed: {e}")
            # Return whatever we found so far instead of 500
            pass
            
        # Sort by mtime descending (newest first)
        image_files.sort(key=lambda x: x['mtime'], reverse=True)

        return {"images": image_files, "query": query}

    except Exception as e:
        # print(f"Error searching images: {e}")
        # Log to file safely if needed, or just return detailed 500
        # Ensure 'e' conversion to string doesn't fail
        error_msg = "Unknown error"
        try:
            error_msg = str(e)
        except:
            pass
        raise HTTPException(status_code=500, detail=error_msg)


# --- Sales Data Integration ---
# --- Sales Data Integration (Global) ---
DATA_DIR = os.path.join(os.path.dirname(__file__), 'data')
SALES_CSV_PATH = os.path.join(DATA_DIR, 'sales_data.csv')
os.makedirs(DATA_DIR, exist_ok=True)

global_sales_df = None

def load_sales_data():
    """Loads sales data from CSV into global DataFrame."""
    global global_sales_df
    if not os.path.exists(SALES_CSV_PATH):
        logging.info("No existing sales data found.")
        return

    try:
        logging.info("Loading sales data from disk...")
        try:
            df = pd.read_csv(SALES_CSV_PATH, encoding='cp932')
        except:
            df = pd.read_csv(SALES_CSV_PATH, encoding='utf-8')
        
        df.columns = [str(col).strip() for col in df.columns]
        
        if '得意先コード' in df.columns:
            df['得意先コード'] = df['得意先コード'].astype(str).str.split('.').str[0]
            global_sales_df = df
            logging.info(f"Sales data loaded successfully. {len(df)} rows.")
        else:
            logging.error("Sales CSV missing '得意先コード' column.")
    except Exception as e:
        logging.error(f"Failed to load sales data: {e}")

# Load on startup
load_sales_data()

@app.post("/sales/upload")
async def upload_sales_csv(file: UploadFile = File(...)):
    """
    Uploads a global sales data CSV file, saves it, and unloads it into memory.
    """
    try:
        logging.info(f"Receiving sales CSV: {file.filename}")
        contents = await file.read()
        
        import io
        try:
            pd.read_csv(io.BytesIO(contents), encoding='cp932')
        except:
            try:
                pd.read_csv(io.BytesIO(contents), encoding='utf-8')
            except Exception as e:
                raise HTTPException(status_code=400, detail="Invalid CSV format. Please use Shift-JIS or UTF-8.")

        with open(SALES_CSV_PATH, "wb") as f:
            f.write(contents)
        
        logging.info("Sales CSV saved to disk.")
        load_sales_data()
        
        return {"message": "Sales data uploaded and processed successfully."}

    except Exception as e:
        logging.error(f"Error uploading sales CSV: {e}")
        raise HTTPException(status_code=500, detail=f"Upload failed: {str(e)}")


@app.get("/sales/all")
async def get_all_sales_data():
    """
    Retrieves ALL sales data as a list.
    """
    if global_sales_df is None:
        return []

    try:
        # Convert NaN to None for JSON compliance
        df_clean = global_sales_df.where(pd.notnull(global_sales_df), None)
        
        # Select relevant columns and rename for consistency
        records = []
        for _, row in df_clean.iterrows():
            records.append({
                "rank": row.get('順位'),
                "rank_class": row.get('ランク'),
                "customer_code": row.get('得意先コード'),
                "customer_name": row.get('得意先名称'),
                "sales_amount": row.get('売上金額'),
                "gross_profit": row.get('粗利金額'),
                "sales_yoy": row.get('前年対比率'),
                "sales_last_year": row.get('前年売上'),
                "profit_last_year": row.get('前年粗利'),
                "sales_2y_ago": row.get('前々年売上'),
                "profit_2y_ago": row.get('前々年粗利'),
                # Attempt to get area from '地域名称' or '地域' or Column M (index 12)
                "area": row.get('地域名称') or row.get('地域') or (row.iloc[12] if len(row) > 12 else None),
                # 担当者 from Column I (index 8)
                "sales_rep": row.get('担当者') or (row.iloc[8] if len(row) > 8 else None),
            })
            
        return records

    except Exception as e:
        logging.error(f"Error retrieving all sales data: {e}")
        raise HTTPException(status_code=500, detail=f"Error retrieving data: {str(e)}")


@app.get("/sales/{customer_code}")
async def get_sales_data(customer_code: str):
    """
    Retrieves sales data for a specific customer from the global dataset.
    """
    if global_sales_df is None:
        return {"found": False, "message": "Sales data not yet uploaded."}
    
    try:
        target_code = str(customer_code).split('.')[0]
        matched_row = global_sales_df[global_sales_df['得意先コード'] == target_code]
        
        if matched_row.empty:
            return {"found": False, "message": "Customer not found in sales data."}
        
        row = matched_row.iloc[0]
        
        def get_val(col):
            val = row.get(col)
            if pd.isna(val):
                return None
            if hasattr(val, 'item'): 
                return val.item() 
            return val

        from datetime import datetime
        data = {
            "found": True,
            "rank": get_val('順位'),
            "rank_class": get_val('ランク'),
            "sales_amount": get_val('売上金額'),
            "gross_profit": get_val('粗利金額'),
            "sales_yoy": get_val('前年対比率'),
            "sales_last_year": get_val('前年売上'),
            "profit_last_year": get_val('前年粗利'),
            "sales_2y_ago": get_val('前々年売上'),
            "profit_2y_ago": get_val('前々年粗利'),
            "customer_name": get_val('得意先名称'),
            "updated_at": datetime.now().isoformat()
        }
        return data

    except Exception as e:
        logging.error(f"Error retrieving sales data: {e}")
        raise HTTPException(status_code=500, detail=f"Error retrieving data: {str(e)}")


# --- Static File Serving for Standalone App ---
STATIC_DIR = os.path.join(BASE_DIR, "static")

if os.path.exists(STATIC_DIR):
    # Mount _next directory for Next.js assets
    # Check if _next exists inside static to avoid error
    if os.path.exists(os.path.join(STATIC_DIR, "_next")):
         app.mount("/_next", StaticFiles(directory=os.path.join(STATIC_DIR, "_next")), name="next_assets")

    @app.get("/")
    async def serve_index():
        index_path = os.path.join(STATIC_DIR, "index.html")
        if os.path.exists(index_path):
            return FileResponse(index_path)
        return {"message": "Daily Report System API (Static files not found)"}

    @app.get("/{full_path:path}")
    async def serve_spa(full_path: str):
        # Check if file exists in static dir
        file_path = os.path.join(STATIC_DIR, full_path)
        if os.path.isfile(file_path):
            return FileResponse(file_path)
        
        # Check if it maps to a .html file (e.g. /design-search -> design-search.html)
        html_path = file_path + ".html"
        if os.path.isfile(html_path):
            return FileResponse(html_path)
            
        # Check if it's a directory with index.html
        index_path = os.path.join(file_path, "index.html")
        if os.path.isfile(index_path):
            return FileResponse(index_path)
        
        # If not found, return index.html for SPA routing
        # (API routes are already handled by precedence)
        spa_index = os.path.join(STATIC_DIR, "index.html")
        if os.path.exists(spa_index):
             return FileResponse(spa_index)
        
        return {"detail": "Not Found"}
# ----------------------------------------------

if __name__ == "__main__":
    import uvicorn
    uvicorn.run(app, host="0.0.0.0", port=8001)



