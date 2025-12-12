import openpyxl
import os

EXCEL_DIR = r"C:\Users\見上　祐司\.gemini\antigravity\scratch\daily-report-system"
DEFAULT_EXCEL_FILE = "本社008　2025年度用日報【見上】.xlsm"
excel_file = os.path.join(EXCEL_DIR, DEFAULT_EXCEL_FILE)

try:
    wb = openpyxl.load_workbook(excel_file, read_only=True, data_only=True)
    ws = wb['営業日報']
    
    print("Column Indices and First Row Data:")
    headers = [str(cell.value).replace('\n', '') for cell in ws[1]]
    first_row = [str(cell.value) for cell in ws[2]]
    
    for i, (header, data) in enumerate(zip(headers, first_row), 1):
        print(f"{i}: {header} (Sample: {data})")
        
    wb.close()
except Exception as e:
    print(f"Error: {e}")
