import pytest
import os
import sys
import tempfile
import openpyxl
from concurrent.futures import ThreadPoolExecutor
from fastapi import HTTPException, BackgroundTasks

sys.path.insert(0, os.path.abspath(os.path.join(os.path.dirname(__file__), '..')))

import config
import excel_io
import routes_reports
import models
import excel_schema

def test_file_write_lock_mutual_exclusion():
    """同一ファイルに対する排他ロックが正しく機能することを検証"""
    temp_dir = tempfile.mkdtemp()
    test_file = os.path.join(temp_dir, "lock_test.xlsx")
    
    # タイムアウト0.1秒で即時競合をテスト
    with excel_io.get_file_write_lock(test_file, timeout=5.0):
        # 同じスレッド内（再帰ロック）は取得可能
        with excel_io.get_file_write_lock(test_file, timeout=1.0):
            pass
        
        # 別スレッドからの取得はタイムアウトして 409 Conflict になること
        def try_acquire():
            with excel_io.get_file_write_lock(test_file, timeout=0.2):
                pass

        with ThreadPoolExecutor(max_workers=1) as executor:
            future = executor.submit(try_acquire)
            with pytest.raises(HTTPException) as exc_info:
                future.result()
            assert exc_info.value.status_code == 409

    try:
        os.rmdir(temp_dir)
    except Exception:
        pass


def test_concurrent_add_reports():
    """複数スレッドからの同時日報追加で管理番号が重複せず連番になることを検証"""
    temp_dir = tempfile.mkdtemp()
    test_filename = "concurrent_reports.xlsm"
    test_filepath = os.path.join(temp_dir, test_filename)

    # テンプレートExcelの作成
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = excel_schema.SHEET_DAILY_REPORT
    # ヘッダー行を作成
    headers = ["管理番号", "日付", "行動内容", "訪問先名"]
    for col_idx, h in enumerate(headers, 1):
        ws.cell(row=1, column=col_idx, value=h)
    
    # 得意先リストシートも作成
    cust_ws = wb.create_sheet(title=excel_schema.SHEET_CUSTOMERS)
    cust_ws.cell(row=1, column=1, value="得意先CD")
    cust_ws.cell(row=1, column=10, value="現目標")

    wb.save(test_filepath)
    wb.close()

    # config.EXCEL_DIR を一時ディレクトリに向ける
    orig_excel_dir = config.EXCEL_DIR
    config.EXCEL_DIR = temp_dir

    num_threads = 8
    bg_tasks = BackgroundTasks()

    try:
        def add_single_report(idx):
            data = {
                "日付": "26/08/18",
                "行動内容": "訪問（新規）",
                "訪問先名": f"株式会社テスト顧客_{idx}",
                "商談内容": f"テスト商談内容_{idx}",
            }
            report_in = models.ReportInput(**data)
            return routes_reports.add_report(report_in, bg_tasks, filename=test_filename)

        with ThreadPoolExecutor(max_workers=num_threads) as executor:
            futures = [executor.submit(add_single_report, i) for i in range(num_threads)]
            results = [f.result() for f in futures]

        # 管理番号の検証
        assigned_nums = [r["management_number"] for r in results]
        assert len(assigned_nums) == num_threads
        # 重複がないこと
        assert len(set(assigned_nums)) == num_threads
        # 1〜num_threads までの連番になっていること
        assert sorted(assigned_nums) == list(range(1, num_threads + 1))

        # 実際にExcelファイルを開いて行数と内容を確認
        saved_wb = openpyxl.load_workbook(test_filepath, data_only=True)
        saved_ws = saved_wb[excel_schema.SHEET_DAILY_REPORT]
        
        # ヘッダー1行 + num_threads行 = num_threads + 1行
        assert saved_ws.max_row == num_threads + 1
        
        excel_nums = [saved_ws.cell(row=r, column=1).value for r in range(2, num_threads + 2)]
        assert sorted(excel_nums) == list(range(1, num_threads + 1))
        saved_wb.close()

    finally:
        config.EXCEL_DIR = orig_excel_dir
        try:
            os.remove(test_filepath)
            # バックアップフォルダがあれば削除
            backup_dir = os.path.join(temp_dir, 'backup')
            if os.path.exists(backup_dir):
                for f in os.listdir(backup_dir):
                    os.remove(os.path.join(backup_dir, f))
                os.rmdir(backup_dir)
            os.rmdir(temp_dir)
        except Exception:
            pass


def test_optimistic_conflict_detection():
    """商談内容が他者により変更された場合に 409 Conflict が発生することを検証"""
    temp_dir = tempfile.mkdtemp()
    test_filename = "conflict_test.xlsm"
    test_filepath = os.path.join(temp_dir, test_filename)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = excel_schema.SHEET_DAILY_REPORT
    ws.cell(row=1, column=1, value="管理番号")
    ws.cell(row=2, column=1, value=1)
    ws.cell(row=2, column=excel_schema.DailyReportColumns.BUSINESS_CONTENT, value="元の商談内容A")

    wb.save(test_filepath)
    wb.close()

    orig_excel_dir = config.EXCEL_DIR
    config.EXCEL_DIR = temp_dir
    bg_tasks = BackgroundTasks()

    try:
        # ユーザーAが取得した古いデータ (商談内容: "異なる商談内容B")
        # 実際のExcelセルは "元の商談内容A" のため不一致で衝突
        stale_report = models.ReportInput(
            日付="26/08/18",
            商談内容="更新したい内容",
            original_values={"商談内容": "異なる商談内容B"}
        )

        with pytest.raises(HTTPException) as exc_info:
            routes_reports.update_report(1, stale_report, bg_tasks, filename=test_filename)

        assert exc_info.value.status_code == 409
        assert "他の方が編集しました" in exc_info.value.detail

        # 一致している場合は正常に更新されること
        valid_report = models.ReportInput(
            日付="26/08/18",
            商談内容="正常に更新された商談内容",
            original_values={"商談内容": "元の商談内容A"}
        )
        res = routes_reports.update_report(1, valid_report, bg_tasks, filename=test_filename)
        assert res["message"] == "Report updated successfully"

    finally:
        config.EXCEL_DIR = orig_excel_dir
        try:
            os.remove(test_filepath)
            backup_dir = os.path.join(temp_dir, 'backup')
            if os.path.exists(backup_dir):
                for f in os.listdir(backup_dir):
                    os.remove(os.path.join(backup_dir, f))
                os.rmdir(backup_dir)
            os.rmdir(temp_dir)
        except Exception:
            pass


def test_reply_conflict_detection():
    """コメント返信で original_values 不一致時に 409 Conflict が発生することを検証"""
    temp_dir = tempfile.mkdtemp()
    test_filename = "reply_conflict_test.xlsm"
    test_filepath = os.path.join(temp_dir, test_filename)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = excel_schema.SHEET_DAILY_REPORT
    ws.cell(row=1, column=1, value="管理番号")
    ws.cell(row=2, column=1, value=1)
    ws.cell(row=2, column=excel_schema.DailyReportColumns.COMMENT_REPLY, value="既存の返信コメントX")

    wb.save(test_filepath)
    wb.close()

    orig_excel_dir = config.EXCEL_DIR
    config.EXCEL_DIR = temp_dir
    bg_tasks = BackgroundTasks()

    try:
        # 古いデータ（"返信コメントY"）をもとに返信しようとすると衝突
        stale_reply = models.ReplyInput(
            コメント返信欄="新しい返信",
            original_values={"コメント返信欄": "返信コメントY"}
        )

        with pytest.raises(HTTPException) as exc_info:
            routes_reports.update_report_reply(1, stale_reply, bg_tasks, filename=test_filename)

        assert exc_info.value.status_code == 409
        assert "他の方が返信コメントを編集しました" in exc_info.value.detail

        # 一致している場合は正常更新
        valid_reply = models.ReplyInput(
            コメント返信欄="正常な返信",
            original_values={"コメント返信欄": "既存の返信コメントX"}
        )
        res = routes_reports.update_report_reply(1, valid_reply, bg_tasks, filename=test_filename)
        assert res["success"] is True

    finally:
        config.EXCEL_DIR = orig_excel_dir
        try:
            os.remove(test_filepath)
            backup_dir = os.path.join(temp_dir, 'backup')
            if os.path.exists(backup_dir):
                for f in os.listdir(backup_dir):
                    os.remove(os.path.join(backup_dir, f))
                os.rmdir(backup_dir)
            os.rmdir(temp_dir)
        except Exception:
            pass
