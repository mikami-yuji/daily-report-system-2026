import os
import sys
import tempfile
import sqlite3
import openpyxl
import pandas as pd
import pytest

sys.path.insert(0, os.path.abspath(os.path.join(os.path.dirname(__file__), '..')))

import config
import cache

def test_sqlite_cache_creation_and_hit():
    """初回読み込みでSQLiteキャッシュが作成され、2回目以降SQLiteから復元されることを検証"""
    temp_dir = tempfile.mkdtemp()
    test_db_path = os.path.join(temp_dir, "test_shadow.db")
    test_excel_filename = "cache_test.xlsm"
    test_excel_path = os.path.join(temp_dir, test_excel_filename)

    # 1. テスト用Excelファイルの作成
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "営業日報"
    ws.cell(row=1, column=1, value="管理番号")
    ws.cell(row=1, column=2, value="商談内容")
    ws.cell(row=2, column=1, value=1)
    ws.cell(row=2, column=2, value="初期商談内容")
    wb.save(test_excel_path)
    wb.close()

    orig_excel_dir = config.EXCEL_DIR
    orig_sqlite_db = config.SQLITE_CACHE_DB
    config.EXCEL_DIR = temp_dir
    config.SQLITE_CACHE_DB = test_db_path

    try:
        # 初回読み込み（Excelから読み込んでSQLiteに保存）
        df1 = cache.get_cached_dataframe(test_excel_filename, "営業日報")
        assert len(df1) == 1
        assert df1.iloc[0]["商談内容"] == "初期商談内容"

        # SQLiteファイルが作成され、メタデータとテーブルが存在することを確認
        assert os.path.exists(test_db_path)
        with sqlite3.connect(test_db_path) as conn:
            cursor = conn.execute("SELECT filename, sheet_name, mtime FROM _cache_meta")
            meta_rows = cursor.fetchall()
            assert len(meta_rows) == 1
            assert meta_rows[0][0] == test_excel_filename
            assert meta_rows[0][1] == "営業日報"

        # インメモリキャッシュをあえて空にして、2回目の取得がSQLiteから行われることを確認
        cache.CACHE.clear()
        df2 = cache.get_cached_dataframe(test_excel_filename, "営業日報")
        assert len(df2) == 1
        assert df2.iloc[0]["商談内容"] == "初期商談内容"

    finally:
        config.EXCEL_DIR = orig_excel_dir
        config.SQLITE_CACHE_DB = orig_sqlite_db
        try:
            if os.path.exists(test_excel_path):
                os.remove(test_excel_path)
            if os.path.exists(test_db_path):
                os.remove(test_db_path)
            # WALファイル等も削除
            for f in os.listdir(temp_dir):
                os.remove(os.path.join(temp_dir, f))
            os.rmdir(temp_dir)
        except Exception:
            pass


def test_sqlite_cache_auto_sync_on_excel_update():
    """Excelファイルが更新された際、SQLiteキャッシュが自動的に再同期されることを検証"""
    temp_dir = tempfile.mkdtemp()
    test_db_path = os.path.join(temp_dir, "test_sync.db")
    test_excel_filename = "sync_test.xlsm"
    test_excel_path = os.path.join(temp_dir, test_excel_filename)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "営業日報"
    ws.cell(row=1, column=1, value="管理番号")
    ws.cell(row=1, column=2, value="商談内容")
    ws.cell(row=2, column=1, value=1)
    ws.cell(row=2, column=2, value="初期内容")
    wb.save(test_excel_path)
    wb.close()

    orig_excel_dir = config.EXCEL_DIR
    orig_sqlite_db = config.SQLITE_CACHE_DB
    config.EXCEL_DIR = temp_dir
    config.SQLITE_CACHE_DB = test_db_path

    try:
        # 初回読み込み
        df1 = cache.get_cached_dataframe(test_excel_filename, "営業日報")
        assert df1.iloc[0]["商談内容"] == "初期内容"

        # Excelを更新（行を追加）し、mtimeを確実に進める
        import time
        time.sleep(0.05)
        wb = openpyxl.load_workbook(test_excel_path)
        ws = wb["営業日報"]
        ws.cell(row=3, column=1, value=2)
        ws.cell(row=3, column=2, value="追加された行")
        wb.save(test_excel_path)
        wb.close()

        # インメモリキャッシュをクリアして再取得
        cache.CACHE.clear()
        df2 = cache.get_cached_dataframe(test_excel_filename, "営業日報")
        assert len(df2) == 2
        assert df2.iloc[1]["商談内容"] == "追加された行"

    finally:
        config.EXCEL_DIR = orig_excel_dir
        config.SQLITE_CACHE_DB = orig_sqlite_db
        try:
            for f in os.listdir(temp_dir):
                os.remove(os.path.join(temp_dir, f))
            os.rmdir(temp_dir)
        except Exception:
            pass


def test_invalidate_cache_clears_memory_and_sqlite():
    """invalidate_cacheがインメモリとSQLiteの両方を破棄することを検証"""
    temp_dir = tempfile.mkdtemp()
    test_db_path = os.path.join(temp_dir, "test_inv.db")
    test_excel_filename = "inv_test.xlsm"
    test_excel_path = os.path.join(temp_dir, test_excel_filename)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "営業日報"
    ws.cell(row=1, column=1, value="管理番号")
    ws.cell(row=2, column=1, value=1)
    wb.save(test_excel_path)
    wb.close()

    orig_excel_dir = config.EXCEL_DIR
    orig_sqlite_db = config.SQLITE_CACHE_DB
    config.EXCEL_DIR = temp_dir
    config.SQLITE_CACHE_DB = test_db_path

    try:
        cache.get_cached_dataframe(test_excel_filename, "営業日報")
        assert (test_excel_filename, "営業日報") in cache.CACHE

        with sqlite3.connect(test_db_path) as conn:
            cursor = conn.execute("SELECT COUNT(*) FROM _cache_meta")
            assert cursor.fetchone()[0] == 1

        # 無効化実行
        cache.invalidate_cache(test_excel_filename, "営業日報")

        # インメモリが消えていること
        assert (test_excel_filename, "営業日報") not in cache.CACHE

        # SQLite側も消えていること
        with sqlite3.connect(test_db_path) as conn:
            cursor = conn.execute("SELECT COUNT(*) FROM _cache_meta")
            assert cursor.fetchone()[0] == 0

    finally:
        config.EXCEL_DIR = orig_excel_dir
        config.SQLITE_CACHE_DB = orig_sqlite_db
        try:
            for f in os.listdir(temp_dir):
                os.remove(os.path.join(temp_dir, f))
            os.rmdir(temp_dir)
        except Exception:
            pass


def test_no_pickle_usage_in_codebase():
    """backendディレクトリ内のプロダクションコード（tests以外）にpickleのimportが一切存在しないことを検証"""
    backend_dir = os.path.abspath(os.path.join(os.path.dirname(__file__), '..'))
    forbidden = ["import " + "pickle", "from " + "pickle import"]
    
    for root, dirs, files in os.walk(backend_dir):
        # 仮想環境やキャッシュディレクトリ、テストフォルダは除外
        if any(ignored in root for ignored in ['.venv', '__pycache__', '.pytest_cache', 'tests']):
            continue
        for file in files:
            if file.endswith('.py'):
                filepath = os.path.join(root, file)
                with open(filepath, 'r', encoding='utf-8', errors='ignore') as f:
                    content = f.read()
                    for f_pattern in forbidden:
                        assert f_pattern not in content, f"Forbidden '{f_pattern}' found in {filepath}"
